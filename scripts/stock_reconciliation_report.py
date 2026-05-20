"""Stock reconciliation report — for trip-prep spot-check.

Per product, shows:
  1. opening_count from Inventory Management - Opening_Stock.csv (physical
     count, ~Feb-Mar 2026 dates, considered ~90% accurate)
  2. unit_csv vs unit_db (flag mismatch — unit conversion bugs are the
     #1 cause of OVER/SHORT noise)
  3. current_stock from stock_levels (right now)
  4. delta_vs_opening = current - opening (raw difference)
  5. Informational: bsn_in / bsn_out / manual_adjust totals since 2024-01-03,
     but NOT used in status (because BSN sync back-dates created_at to BSN
     doc dates → can't reliably filter 'transactions since count date')

Status logic (simplified):
  UNIT_MISMATCH   unit_csv != unit_db (delta is meaningless — fix unit first)
  MISSING_PRODUCT SKU in CSV not in active products
  OK              |delta_vs_opening| < 50 AND <25% of max
  CHANGED         everything else (Put spot-checks: was the change expected?)

Output: data/exports/stock_reconciliation_2026-05-08.xlsx with color coding.
Sorted by |delta| DESC.
"""
from __future__ import annotations

import csv
import os
import sqlite3
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "inventory_app" / "instance" / "inventory.db"
_INPUT_DIR = Path(os.environ.get("SENDY_INPUT_DIR", os.path.expanduser("~/Downloads")))
OPENING_CSV = _INPUT_DIR / "Inventory Management - Opening_Stock.csv"
OUT_XLSX = ROOT / "data" / "exports" / "stock_reconciliation_2026-05-08.xlsx"

DELTA_THRESHOLD_ABS = 50      # absolute units
DELTA_THRESHOLD_PCT = 0.25    # 25% of max(opening, current)


def parse_int(s):
    if not s: return 0
    s = str(s).replace(",", "").strip()
    try: return int(float(s))
    except: return 0


def main():
    if not OPENING_CSV.exists():
        raise SystemExit(f"opening CSV not found: {OPENING_CSV}")
    if not DB_PATH.exists():
        raise SystemExit(f"DB not found: {DB_PATH}")

    # Load opening stock
    opening = {}  # sku → {pieces, name, count_date, unit, note}
    with open(OPENING_CSV, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            sku_raw = (r.get("SKU (Order)") or "").strip()
            try:
                sku = int(sku_raw)
            except: continue
            opening[sku] = {
                "pieces": parse_int(r.get("Pieces")),
                "name_in_csv": r.get("รายการ", "").strip(),
                "count_date": r.get("Date", "").strip(),
                "unit_in_csv": r.get("หน่วย", "").strip(),
                "remark": (r.get("Remark") or "").strip(),
            }

    print(f"Opening counts loaded: {len(opening)} SKUs")

    # DB query
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row

    # Product master
    prod_rows = conn.execute("""
        SELECT p.id, p.sku, p.product_name, p.unit_type,
               p.cost_price, p.base_sell_price,
               COALESCE(s.quantity, 0) AS current_qty
          FROM products p
          LEFT JOIN stock_levels s ON s.product_id = p.id
         WHERE p.is_active = 1
    """).fetchall()
    by_sku = {r["sku"]: dict(r) for r in prod_rows}

    # Load unit_conversions per product. bsn_unit is often a truncated form
    # (กล for กล่อง, ซง for ซอง, etc.) due to BSN cp874 truncation. We normalize
    # CSV's full unit names (กล่อง/ซอง/แพ็ค) to match.
    UNIT_ALIAS = {
        "กล่อง": ("กล", "กล่อง"),
        "ซอง":   ("ซง", "ซอง"),
        "แพ็ค":  ("แพ", "แพ็ค", "แพค"),
        "ลัง":   ("ลัง",),
        "โหล":   ("โหล", "ดซ"),
        "ตัว":   ("ตัว",),
        "แผง":   ("แผง",),
        "ดอก":   ("ดอก",),
        "ใบ":    ("ใบ",),
        "ม้วน":  ("ม้วน",),
        "อัน":   ("อัน",),
        "ผืน":   ("ผืน",),
        "ชิ้น":  ("ชิ้น",),
    }
    unit_conv = {}  # (product_id, bsn_unit) → ratio
    for r in conn.execute("SELECT product_id, bsn_unit, ratio FROM unit_conversions"):
        unit_conv[(r[0], r[1])] = r[2]

    def find_ratio(product_id, csv_unit):
        """Convert CSV unit (full form) to DB unit_type via unit_conversions.
        Returns ratio (DB units per CSV unit) or None."""
        if not csv_unit:
            return None
        # Try exact match first
        if (product_id, csv_unit) in unit_conv:
            return unit_conv[(product_id, csv_unit)]
        # Try aliases (CSV full form → BSN truncated form)
        for full, alts in UNIT_ALIAS.items():
            if csv_unit == full:
                for a in alts:
                    if (product_id, a) in unit_conv:
                        return unit_conv[(product_id, a)]
        return None

    # Per-product txn sums — EXCLUDE all ADJUSTs on 2024-01-03 since those
    # are the original opening-balance setup (Put's previous Claude session
    # set them up via 'ยอดต้นปี back-solved') AND today's compensating fix
    # (note='opening adjust auto-corrected...'). Both represent opening; we
    # account for opening separately via the CSV count, so don't double-count.
    txn_sums = {}
    for r in conn.execute("""
        SELECT product_id,
               SUM(CASE WHEN txn_type='IN'  THEN quantity_change ELSE 0 END) AS in_qty,
               SUM(CASE WHEN txn_type='OUT' THEN -quantity_change ELSE 0 END) AS out_qty,
               SUM(CASE WHEN txn_type='ADJUST'
                         AND date(created_at) != '2024-01-03'
                        THEN quantity_change ELSE 0 END) AS adjust_qty
          FROM transactions
         WHERE created_at >= '2024-01-03'
         GROUP BY product_id
    """).fetchall():
        txn_sums[r["product_id"]] = {
            "in": r["in_qty"] or 0,
            "out": r["out_qty"] or 0,
            "adjust": r["adjust_qty"] or 0,
        }

    conn.close()

    # Compose report rows
    rows = []
    for sku, opn in opening.items():
        p = by_sku.get(sku)
        if not p:
            rows.append({
                "sku": sku,
                "product_name_db": "(NOT IN ACTIVE PRODUCTS)",
                "name_in_csv": opn["name_in_csv"],
                "unit_db": "",
                "unit_csv": opn["unit_in_csv"],
                "unit_match": "?",
                "ratio_used": "",
                "opening_count": opn["pieces"],
                "opening_in_db_unit": "",
                "current_stock": "",
                "delta_vs_opening": "",
                "abs_delta": opn["pieces"],
                "bsn_in_since_2024": "",
                "bsn_out_since_2024": "",
                "manual_adjust_since_2024": "",
                "status": "MISSING_PRODUCT",
                "count_date": opn["count_date"],
                "csv_remark": opn["remark"],
                "spot_check_actual": "",
                "notes": "",
            })
            continue
        sums = txn_sums.get(p["id"], {"in": 0, "out": 0, "adjust": 0})

        # Convert opening_count to DB unit using unit_conversions when units differ
        unit_mismatch = (opn["unit_in_csv"]
                         and p["unit_type"]
                         and opn["unit_in_csv"] != p["unit_type"])
        opening_in_db_unit = opn["pieces"]
        ratio_used = None
        if unit_mismatch:
            ratio = find_ratio(p["id"], opn["unit_in_csv"])
            if ratio:
                opening_in_db_unit = opn["pieces"] * ratio
                ratio_used = ratio
            else:
                opening_in_db_unit = None  # can't compare without conversion

        # Recompute expected + delta using converted opening
        if opening_in_db_unit is not None:
            expected = opening_in_db_unit + sums["in"] - sums["out"] + sums["adjust"]
            delta_expected = p["current_qty"] - expected
            delta_opening = p["current_qty"] - opening_in_db_unit
        else:
            expected = None
            delta_expected = None
            delta_opening = None

        # Status: prefer delta_opening with converted opening when we have ratio
        if delta_opening is None:
            # No conversion ratio for unit mismatch
            status = "NEEDS_RATIO"
        else:
            ref = max(abs(opening_in_db_unit or 0), abs(p["current_qty"]), 1)
            pct = abs(delta_opening) / ref
            if abs(delta_opening) < DELTA_THRESHOLD_ABS and pct < DELTA_THRESHOLD_PCT:
                status = "OK"
            else:
                status = "CHANGED"

        rows.append({
            "sku": sku,
            "product_name_db": p["product_name"],
            "name_in_csv": opn["name_in_csv"],
            "unit_db": p["unit_type"],
            "unit_csv": opn["unit_in_csv"],
            "unit_match": "✓" if (not unit_mismatch and opn["unit_in_csv"] and p["unit_type"]) else ("✗" if unit_mismatch else "?"),
            "ratio_used": ratio_used or "",
            "opening_count": opn["pieces"],
            "opening_in_db_unit": opening_in_db_unit if opening_in_db_unit is not None else "",
            "current_stock": p["current_qty"],
            "delta_vs_opening": delta_opening if delta_opening is not None else "",
            "bsn_in_since_2024": sums["in"],
            "bsn_out_since_2024": sums["out"],
            "manual_adjust_since_2024": sums["adjust"],
            "abs_delta": abs(delta_opening) if delta_opening is not None else 0,
            "status": status,
            "count_date": opn["count_date"],
            "csv_remark": opn["remark"],
            "spot_check_actual": "",
            "notes": "",
        })

    # Filter to "interesting" rows
    flagged = [r for r in rows if r["status"] in ("CHANGED", "MISSING_PRODUCT", "NEEDS_RATIO")]
    flagged.sort(key=lambda r: -r["abs_delta"])

    # Stats
    n_total = len(rows)
    by_st = {}
    for r in rows:
        by_st[r["status"]] = by_st.get(r["status"], 0) + 1
    print(f"\nTotals: {n_total} SKUs in opening file")
    for st in ("OK", "CHANGED", "NEEDS_RATIO", "MISSING_PRODUCT"):
        n = by_st.get(st, 0)
        if n:
            print(f"  {st:<18}  {n}  ({n*100//n_total}%)")
    print(f"  → Flagged:        {len(flagged)} rows for spot-check")

    # Write Excel
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    headers = [
        "sku", "product_name_db", "name_in_csv",
        "unit_db", "unit_csv", "unit_match", "ratio_used",
        "opening_count", "opening_in_db_unit", "current_stock",
        "delta_vs_opening", "abs_delta",
        "bsn_in_since_2024", "bsn_out_since_2024", "manual_adjust_since_2024",
        "status", "count_date", "csv_remark",
        "spot_check_actual", "notes",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Color rules
    fill_short = PatternFill("solid", fgColor="FFD6D6")    # red-ish
    fill_over  = PatternFill("solid", fgColor="FFEFC2")    # yellow-ish
    fill_missing = PatternFill("solid", fgColor="C8C8E0")  # purple-ish
    fill_unit = PatternFill("solid", fgColor="DCEAF8")     # blue-ish

    for r in flagged:
        row = [r[h] for h in headers]
        ws.append(row)
        last = ws.max_row
        # Color: red if delta < 0 (lost stock), yellow if delta > 0 (gained),
        # blue if needs ratio, purple if missing product
        if r["status"] == "NEEDS_RATIO":
            fill = fill_unit
        elif r["status"] == "MISSING_PRODUCT":
            fill = fill_missing
        elif isinstance(r.get("delta_vs_opening"), (int, float)) and r["delta_vs_opening"] < 0:
            fill = fill_short
        else:
            fill = fill_over
        for c in range(1, len(headers)+1):
            ws.cell(row=last, column=c).fill = fill

    # Column widths
    widths = {
        "A": 8, "B": 42, "C": 42,         # sku, product_name, csv_name
        "D": 8, "E": 8, "F": 8, "G": 8,   # unit_db, unit_csv, unit_match, ratio
        "H": 10, "I": 14, "J": 11,        # opening, opening_in_db_unit, current
        "K": 12, "L": 9,                  # delta, abs_delta
        "M": 10, "N": 10, "O": 12,        # bsn_in/out, manual
        "P": 16, "Q": 12, "R": 22,        # status, date, remark
        "S": 15, "T": 25,                 # spot_check, notes
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "C2"

    wb.save(OUT_XLSX)
    print(f"\nOutput: {OUT_XLSX.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
