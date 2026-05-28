"""Export product data for Google Sheet review — 2 sheets:

Sheet 1 'stock_and_mapping' — one row per (product, bsn_code) pair:
  sku, sku_code, product_name, old_product_name, base_unit, stock,
  bsn_code, bsn_name, bsn_unit, ratio_to_base

Sheet 2 'sku_code_parts' — one row per product, structured columns
  composing sku_code:
  sku, sku_code, category_short, category_th, brand_short, brand,
  series, model, size, color_code, color_th, packaging, condition, pack_variant

Output: data/exports/product_review_2026-05-08.xlsx
"""
from __future__ import annotations

import json
import sqlite3
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "inventory_app" / "instance" / "inventory.db"
OUT = ROOT / "data" / "exports" / "product_review_2026-05-08.xlsx"


def get_old_product_names(conn):
    """For each product_id, find the OLDEST recorded old_value of product_name
    in audit_log. That's the closest we have to 'original' pre-rename name."""
    rows = conn.execute("""
        SELECT row_id,
               json_extract(changed_fields, '$.product_name') AS pn_pair
          FROM audit_log
         WHERE table_name = 'products'
           AND action = 'UPDATE'
           AND changed_fields LIKE '%product_name%'
         ORDER BY id ASC
    """).fetchall()
    earliest = {}  # product_id → first old_value seen
    for r in rows:
        pid = r[0]
        try:
            pair = json.loads(r[1] or "[]")
        except Exception:
            continue
        if not isinstance(pair, list) or len(pair) < 1:
            continue
        if pid not in earliest:
            earliest[pid] = pair[0]  # old value of the first UPDATE
    return earliest


def main():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row

    old_names = get_old_product_names(conn)
    print(f"Old product names captured from audit_log: {len(old_names)}")

    # Sheet 1: stock + BSN mapping (one row per product-bsn pair)
    sheet1 = conn.execute("""
        SELECT p.id AS product_id, p.sku, p.sku_code, p.product_name,
               p.unit_type AS base_unit,
               COALESCE(s.quantity, 0) AS stock,
               m.bsn_code, m.bsn_name,
               uc.bsn_unit, uc.ratio
          FROM products p
          LEFT JOIN stock_levels s ON s.product_id = p.id
          LEFT JOIN product_code_mapping m ON m.product_id = p.id AND COALESCE(m.is_ignored, 0) = 0
          LEFT JOIN unit_conversions uc ON uc.product_id = p.id AND uc.bsn_unit IS NOT NULL
         WHERE p.is_active = 1
         ORDER BY p.sku, m.bsn_code, uc.bsn_unit
    """).fetchall()

    # Sheet 2: sku_code breakdown
    sheet2 = conn.execute("""
        SELECT p.sku, p.sku_code, p.product_name,
               c.short_code AS category_short, c.name_th AS category_th,
               p.sub_category,
               b.short_code AS brand_short, b.name AS brand,
               p.series, p.model, p.size,
               p.color_code, cf.name_th AS color_th,
               p.packaging_th AS packaging, p.condition, p.pack_variant
          FROM products p
          LEFT JOIN categories c ON c.id = p.category_id
          LEFT JOIN brands b ON b.id = p.brand_id
          LEFT JOIN color_finish_codes cf ON cf.code = p.color_code
         WHERE p.is_active = 1
         ORDER BY p.sku
    """).fetchall()

    conn.close()

    # Build workbook
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "stock_and_mapping"

    # Sheet 1 headers
    h1 = ["sku", "sku_code", "product_name", "old_product_name",
          "base_unit", "stock",
          "bsn_code", "bsn_name", "bsn_unit", "ratio_to_base"]
    ws1.append(h1)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(horizontal="center")

    seen_pids = set()  # for tracking products already written (any row)
    for r in sheet1:
        old_name = old_names.get(r["product_id"], "")
        ws1.append([
            r["sku"], r["sku_code"], r["product_name"], old_name,
            r["base_unit"], r["stock"],
            r["bsn_code"] or "", r["bsn_name"] or "",
            r["bsn_unit"] or "", r["ratio"] or "",
        ])
        seen_pids.add(r["product_id"])

    # Column widths sheet 1
    widths1 = {"A": 8, "B": 28, "C": 50, "D": 50, "E": 10, "F": 10,
               "G": 14, "H": 32, "I": 10, "J": 10}
    for col, w in widths1.items():
        ws1.column_dimensions[col].width = w
    ws1.freeze_panes = "B2"

    # Sheet 2
    ws2 = wb.create_sheet("sku_code_parts")
    h2 = ["sku", "sku_code", "product_name",
          "category_short", "category_th", "sub_category",
          "brand_short", "brand", "series", "model", "size",
          "color_code", "color_th", "packaging", "condition", "pack_variant"]
    ws2.append(h2)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(horizontal="center")

    for r in sheet2:
        ws2.append([
            r["sku"], r["sku_code"], r["product_name"],
            r["category_short"] or "", r["category_th"] or "", r["sub_category"] or "",
            r["brand_short"] or "", r["brand"] or "",
            r["series"] or "", r["model"] or "", r["size"] or "",
            r["color_code"] or "", r["color_th"] or "",
            r["packaging"] or "", r["condition"] or "", r["pack_variant"] or "",
        ])

    widths2 = {"A": 8, "B": 28, "C": 50,
               "D": 8, "E": 22, "F": 18,
               "G": 8, "H": 14, "I": 14, "J": 12, "K": 14,
               "L": 10, "M": 14, "N": 12, "O": 10, "P": 10}
    for col, w in widths2.items():
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "B2"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)

    print(f"\nSheet 1 'stock_and_mapping': {ws1.max_row - 1} rows ({len(seen_pids)} unique products)")
    print(f"Sheet 2 'sku_code_parts':    {ws2.max_row - 1} rows")
    print(f"\nOutput: {OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
