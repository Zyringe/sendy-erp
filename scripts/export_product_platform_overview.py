"""
Export an Excel summarising every internal product that is mapped to at least
one Shopee or Lazada listing.

One row per internal product. Aggregated across variations:
  - Sendy:  product_id, name, internal stock
  - Shopee: variation count, names (joined), total stock, active flag
  - Lazada: variation count, names (joined), total stock, active flag
  - Mapping coverage flag (both / shopee_only / lazada_only / unmapped_listed)

Optional --shopee-inactive PATH points to a Shopee mass_update_basic_info xlsx
that the user filtered to "not visible" — every parent product_id in that file
is treated as inactive on Shopee.

Run: python scripts/export_product_platform_overview.py
Output: data/exports/product_platform_overview_YYYYMMDD.xlsx
"""
import argparse
import datetime
import json
import sqlite3
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / 'inventory_app' / 'instance' / 'inventory.db'
EXPORTS = ROOT / 'data' / 'exports'
sys.path.insert(0, str(ROOT / 'inventory_app'))


def lazada_active(raw_json):
    if not raw_json:
        return None
    try:
        return json.loads(raw_json).get('status')
    except Exception:
        return None


def load_shopee_inactive(path):
    if not path:
        return set()
    from parse_platform import parse_shopee
    return {r['product_id_str'] for r in parse_shopee(path) if r.get('product_id_str')}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--shopee-inactive', default=None,
                    help='Shopee mass_update_basic_info xlsx of products NOT visible on Shopee')
    args = ap.parse_args()

    shopee_inactive_ids = load_shopee_inactive(args.shopee_inactive)
    if shopee_inactive_ids:
        print(f'Shopee inactive parent IDs: {len(shopee_inactive_ids)}')

    EXPORTS.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    rows = conn.execute('''
        SELECT p.id, p.product_name, COALESCE(sl.quantity, 0) AS internal_stock
        FROM products p
        LEFT JOIN stock_levels sl ON sl.product_id = p.id
        WHERE p.is_active = 1
          AND EXISTS (SELECT 1 FROM platform_skus ps WHERE ps.internal_product_id = p.id)
        ORDER BY p.id
    ''').fetchall()

    # Unmapped Shopee parents (one row per distinct product_id_str)
    unmapped_shopee_parents = conn.execute('''
        SELECT product_id_str
        FROM platform_skus
        WHERE platform='shopee' AND internal_product_id IS NULL
          AND product_id_str IS NOT NULL
        GROUP BY product_id_str
        ORDER BY MIN(product_name)
    ''').fetchall()

    unmapped_lazada_parents = conn.execute('''
        SELECT product_id_str
        FROM platform_skus
        WHERE platform='lazada' AND internal_product_id IS NULL
          AND product_id_str IS NOT NULL
        GROUP BY product_id_str
        ORDER BY MIN(product_name)
    ''').fetchall()

    output = []
    for p in rows:
        shopee = conn.execute('''
            SELECT product_id_str, product_name, variation_name, stock, raw_json, qty_per_sale
            FROM platform_skus
            WHERE platform='shopee' AND internal_product_id = ?
            ORDER BY product_name, variation_name
        ''', (p['id'],)).fetchall()
        lazada = conn.execute('''
            SELECT product_name, variation_name, stock, raw_json, qty_per_sale
            FROM platform_skus
            WHERE platform='lazada' AND internal_product_id = ?
            ORDER BY product_name, variation_name
        ''', (p['id'],)).fetchall()

        def name_with_variation(r):
            base = r['product_name'] + (f" — {r['variation_name']}" if r['variation_name'] else '')
            qps = r['qty_per_sale']
            if qps and qps != 1:
                base += f"  [×{qps:g}]"
            return base

        shopee_names = sorted({s['product_name'] for s in shopee})
        lazada_names = sorted({l['product_name'] for l in lazada})

        shopee_stock = sum((s['stock'] or 0) for s in shopee)
        lazada_stock = sum((l['stock'] or 0) for l in lazada)

        # Lazada has explicit status flag in raw_json
        lz_active = any(lazada_active(l['raw_json']) == 'active' for l in lazada)
        lz_status = ('active' if lz_active
                     else ('inactive' if lazada else '—'))
        # Shopee status: file of "not visible" parents is authoritative if provided;
        # otherwise fall back to stock > 0 proxy.
        if shopee:
            if shopee_inactive_ids:
                sp_active = any(s['product_id_str'] not in shopee_inactive_ids for s in shopee)
                sp_status = 'active' if sp_active else 'inactive'
            else:
                sp_status = 'listed' if any((s['stock'] or 0) > 0 for s in shopee) else 'zero stock'
        else:
            sp_status = '—'

        if shopee and lazada:
            cov = 'both'
        elif shopee:
            cov = 'shopee only'
        elif lazada:
            cov = 'lazada only'
        else:
            cov = '—'

        output.append({
            'product_id': p['id'],
            'sendy_name': p['product_name'],
            'internal_stock': p['internal_stock'],
            'mapping': cov,
            'shopee_count': len(shopee),
            'shopee_names': '\n'.join(name_with_variation(s) for s in shopee),
            'shopee_stock': shopee_stock,
            'shopee_status': sp_status,
            'lazada_count': len(lazada),
            'lazada_names': '\n'.join(name_with_variation(l) for l in lazada),
            'lazada_stock': lazada_stock,
            'lazada_status': lz_status,
        })

    # Append unmapped Shopee parents (one row per parent product_id_str)
    for u in unmapped_shopee_parents:
        variations = conn.execute('''
            SELECT product_id_str, product_name, variation_name, stock, raw_json, qty_per_sale
            FROM platform_skus
            WHERE platform='shopee' AND product_id_str = ?
            ORDER BY variation_name
        ''', (u['product_id_str'],)).fetchall()
        names = sorted({v['product_name'] for v in variations})
        stock = sum((v['stock'] or 0) for v in variations)
        if shopee_inactive_ids:
            sp_active = u['product_id_str'] not in shopee_inactive_ids
            sp_status = 'active' if sp_active else 'inactive'
        else:
            sp_status = 'listed' if stock > 0 else 'zero stock'
        output.append({
            'product_id': '', 'sendy_name': '(unmapped)', 'internal_stock': '',
            'mapping': 'shopee unmapped',
            'shopee_count': len(variations),
            'shopee_names': '\n'.join(
                v['product_name'] + (f" — {v['variation_name']}" if v['variation_name'] else '')
                for v in variations),
            'shopee_stock': stock,
            'shopee_status': sp_status,
            'lazada_count': 0, 'lazada_names': '', 'lazada_stock': '', 'lazada_status': '—',
        })

    for u in unmapped_lazada_parents:
        variations = conn.execute('''
            SELECT product_id_str, product_name, variation_name, stock, raw_json, qty_per_sale
            FROM platform_skus
            WHERE platform='lazada' AND product_id_str = ?
            ORDER BY variation_name
        ''', (u['product_id_str'],)).fetchall()
        stock = sum((v['stock'] or 0) for v in variations)
        lz_active = any(lazada_active(v['raw_json']) == 'active' for v in variations)
        lz_status = 'active' if lz_active else 'inactive'
        output.append({
            'product_id': '', 'sendy_name': '(unmapped)', 'internal_stock': '',
            'mapping': 'lazada unmapped',
            'shopee_count': 0, 'shopee_names': '', 'shopee_stock': '', 'shopee_status': '—',
            'lazada_count': len(variations),
            'lazada_names': '\n'.join(
                v['product_name'] + (f" — {v['variation_name']}" if v['variation_name'] else '')
                for v in variations),
            'lazada_stock': stock,
            'lazada_status': lz_status,
        })

    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Product × Platform'

    cols = [
        ('Product ID', 'product_id', 10),
        ('Sendy name', 'sendy_name', 38),
        ('Stock (Sendy)', 'internal_stock', 12),
        ('Mapping', 'mapping', 12),
        ('Shopee variations', 'shopee_count', 10),
        ('Shopee name(s)', 'shopee_names', 42),
        ('Stock (Shopee)', 'shopee_stock', 12),
        ('Shopee status', 'shopee_status', 12),
        ('Lazada variations', 'lazada_count', 10),
        ('Lazada name(s)', 'lazada_names', 42),
        ('Stock (Lazada)', 'lazada_stock', 12),
        ('Lazada status', 'lazada_status', 12),
    ]

    hdr_fill = PatternFill('solid', start_color='1f3a5f')
    hdr_font = Font(bold=True, color='FFFFFF')
    for ci, (label, _, w) in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=label)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30

    wrap = Alignment(wrap_text=True, vertical='top')
    for ri, r in enumerate(output, 2):
        for ci, (_, key, _) in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci, value=r[key])
            cell.alignment = wrap

    ws.freeze_panes = 'B2'
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: To Map (active + in-stock unmapped variations) ──────────────
    conn2 = sqlite3.connect(DB_PATH)
    conn2.row_factory = sqlite3.Row
    to_map = []
    sp_rows = conn2.execute('''
        SELECT id, product_id_str, product_name, variation_id, variation_name,
               seller_sku, stock, qty_per_sale
        FROM platform_skus
        WHERE platform='shopee' AND internal_product_id IS NULL
          AND COALESCE(stock,0) > 0
        ORDER BY product_name, variation_name
    ''').fetchall()
    for r in sp_rows:
        if shopee_inactive_ids and r['product_id_str'] in shopee_inactive_ids:
            continue
        to_map.append(('shopee', r))
    lz_rows = conn2.execute('''
        SELECT id, product_id_str, product_name, variation_id, variation_name,
               seller_sku, stock, qty_per_sale, raw_json
        FROM platform_skus
        WHERE platform='lazada' AND internal_product_id IS NULL
          AND COALESCE(stock,0) > 0
        ORDER BY product_name, variation_name
    ''').fetchall()
    for r in lz_rows:
        if lazada_active(r['raw_json']) != 'active':
            continue
        to_map.append(('lazada', r))
    conn2.close()

    # Fuzzy-suggest internal_sku for each to-map row using existing models helper.
    # (Cheap to import here — only loaded when --suggest is used.)
    suggestions = {}
    if to_map:
        try:
            sys.path.insert(0, str(ROOT / 'inventory_app'))
            from rapidfuzz import fuzz
            from rapidfuzz.process import cdist
            import models as mdl
            with sqlite3.connect(DB_PATH) as c3:
                c3.row_factory = sqlite3.Row
                products_all = c3.execute(
                    'SELECT id, product_name FROM products WHERE is_active=1'
                ).fetchall()
            corpus = [mdl._clean_for_match(p['product_name']) for p in products_all]
            queries = [
                mdl._clean_for_match(
                    f"{r['product_name']} {r['variation_name'] or ''} {r['seller_sku'] or ''}"
                )
                for _, r in to_map
            ]
            matrix = cdist(queries, corpus, scorer=fuzz.token_set_ratio, workers=-1)
            best_idx = matrix.argmax(axis=1)
            best_score = matrix.max(axis=1)
            for i, (_, r) in enumerate(to_map):
                score = int(best_score[i])
                if score < 35:
                    continue
                p = products_all[best_idx[i]]
                suggestions[r['id']] = (p['id'], p['product_name'], score)
        except Exception as e:
            print(f'  (fuzzy suggest skipped: {e})')

    ws2 = wb.create_sheet('To Map (Active + In Stock)')
    map_cols = [
        ('Platform', 12),
        ('platform_sku_id', 14),
        ('Parent ID', 14),
        ('Variation ID', 16),
        ('Listing name', 38),
        ('Variation', 22),
        ('Seller SKU', 18),
        ('Stock', 8),
        ('Suggested product_id', 16),
        ('Suggested name', 38),
        ('Confidence', 10),
        ('→ product_id (fill in)', 18),
        ('→ qty_per_sale (default 1)', 14),
    ]
    hdr_fill2 = PatternFill('solid', start_color='2e7d3a')
    for ci, (label, w) in enumerate(map_cols, 1):
        c = ws2.cell(row=1, column=ci, value=label)
        c.fill = hdr_fill2
        c.font = hdr_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[1].height = 30

    fill_yellow = PatternFill('solid', start_color='FFF59D')
    for ri, (platform, r) in enumerate(to_map, 2):
        suggested = suggestions.get(r['id'])
        ws2.cell(row=ri, column=1, value=platform).alignment = wrap
        ws2.cell(row=ri, column=2, value=r['id']).alignment = wrap
        ws2.cell(row=ri, column=3, value=r['product_id_str']).alignment = wrap
        ws2.cell(row=ri, column=4, value=r['variation_id']).alignment = wrap
        ws2.cell(row=ri, column=5, value=r['product_name']).alignment = wrap
        ws2.cell(row=ri, column=6, value=r['variation_name']).alignment = wrap
        ws2.cell(row=ri, column=7, value=r['seller_sku']).alignment = wrap
        ws2.cell(row=ri, column=8, value=r['stock']).alignment = wrap
        if suggested:
            ws2.cell(row=ri, column=9, value=suggested[0]).alignment = wrap
            ws2.cell(row=ri, column=10, value=suggested[1]).alignment = wrap
            ws2.cell(row=ri, column=11, value=suggested[2]).alignment = wrap
        # User-fillable columns highlighted yellow
        c12 = ws2.cell(row=ri, column=12, value=suggested[0] if suggested and suggested[2] >= 70 else '')
        c12.fill = fill_yellow
        c12.alignment = wrap
        c13 = ws2.cell(row=ri, column=13, value=1)
        c13.fill = fill_yellow
        c13.alignment = wrap

    ws2.freeze_panes = 'B2'
    ws2.auto_filter.ref = ws2.dimensions

    fname = f'product_platform_overview_{datetime.date.today().strftime("%Y%m%d")}.xlsx'
    path = EXPORTS / fname
    wb.save(path)

    counts = {
        'rows': len(output),
        'both': sum(1 for r in output if r['mapping'] == 'both'),
        'shopee_only': sum(1 for r in output if r['mapping'] == 'shopee only'),
        'lazada_only': sum(1 for r in output if r['mapping'] == 'lazada only'),
        'shopee_unmapped': sum(1 for r in output if r['mapping'] == 'shopee unmapped'),
        'lazada_unmapped': sum(1 for r in output if r['mapping'] == 'lazada unmapped'),
        'to_map_rows': len(to_map),
        'to_map_pre_filled': sum(1 for _, r in to_map if suggestions.get(r['id'], (None,None,0))[2] >= 70),
    }
    print(f'Wrote {path}')
    for k, v in counts.items():
        print(f'  {k:15s} {v}')


if __name__ == '__main__':
    main()
