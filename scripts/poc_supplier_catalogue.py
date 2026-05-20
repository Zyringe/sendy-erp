"""
POC: parse supplier catalogue (ศรีไทยเจริญโลหะกิจ, ไฟล์ 2-69.xlsx)
Goal: validate that we can extract name/unit/price/discount + font-color flag
      and roughly match catalogue items to ERP purchase_transactions.

Run:
  python scripts/poc_supplier_catalogue.py

No DB writes. Output: JSON to stdout + match report.
"""
import json
import re
import sqlite3
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

CATALOGUE_PATH = Path("/Volumes/ZYRINGE/ใบราคาสินค้าเดือน 2-69.xlsx")
DB_PATH = Path(__file__).resolve().parent.parent / "inventory_app" / "instance" / "inventory.db"
SUPPLIER_NAME = "ศรีไทยเจริญโลหะกิจ"
TARGET_SHEETS = ["ก", "ข"]

# Font-color → semantic flag (per instruction sheet of the file)
COLOR_FLAG = {
    "FFFF0000": "changed",   # red
    "FF000000": "same",      # black (or blank)
    "FF0000FF": "new",       # blue
    "FF00B050": "preorder",  # green (Excel default green)
    "FF008000": "preorder",  # alt green
    "FF179517": "preorder",  # alt green seen in this file (2 rows)
}


def color_to_flag(color):
    if color is None:
        return "same"
    rgb = getattr(color, "rgb", None)
    if not rgb or not isinstance(rgb, str):
        return "same"
    return COLOR_FLAG.get(rgb.upper(), f"unknown:{rgb}")


def parse_sheet(ws):
    """Return list of dicts. Header at row 2 (1-indexed), data from row 3."""
    rows = []
    sub_category = None
    # rows are 1-indexed in openpyxl
    for row_idx in range(3, ws.max_row + 1):
        c_name = ws.cell(row=row_idx, column=1)
        c_min  = ws.cell(row=row_idx, column=2)
        c_unit = ws.cell(row=row_idx, column=3)
        c_price = ws.cell(row=row_idx, column=4)
        c_disc = ws.cell(row=row_idx, column=5)
        c_cash = ws.cell(row=row_idx, column=6)

        name = c_name.value
        if name is None:
            continue
        name = str(name).strip()
        if not name:
            continue

        # Sub-category marker like **หมวดย่อย**
        if name.startswith("**") and name.endswith("**"):
            sub_category = name.strip("*").strip()
            continue

        # Skip if no price (likely a header/spacer row)
        if c_price.value is None:
            continue

        rows.append({
            "name_raw": name,
            "sub_category": sub_category,
            "min_order_qty": c_min.value,
            "unit": (str(c_unit.value).strip() if c_unit.value else None),
            "list_price": c_price.value,
            "trade_discount": c_disc.value,
            "cash_discount": c_cash.value,
            "price_change_flag": color_to_flag(c_price.font.color) if c_price.font else "same",
            "name_color_flag": color_to_flag(c_name.font.color) if c_name.font else "same",
            "row": row_idx,
        })
    return rows


def fetch_purchase_history(conn, supplier=SUPPLIER_NAME, sheet_filter=None):
    """Fetch distinct purchased items from this supplier.
    Optionally filter to items whose name starts with one of the given Thai
    consonants (so we only compare against parsed sheets).
    """
    cur = conn.execute(
        """
        SELECT product_name_raw, COUNT(*) AS n_lines, MAX(date_iso) AS last_seen
        FROM purchase_transactions
        WHERE supplier = ?
        GROUP BY product_name_raw
        ORDER BY n_lines DESC
        """,
        (supplier,),
    )
    rows = [dict(zip([c[0] for c in cur.description], r)) for r in cur.fetchall()]
    if sheet_filter:
        prefixes = tuple(sheet_filter)
        rows = [r for r in rows if (r["product_name_raw"] or "").lstrip().startswith(prefixes)]
    return rows


_BRACKET_RE = re.compile(r"[\[\(].*?[\]\)]")  # [รุ่นใหม่], (xxx)


def normalize(s):
    if s is None:
        return ""
    s = str(s).strip()
    # strip annotations like [รุ่นใหม่] or (xxx)
    s = _BRACKET_RE.sub(" ", s)
    # normalize inch (handle both " and ”)
    s = s.replace('"', "นิ้ว").replace("”", "นิ้ว").replace("“", "นิ้ว")
    # normalize fullwidth/space variants
    s = s.replace(" ", " ").replace("　", " ")
    # collapse whitespace
    s = " ".join(s.split())
    # case-fold latin parts (STANLEY vs stanley)
    s = s.lower()
    return s


def match_report(catalogue_rows, purchase_rows):
    """Match direction: for each PURCHASED item, is it in the catalogue?
    User only cares about products they've actually bought.
    """
    catalogue_norm = {}
    for r in catalogue_rows:
        key = normalize(r["name_raw"])
        if key and key not in catalogue_norm:
            catalogue_norm[key] = r

    matched = 0
    samples_match = []
    samples_miss = []
    matched_catalogue_keys = set()
    for p in purchase_rows:
        key = normalize(p["product_name_raw"])
        if key in catalogue_norm:
            matched += 1
            matched_catalogue_keys.add(key)
            if len(samples_match) < 5:
                samples_match.append({
                    "purchase": p["product_name_raw"],
                    "catalogue": catalogue_norm[key]["name_raw"],
                    "n_lines": p["n_lines"],
                    "last_seen": p["last_seen"],
                })
        else:
            if len(samples_miss) < 10:
                samples_miss.append({
                    "purchase": p["product_name_raw"],
                    "n_lines": p["n_lines"],
                    "last_seen": p["last_seen"],
                })
    return {
        "purchased_distinct": len(purchase_rows),
        "catalogue_distinct": len(catalogue_norm),
        "purchased_found_in_catalogue": matched,
        "match_rate": round(matched / len(purchase_rows), 3) if purchase_rows else 0,
        "catalogue_unmatched_count": len(catalogue_norm) - len(matched_catalogue_keys),
        "samples_match": samples_match,
        "samples_miss_purchased_not_in_catalogue": samples_miss,
    }


def main():
    if not CATALOGUE_PATH.exists():
        print(f"ERROR: catalogue not found at {CATALOGUE_PATH}", file=sys.stderr)
        sys.exit(1)
    if not DB_PATH.exists():
        print(f"ERROR: DB not found at {DB_PATH}", file=sys.stderr)
        sys.exit(1)

    print(f"Loading workbook: {CATALOGUE_PATH.name}", file=sys.stderr)
    wb = load_workbook(CATALOGUE_PATH, data_only=True)
    print(f"Sheets: {wb.sheetnames}", file=sys.stderr)

    all_rows = []
    color_counter = Counter()
    unit_counter = Counter()
    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  skip: sheet '{sheet_name}' not found", file=sys.stderr)
            continue
        ws = wb[sheet_name]
        rows = parse_sheet(ws)
        print(f"  sheet '{sheet_name}': {len(rows)} item rows", file=sys.stderr)
        for r in rows:
            color_counter[r["price_change_flag"]] += 1
            unit_counter[r["unit"]] += 1
        all_rows.extend(rows)

    conn = sqlite3.connect(DB_PATH)
    # Restrict purchase history to items whose name starts with a parsed sheet
    # consonant — so we don't penalize the match rate for sheets we haven't read.
    sheet_prefixes = []
    for s in TARGET_SHEETS:
        sheet_prefixes.extend(s.split("-"))
    purchases = fetch_purchase_history(conn, sheet_filter=sheet_prefixes)
    conn.close()

    report = {
        "sheets_parsed": TARGET_SHEETS,
        "rows_extracted": len(all_rows),
        "color_flag_distribution": dict(color_counter),
        "unit_distribution": dict(unit_counter.most_common(15)),
        "match_report": match_report(all_rows, purchases),
        "sample_rows": all_rows[:5],
    }

    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
