#!/usr/bin/env python
"""Export the customer contact-review (all 819 staged rows) to an Excel sheet for Put to audit
in Excel and hand back. Reference columns are read-only; the ✏️ columns are where Put types
corrections. Phone columns are forced to TEXT so Excel can't drop leading zeros / sci-notate.

Pairs with import_customer_review_xlsx.py (preview + apply).
Run: python scripts/export_customer_review_xlsx.py [--db PATH] [--out PATH]
"""
import argparse
import json
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
APP = os.path.join(HERE, '..', 'inventory_app')
import sqlite3  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

DEFAULT_DB = os.path.join(APP, 'instance', 'inventory.db')
DEFAULT_OUT = os.path.join(
    HERE, '..', '..', 'Operations', '05_analysis-reports', 'data-quality',
    'customer-contact-review_2026-06-16.xlsx')

ISSUE_TH = {
    'assumed_bangkok': 'เดาว่าเป็นเบอร์ กทม. (เติม 02 ให้)',
    'legacy_mobile': 'มือถือแบบเก่า → แปลงเป็น 08x (อาจเลิกใช้แล้ว)',
    'inferred_area_code': 'เติมรหัสพื้นที่จากเบอร์ข้างเคียง',
    'undialable_phone': '⚠ เบอร์ไม่ครบ/โทรไม่ได้',
    'undialable_contact': '⚠ เบอร์ไม่ครบในช่องผู้ติดต่อ',
    'leftover_in_phone': '⚠ มีข้อความ/เบอร์แปลกๆ ในช่องเบอร์',
    'leftover_in_contact': '⚠ มีข้อความ/เบอร์แปลกๆ ในช่องผู้ติดต่อ',
    'person_in_phone': 'มีชื่อคนปนในช่องเบอร์ → ย้ายไปผู้ติดต่อ',
    'person_in_contact': 'มีชื่อคน+เบอร์ในช่องผู้ติดต่อ',
    'phone_in_contact': 'มีเบอร์ในช่องผู้ติดต่อ',
    'phone_in_name': 'มีเบอร์ปนในช่องชื่อ',
    'name_changed': 'ดึงเบอร์/จัดชื่อใหม่',
    'fax_in_phone': 'แยกแฟกซ์ออกจากเบอร์',
    'fax_in_contact': 'แยกแฟกซ์ออกจากผู้ติดต่อ',
    'line_in_phone': 'แยก Line ออก',
    'line_in_contact': 'แยก Line ออก',
    'note_in_phone': 'แยกหมายเหตุ (วันวางบิล ฯลฯ) ออก',
    'note_in_contact': 'แยกหมายเหตุออก',
}
STATUS_TH = {'applied': 'ระบบจัดให้อัตโนมัติ', 'confirmed': 'ระบบจัดให้ (กลุ่มยืนยัน)',
             'pending': '⚠ รอแก้มือ'}

HEADERS = [
    ('รหัส', 10), ('ชื่อร้าน', 32), ('สถานะ', 18), ('ทำไมต้องดู', 34),
    ('เบอร์เดิม', 22), ('เบอร์ตอนนี้', 22), ('ผู้ติดต่อเดิม', 24), ('ผู้ติดต่อตอนนี้', 24),
    ('แฟกซ์', 16), ('หมายเหตุ', 18),
    ('✏️แก้เบอร์', 22), ('✏️แก้ผู้ติดต่อ', 22), ('✏️แก้แฟกซ์', 16),
    ('✏️แก้หมายเหตุ', 18), ('✏️ลบ? ใส่ x', 12),
]
# 1-based column indexes that hold phone numbers → force TEXT so Excel never mangles them.
PHONE_COLS = [5, 6, 9, 11, 13]
EDIT_FIRST_COL = 11  # ✏️ columns start here


def run(db_path, out_path):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT r.customer_code, r.status, r.issues_json, r.original_json,
               c.name AS cur_name, c.phone AS cur_phone, c.fax AS cur_fax,
               c.contact AS cur_contact, c.contact_note AS cur_note
        FROM customer_contact_review r JOIN customers c ON c.code = r.customer_code
        ORDER BY CASE r.status WHEN 'pending' THEN 0 WHEN 'confirmed' THEN 1 ELSE 2 END,
                 r.customer_code
    """).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'ตรวจข้อมูลลูกค้า'

    hdr_fill = PatternFill('solid', fgColor='1F4E78')
    edit_fill = PatternFill('solid', fgColor='FFF2CC')
    pending_fill = PatternFill('solid', fgColor='FCE4D6')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # instruction row
    ws.cell(1, 1, "วิธีใช้: ดูคอลัมน์ซ้าย (อ่านอย่างเดียว). ถ้าจะแก้ → พิมพ์ในคอลัมน์ ✏️ สีเหลือง. "
                  "เบอร์ให้พิมพ์เป็นข้อความ เก็บเลข 0 หน้าไว้ (เช่น 081-2345678). ลบทิ้ง = ใส่ x. "
                  "ช่องว่าง = เก็บค่าเดิม. ส่งไฟล์กลับให้ Claudy import.")
    ws.cell(1, 1).font = Font(italic=True, color='C00000')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    hr = 2
    for ci, (title, width) in enumerate(HEADERS, start=1):
        c = ws.cell(hr, ci, title)
        c.fill = edit_fill if ci >= EDIT_FIRST_COL else hdr_fill
        c.font = Font(bold=True, color=('7F6000' if ci >= EDIT_FIRST_COL else 'FFFFFF'))
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = width

    for ri, r in enumerate(rows, start=hr + 1):
        o = json.loads(r['original_json'] or '{}')
        issues = json.loads(r['issues_json'] or '[]')
        issue_txt = ' · '.join(ISSUE_TH.get(i, i) for i in issues)
        vals = [r['customer_code'], r['cur_name'], STATUS_TH.get(r['status'], r['status']),
                issue_txt, o.get('phone', ''), r['cur_phone'] or '', o.get('contact', ''),
                r['cur_contact'] or '', r['cur_fax'] or '', r['cur_note'] or '',
                '', '', '', '', '']
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(ri, ci, v)
            c.border = border
            c.alignment = Alignment(vertical='top', wrap_text=(ci in (2, 4, 7, 8)))
            if ci in PHONE_COLS:
                c.number_format = '@'
            if ci >= EDIT_FIRST_COL:
                c.fill = edit_fill
            elif r['status'] == 'pending':
                c.fill = pending_fill

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(HEADERS))}{hr + len(rows)}"
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    print("wrote %d rows -> %s" % (len(rows), os.path.abspath(out_path)))


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--db', default=DEFAULT_DB)
    ap.add_argument('--out', default=DEFAULT_OUT)
    run(ap.parse_args().db, ap.parse_args().out)
