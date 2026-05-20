"""
Phase 2 importer — populate supplier_catalogue_* tables from an Excel
price-list file (e.g. ใบราคาสินค้าเดือน 2-69.xlsx).

What it does (one transaction per file):
  1. Resolves the supplier (must already exist in `suppliers`).
  2. Inserts (or reuses) one `supplier_catalogue_versions` row keyed by
     (supplier_id, source_file, catalogue_date). Re-running on the same
     file → reuses the version row.
  3. Parses every sheet (skipping the instructions sheet by default).
     Sub-categories (rows like **...**) are kept as `category_hint`.
  4. UPSERTs into `supplier_catalogue_items` keyed by
     (supplier_id, name_normalized). On insert: sets first_seen_version_id.
     On update: refreshes name_raw/unit/list_price/etc. and
     last_seen_version_id; only flips is_active=1 if it had been deactivated.
  5. INSERT OR REPLACE into `supplier_catalogue_price_history` for
     (item_id, version_id) — guarantees one history row per item per version.

Run:
  python scripts/import_supplier_catalogue.py \
      --file "/Volumes/ZYRINGE/ใบราคาสินค้าเดือน 2-69.xlsx" \
      --supplier "ศรีไทยเจริญโลหะกิจ" \
      --catalogue-date 2026-02 \
      [--dry-run]

Idempotent: re-running on the same file yields the same row counts (no dupes).
"""
from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

DB_PATH = Path(__file__).resolve().parent.parent / "inventory_app" / "instance" / "inventory.db"

# Sheets that contain instructions/legends, not product rows.
SKIP_SHEETS = {"ข้อแนะนำการดู CD"}

COLOR_FLAG = {
    "FFFF0000": "changed",
    "FF000000": "same",
    "FF0000FF": "new",
    "FF00B050": "preorder",
    "FF008000": "preorder",
    "FF179517": "preorder",
    "FF006600": "preorder",  # dark-green variant seen on 10 rows in 2-69
}

_BRACKET_RE = re.compile(r"[\[\(].*?[\]\)]")
_FILE_DATE_RE = re.compile(r"(\d{1,2})[-_](\d{2})")  # e.g. "2-69" or "02-69"


def color_to_flag(color):
    if color is None:
        return "same"
    rgb = getattr(color, "rgb", None)
    if not rgb or not isinstance(rgb, str):
        return "same"
    return COLOR_FLAG.get(rgb.upper(), "unknown")


def normalize(s):
    if s is None:
        return ""
    s = str(s).strip()
    s = _BRACKET_RE.sub(" ", s)
    s = s.replace('"', "นิ้ว").replace("”", "นิ้ว").replace("“", "นิ้ว")
    s = s.replace(" ", " ").replace("　", " ")
    s = " ".join(s.split())
    s = s.lower()
    return s


def tokenize(s):
    return [t for t in normalize(s).split() if t]


def parse_percent(value):
    """Coerce things like '25%', 25, 0.25, '0.05' into a percent number (e.g. 25.0)."""
    if value is None or value == "":
        return None
    if isinstance(value, str):
        v = value.strip().replace("%", "").replace(",", "")
        if not v:
            return None
        try:
            n = float(v)
        except ValueError:
            return None
        # If user wrote "25%" it parses to 25.0 — that's correct.
        # If they wrote "0.25" we treat as fraction → 25.
        if n <= 1:
            return n * 100.0
        return n
    if isinstance(value, (int, float)):
        # Numeric Excel: convention is 0.05 = 5%, 25 = 25%
        if value <= 1:
            return float(value) * 100.0
        return float(value)
    return None


def compute_net_cash(list_price, trade_pct, cash_pct):
    if list_price is None:
        return None
    p = float(list_price)
    if trade_pct:
        p = p * (1.0 - trade_pct / 100.0)
    if cash_pct:
        p = p * (1.0 - cash_pct / 100.0)
    return round(p, 4)


def parse_sheet(ws, sheet_name):
    """Yield dicts of parsed rows. Header at row 2; data starts at row 3."""
    sub_category = None
    for row_idx in range(3, ws.max_row + 1):
        c_name = ws.cell(row=row_idx, column=1)
        c_min = ws.cell(row=row_idx, column=2)
        c_unit = ws.cell(row=row_idx, column=3)
        c_price = ws.cell(row=row_idx, column=4)
        c_disc = ws.cell(row=row_idx, column=5)
        c_cash = ws.cell(row=row_idx, column=6)

        name = c_name.value
        if name is None:
            continue
        name = str(name).strip()
        if not name:
            continue

        if name.startswith("**") and name.endswith("**"):
            sub_category = name.strip("*").strip()
            continue

        if c_price.value is None:
            continue

        try:
            list_price = float(c_price.value)
        except (TypeError, ValueError):
            continue

        trade_pct = parse_percent(c_disc.value)
        cash_pct = parse_percent(c_cash.value)

        yield {
            "name_raw": name,
            "name_normalized": normalize(name),
            "name_tokens": json.dumps(tokenize(name), ensure_ascii=False),
            "category_hint": sub_category,
            "sheet_name": sheet_name,
            "unit": (str(c_unit.value).strip() if c_unit.value else None),
            "min_order_qty": _safe_float(c_min.value),
            "list_price": list_price,
            "trade_discount_pct": trade_pct,
            "cash_discount_pct": cash_pct,
            "net_cash_price": compute_net_cash(list_price, trade_pct, cash_pct),
            "price_change_flag": color_to_flag(c_price.font.color) if c_price.font else "same",
        }


def _safe_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def derive_catalogue_date(file_path):
    """Try to pull 'YYYY-MM' out of the filename like '2-69' = ก.พ. 2569 = 2026-02."""
    m = _FILE_DATE_RE.search(Path(file_path).stem)
    if not m:
        return None
    month = int(m.group(1))
    be_year_2 = int(m.group(2))  # 2-digit Buddhist Era year (e.g. 69)
    # 2569 BE = 2026 CE. Map 2-digit BE year → CE.
    ce_year = 2500 + be_year_2 - 543
    return f"{ce_year:04d}-{month:02d}"


def get_or_create_version(conn, supplier_id, source_file, catalogue_date, imported_by, note):
    cur = conn.execute(
        """
        SELECT id FROM supplier_catalogue_versions
        WHERE supplier_id = ? AND source_file = ?
          AND COALESCE(catalogue_date,'') = COALESCE(?, '')
        """,
        (supplier_id, source_file, catalogue_date),
    )
    row = cur.fetchone()
    if row:
        return row[0], False
    cur = conn.execute(
        """
        INSERT INTO supplier_catalogue_versions
          (supplier_id, source_file, catalogue_date, imported_by, note)
        VALUES (?, ?, ?, ?, ?)
        """,
        (supplier_id, source_file, catalogue_date, imported_by, note),
    )
    return cur.lastrowid, True


def upsert_item(conn, supplier_id, version_id, row):
    cur = conn.execute(
        """
        SELECT id FROM supplier_catalogue_items
        WHERE supplier_id = ? AND name_normalized = ?
        """,
        (supplier_id, row["name_normalized"]),
    )
    existing = cur.fetchone()
    if existing:
        item_id = existing[0]
        conn.execute(
            """
            UPDATE supplier_catalogue_items SET
              name_raw = ?,
              name_tokens = ?,
              category_hint = ?,
              sheet_name = ?,
              unit = ?,
              min_order_qty = ?,
              list_price = ?,
              trade_discount_pct = ?,
              cash_discount_pct = ?,
              net_cash_price = ?,
              price_change_flag = ?,
              last_seen_version_id = ?,
              is_active = 1,
              updated_at = datetime('now','localtime')
            WHERE id = ?
            """,
            (
                row["name_raw"], row["name_tokens"], row["category_hint"], row["sheet_name"],
                row["unit"], row["min_order_qty"], row["list_price"],
                row["trade_discount_pct"], row["cash_discount_pct"], row["net_cash_price"],
                row["price_change_flag"], version_id, item_id,
            ),
        )
        action = "updated"
    else:
        cur = conn.execute(
            """
            INSERT INTO supplier_catalogue_items
              (supplier_id, name_raw, name_normalized, name_tokens, category_hint,
               sheet_name, unit, min_order_qty, list_price,
               trade_discount_pct, cash_discount_pct, net_cash_price,
               price_change_flag, first_seen_version_id, last_seen_version_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                supplier_id, row["name_raw"], row["name_normalized"], row["name_tokens"],
                row["category_hint"], row["sheet_name"], row["unit"], row["min_order_qty"],
                row["list_price"], row["trade_discount_pct"], row["cash_discount_pct"],
                row["net_cash_price"], row["price_change_flag"], version_id, version_id,
            ),
        )
        item_id = cur.lastrowid
        action = "inserted"

    conn.execute(
        """
        INSERT OR REPLACE INTO supplier_catalogue_price_history
          (item_id, version_id, list_price, trade_discount_pct, cash_discount_pct,
           net_cash_price, unit, price_change_flag)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            item_id, version_id, row["list_price"], row["trade_discount_pct"],
            row["cash_discount_pct"], row["net_cash_price"], row["unit"],
            row["price_change_flag"],
        ),
    )
    return action


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True, help="Excel catalogue path")
    ap.add_argument("--supplier", required=True, help="Supplier name (must exist in suppliers.name)")
    ap.add_argument("--catalogue-date", default=None, help="ISO month YYYY-MM (e.g. 2026-02). Inferred from filename if omitted.")
    ap.add_argument("--imported-by", default=None)
    ap.add_argument("--note", default=None)
    ap.add_argument("--dry-run", action="store_true", help="Parse + report; do not write to DB")
    ap.add_argument("--db", default=str(DB_PATH))
    args = ap.parse_args()

    file_path = Path(args.file)
    if not file_path.exists():
        print(f"ERROR: file not found: {file_path}", file=sys.stderr)
        sys.exit(1)

    catalogue_date = args.catalogue_date or derive_catalogue_date(file_path)

    print(f"Loading workbook: {file_path.name}", file=sys.stderr)
    wb = load_workbook(file_path, data_only=True)

    parsed_rows = []
    color_counter = Counter()
    sheet_counter = Counter()
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        n = 0
        for row in parse_sheet(ws, sheet_name):
            parsed_rows.append(row)
            color_counter[row["price_change_flag"]] += 1
            n += 1
        sheet_counter[sheet_name] = n
        print(f"  sheet '{sheet_name}': {n} item rows", file=sys.stderr)

    print(f"Total rows parsed: {len(parsed_rows)}", file=sys.stderr)
    print(f"Color flags: {dict(color_counter)}", file=sys.stderr)
    print(f"Catalogue date: {catalogue_date or '(not derived)'}", file=sys.stderr)

    if args.dry_run:
        print("DRY RUN — no DB writes.", file=sys.stderr)
        return

    conn = sqlite3.connect(args.db)
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        with conn:
            cur = conn.execute("SELECT id FROM suppliers WHERE name = ?", (args.supplier,))
            srow = cur.fetchone()
            if not srow:
                print(f"ERROR: supplier not found: {args.supplier!r}", file=sys.stderr)
                sys.exit(2)
            supplier_id = srow[0]

            version_id, created = get_or_create_version(
                conn, supplier_id, str(file_path), catalogue_date,
                args.imported_by, args.note,
            )
            print(f"version_id={version_id} ({'NEW' if created else 'reused'})", file=sys.stderr)

            actions = Counter()
            for row in parsed_rows:
                actions[upsert_item(conn, supplier_id, version_id, row)] += 1

        print(f"DONE. items inserted={actions['inserted']} updated={actions['updated']}", file=sys.stderr)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
