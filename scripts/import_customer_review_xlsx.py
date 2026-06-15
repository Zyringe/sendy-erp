#!/usr/bin/env python
"""Import Put's edited customer-review Excel back into Sendy.

Reads ONLY the ✏️ columns (what Put typed) — the read-only reference columns are ignored, so
Excel mangling them on save can't hurt. Validates every corrected phone (warns hard if Excel
dropped a leading zero or sci-notated it). Default = preview (diff + warnings); --apply commits
in a transaction. Originals stay frozen in contact_orig_json (reversible). Edited rows are marked
'confirmed'.

Edit columns (1-based): 11=phone, 12=contact, 13=fax, 14=note, 15=clear?(x → wipe phone+fax).
Run: python scripts/import_customer_review_xlsx.py FILE.xlsx [--db PATH] [--apply] [--user NAME]
"""
import argparse
import json
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
APP = os.path.join(HERE, '..', 'inventory_app')
sys.path.insert(0, APP)
import sqlite3  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from customer_contact_normalize import is_valid_thai_phone  # noqa: E402

DEFAULT_DB = os.path.join(APP, 'instance', 'inventory.db')
COL = {'code': 1, 'phone': 11, 'contact': 12, 'fax': 13, 'note': 14, 'clear': 15}
DATA_START = 3


def _cell(ws, r, c):
    v = ws.cell(r, c).value
    return v


def _phone_warnings(raw):
    """Flag values that look like Excel mangled them, or aren't dialable Thai numbers."""
    warns = []
    if isinstance(raw, (int, float)):
        warns.append("Excel เก็บเป็นตัวเลข (เลข 0 หน้าน่าจะหาย) — พิมพ์ใหม่เป็นข้อความ")
        raw = str(int(raw))
    s = str(raw).strip()
    for tok in re.split(r'[\s,]+', s):
        d = re.sub(r'\D', '', tok)
        if len(d) < 6:
            continue
        if not d.startswith('0'):
            warns.append("'%s' ไม่ขึ้นต้นด้วย 0 (เลข 0 หน้าหาย?)" % tok)
        elif not is_valid_thai_phone(tok):
            warns.append("'%s' รูปแบบเบอร์ไม่ปกติ" % tok)
    return warns


def run(path, db_path, apply, user):
    ws = load_workbook(path, data_only=True).active
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    edits = []          # (code, dict-of-changes, warnings)
    r = DATA_START
    while True:
        code = _cell(ws, r, COL['code'])
        if code is None:
            break
        code = str(code).strip()
        ep = _cell(ws, r, COL['phone']); ec = _cell(ws, r, COL['contact'])
        ef = _cell(ws, r, COL['fax']);   en = _cell(ws, r, COL['note'])
        clr = _cell(ws, r, COL['clear'])
        has_edit = any(v not in (None, '') for v in (ep, ec, ef, en, clr))
        if has_edit:
            cur = conn.execute(
                "SELECT phone,fax,contact,contact_note FROM customers WHERE code=?",
                (code,)).fetchone()
            ch, warns = {}, []
            if str(clr).strip().lower() == 'x':
                ch['phone'] = None
                ch['fax'] = None
            if ep not in (None, ''):
                ch['phone'] = str(ep).strip()
                warns += _phone_warnings(ep)
            if ef not in (None, ''):
                ch['fax'] = str(ef).strip()
                warns += _phone_warnings(ef)
            if ec not in (None, ''):
                ch['contact'] = str(ec).strip()
            if en not in (None, ''):
                ch['note'] = str(en).strip()
            edits.append((code, ch, warns, cur))
        r += 1

    print("rows with edits: %d" % len(edits))
    warn_rows = [e for e in edits if e[2]]
    print("rows with warnings: %d\n" % len(warn_rows))
    for code, ch, warns, cur in edits[:40]:
        cur_d = dict(cur) if cur else {}
        diff = ", ".join("%s: %r→%r" % (k, cur_d.get(k if k != 'note' else 'contact_note'),
                                        ch[k]) for k in ch)
        print("  [%s] %s%s" % (code, diff, ("  ⚠ " + "; ".join(warns)) if warns else ""))
    if len(edits) > 40:
        print("  ... (%d more)" % (len(edits) - 40))

    if not apply:
        print("\nPREVIEW — nothing written. Fix any ⚠ rows, then re-run with --apply.")
        conn.close()
        return
    if warn_rows:
        print("\n!! %d rows still have warnings — resolve them or they apply as-is. "
              "Re-run with --apply to proceed anyway." % len(warn_rows))

    try:
        conn.execute('BEGIN IMMEDIATE')
        for code, ch, warns, cur in edits:
            sets, params = [], []
            colmap = {'phone': 'phone', 'fax': 'fax', 'contact': 'contact', 'note': 'contact_note'}
            for k, v in ch.items():
                sets.append("%s=?" % colmap[k])
                params.append(v or None)
            sets.append("contact_normalized_at=datetime('now','localtime')")
            sets.append("contact_normalized_by=?")
            params.append(user)
            params.append(code)
            conn.execute("UPDATE customers SET %s WHERE code=?" % ",".join(sets), params)
            conn.execute("""UPDATE customer_contact_review
                            SET status='confirmed', reviewed_by=?,
                                reviewed_at=datetime('now','localtime')
                            WHERE customer_code=?""", (user, code))
        conn.execute('COMMIT')
    except Exception:
        conn.execute('ROLLBACK')
        conn.close()
        raise
    print("\nAPPLIED %d edited rows." % len(edits))
    conn.close()


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('file')
    ap.add_argument('--db', default=DEFAULT_DB)
    ap.add_argument('--apply', action='store_true')
    ap.add_argument('--user', default='excel_review')
    a = ap.parse_args()
    run(a.file, a.db, a.apply, a.user)
