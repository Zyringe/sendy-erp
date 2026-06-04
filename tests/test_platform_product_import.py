"""
TDD tests for the Full Shopee + Lazada product-info import.

Spec: projects/pack-unpack-and-e-commerce/marketplace-product-import-spec.md
Schema: migration 094 (already applied — platform_products + 8 platform_skus cols)

Tests:
  1. parse_lazada English-header: name/variation_id/price/stock all populated
  2. Multi-file merge: Shopee product gets description + cover_image; variation gets weight; image collapse
  3. Idempotency/regression: import real files twice → internal_product_id + qty_per_sale unchanged, row counts stable
  4. UPSERT contract: _propagate never overwrites non-NULL internal_product_id
  5. import_platform_products upsert: conflict updates fields, never nulls
"""
import io
import json
import os
import shutil
import sqlite3

import pandas as pd
import pytest

# Allow importing without starting the app server
os.environ.setdefault('SKIP_DB_INIT', '1')
os.environ.setdefault('SECRET_KEY', 'test-only-secret')
os.environ.setdefault('ADMIN_PASSWORD', 'test-only-admin')

import parse_platform as pp
from parse_platform import (
    parse_lazada,
    parse_shopee_product_files,
    parse_lazada_product_files,
)
import models

# ── Real-file paths ─────────────────────────────────────────────────────────────

_REPO = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
_SHOPEE_DIR = os.path.join(_REPO, 'E-Commerce', 'Shopee', 'sendaibyboonsawat', '01_product-info')
_LAZADA_DIR = os.path.join(_REPO, 'E-Commerce', 'Lazada', '01_product-info')

_HAS_REAL_FILES = os.path.isdir(_SHOPEE_DIR) and os.path.isdir(_LAZADA_DIR)


# ── Helpers ─────────────────────────────────────────────────────────────────────

def _make_lazada_pricestock_xlsx(rows_data):
    """
    Build a minimal in-memory Lazada pricestock xlsx with English headers.
    rows_data: list of dicts with keys from pricestock columns.
    Row 0 = headers (English), rows 1-3 = instruction rows (non-digit Product ID), rows 4+ = data.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'template'

    cols = ['Product ID', 'catId', 'Product Name', 'currencyCode', 'sku.skuId',
            'status', 'Shop SKU', 'SpecialPrice', 'SpecialPrice Start',
            'SpecialPrice End', 'Price', 'SellerSKU', 'บุญสวัสดิ์นำชัย',
            'Variations Combo', 'tr(s-wb-product@md5key)']

    # Row 1: English headers
    for ci, col in enumerate(cols, 1):
        ws.cell(row=1, column=ci, value=col)

    # Rows 2-4: instruction rows (non-digit Product ID so the filter skips them)
    for instr_row in range(2, 5):
        ws.cell(row=instr_row, column=1, value='Mandatory/Optional instruction')

    # Data rows start at row 5
    for ri, r in enumerate(rows_data, 5):
        for ci, col in enumerate(cols, 1):
            ws.cell(row=ri, column=ci, value=r.get(col))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _make_lazada_pricestock_xlsx_thai(rows_data):
    """Build a Lazada pricestock xlsx with LEGACY THAI headers (for backward-compat test)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'template'

    # Old Thai columns — only the ones the legacy parse_lazada used
    cols = ['Product ID', 'catId', 'ชื่อสินค้า', 'currencyCode', 'sku.skuId',
            'status', 'ร้าน sku', 'SpecialPrice', 'SpecialPrice Start',
            'SpecialPrice End', 'ราคา', 'SellerSKU', 'บุญสวัสดิ์นำชัย',
            'Variations Combo', 'tr(s-wb-product@md5key)']

    for ci, col in enumerate(cols, 1):
        ws.cell(row=1, column=ci, value=col)

    for instr_row in range(2, 5):
        ws.cell(row=instr_row, column=1, value='instruction')

    for ri, r in enumerate(rows_data, 5):
        for ci, col in enumerate(cols, 1):
            ws.cell(row=ri, column=ci, value=r.get(col))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── 1. parse_lazada English-header fix ──────────────────────────────────────────

class TestParseLazadaEnglishHeaders:

    SAMPLE_ROWS = [
        {
            'Product ID': '421038989',
            'catId': '62556202',
            'Product Name': 'SENDAI ลูกบิดประตู หัวคริสตัลอย่างดี',
            'sku.skuId': '803226494',
            'Shop SKU': '421038989_TH-803226494',
            'Price': '720.00',
            'SpecialPrice': '599.00',
            'SellerSKU': 'Brown',
            'บุญสวัสดิ์นำชัย': '2',
            'Variations Combo': 'Brown',
        },
        {
            'Product ID': '421038990',
            'catId': '62556202',
            'Product Name': 'SENDAI กลอนห้องน้ำ STL-430',
            'sku.skuId': '803226495',
            'Shop SKU': '421038990_TH-803226495',
            'Price': '150.00',
            'SpecialPrice': None,
            'SellerSKU': '431SN',
            'บุญสวัสดิ์นำชัย': '5',
            'Variations Combo': None,
        },
    ]

    def test_english_headers_product_name_populated(self):
        xlsx = _make_lazada_pricestock_xlsx(self.SAMPLE_ROWS)
        records = parse_lazada(xlsx)
        names = [r['product_name'] for r in records]
        assert all(n and n != '0' and len(n) > 2 for n in names), \
            f"product_name empty/zero with English headers: {names}"

    def test_english_headers_variation_id_populated(self):
        xlsx = _make_lazada_pricestock_xlsx(self.SAMPLE_ROWS)
        records = parse_lazada(xlsx)
        vids = [r['variation_id'] for r in records]
        assert all(v and '_TH-' in str(v) for v in vids), \
            f"variation_id not Shop-SKU form with English headers: {vids}"

    def test_english_headers_price_populated(self):
        xlsx = _make_lazada_pricestock_xlsx(self.SAMPLE_ROWS)
        records = parse_lazada(xlsx)
        prices = [r['price'] for r in records]
        assert all(isinstance(p, float) and p > 0 for p in prices), \
            f"price not parsed with English headers: {prices}"

    def test_english_headers_stock_populated(self):
        xlsx = _make_lazada_pricestock_xlsx(self.SAMPLE_ROWS)
        records = parse_lazada(xlsx)
        stocks = [r['stock'] for r in records]
        assert all(s is not None for s in stocks), \
            f"stock None with English headers: {stocks}"

    def test_returns_correct_row_count(self):
        xlsx = _make_lazada_pricestock_xlsx(self.SAMPLE_ROWS)
        records = parse_lazada(xlsx)
        assert len(records) == 2

    def test_legacy_thai_headers_still_work(self):
        """Backward-compat: old Thai-header files must still parse correctly."""
        rows_thai = [
            {
                'Product ID': '111000001',
                'catId': '12345',
                'ชื่อสินค้า': 'สินค้าทดสอบ ภาษาไทย',
                'sku.skuId': '900000001',
                'ร้าน sku': '111000001_TH-900000001',
                'ราคา': '100.00',
                'SellerSKU': 'TH-OLD',
                'บุญสวัสดิ์นำชัย': '10',
                'Variations Combo': 'สีแดง',
            }
        ]
        xlsx = _make_lazada_pricestock_xlsx_thai(rows_thai)
        records = parse_lazada(xlsx)
        assert len(records) == 1
        r = records[0]
        assert r['product_name'] == 'สินค้าทดสอบ ภาษาไทย'
        assert r['variation_id'] == '111000001_TH-900000001'
        assert r['price'] == 100.0
        assert r['stock'] == 10

    @pytest.mark.skipif(not _HAS_REAL_FILES, reason="real Lazada files not present")
    def test_real_lazada_pricestock_all_rows_populated(self):
        """Integration: real pricestock file must yield 396 rows with non-empty key fields."""
        import glob
        files = glob.glob(os.path.join(_LAZADA_DIR, 'pricestock*.xlsx'))
        assert files, "no pricestock*.xlsx found"
        with open(files[0], 'rb') as fh:
            records = parse_lazada(io.BytesIO(fh.read()))
        assert len(records) == 396, f"expected 396, got {len(records)}"
        names = [r for r in records if not r['product_name']]
        vids = [r for r in records if not r['variation_id']]
        prices = [r for r in records if r['price'] is None]
        stocks = [r for r in records if r['stock'] is None]
        assert not names, f"{len(names)} rows with empty product_name"
        assert not vids, f"{len(vids)} rows with empty variation_id"
        assert not prices, f"{len(prices)} rows with None price"
        assert not stocks, f"{len(stocks)} rows with None stock"


# ── 2. Multi-file merge ──────────────────────────────────────────────────────────

@pytest.mark.skipif(not _HAS_REAL_FILES, reason="real product-info files not present")
class TestMultiFileMerge:

    def test_shopee_product_rows_populated(self):
        product_rows, _ = parse_shopee_product_files(_SHOPEE_DIR)
        assert len(product_rows) >= 280, f"expected >=280 product rows, got {len(product_rows)}"
        # Every row must have product_id_str
        missing_pid = [r for r in product_rows if not r.get('product_id_str')]
        assert not missing_pid, f"{len(missing_pid)} Shopee product rows missing product_id_str"

    def test_shopee_variation_rows_populated(self):
        _, var_rows = parse_shopee_product_files(_SHOPEE_DIR)
        assert len(var_rows) == 661, f"expected 661 variation rows, got {len(var_rows)}"

    def test_shopee_product_gets_cover_image(self):
        product_rows, _ = parse_shopee_product_files(_SHOPEE_DIR)
        with_cover = [r for r in product_rows if r.get('cover_image_url')]
        assert len(with_cover) > 0, "no Shopee products have cover_image_url"
        # Cover URLs should look like https://...
        for r in with_cover[:5]:
            assert r['cover_image_url'].startswith('http'), \
                f"cover_image_url not a URL: {r['cover_image_url']}"

    def test_shopee_product_description_from_basic(self):
        product_rows, _ = parse_shopee_product_files(_SHOPEE_DIR)
        with_desc = [r for r in product_rows if r.get('description')]
        assert len(with_desc) > 0, "no Shopee products have description"

    def test_shopee_variation_gets_weight(self):
        _, var_rows = parse_shopee_product_files(_SHOPEE_DIR)
        with_weight = [r for r in var_rows if r.get('weight_kg') is not None]
        assert len(with_weight) > 600, \
            f"expected >600 variations with weight_kg, got {len(with_weight)}"

    def test_shopee_image_urls_json_array(self):
        product_rows, _ = parse_shopee_product_files(_SHOPEE_DIR)
        for r in product_rows[:20]:
            assert isinstance(r.get('image_urls'), str), \
                f"image_urls should be JSON string, got {type(r.get('image_urls'))}"
            arr = json.loads(r['image_urls'])
            assert isinstance(arr, list), f"image_urls decoded to non-list: {arr}"
            # All entries should be URLs
            for url in arr:
                assert url.startswith('http'), f"image in array not a URL: {url}"

    def test_lazada_product_rows_populated(self):
        product_rows, _ = parse_lazada_product_files(_LAZADA_DIR)
        assert len(product_rows) == 179, f"expected 179, got {len(product_rows)}"

    def test_lazada_variation_rows_populated(self):
        _, var_rows = parse_lazada_product_files(_LAZADA_DIR)
        assert len(var_rows) == 396, f"expected 396, got {len(var_rows)}"

    def test_lazada_variation_gets_weight(self):
        _, var_rows = parse_lazada_product_files(_LAZADA_DIR)
        with_weight = [r for r in var_rows if r.get('weight_kg') is not None]
        assert len(with_weight) >= 390, \
            f"expected >=390 Lazada variations with weight_kg, got {len(with_weight)}"

    def test_lazada_product_cover_image(self):
        product_rows, _ = parse_lazada_product_files(_LAZADA_DIR)
        with_cover = [r for r in product_rows if r.get('cover_image_url')]
        assert len(with_cover) > 0, "no Lazada products have cover_image_url"

    def test_lazada_image_urls_json_array(self):
        product_rows, _ = parse_lazada_product_files(_LAZADA_DIR)
        for r in product_rows[:20]:
            arr = json.loads(r.get('image_urls', '[]'))
            assert isinstance(arr, list)

    def test_lazada_attribute_brand_populated(self):
        product_rows, _ = parse_lazada_product_files(_LAZADA_DIR)
        with_brand = [r for r in product_rows if r.get('brand')]
        # Most products have a brand from attribute sheets
        assert len(with_brand) > 100, \
            f"expected >100 products with brand, got {len(with_brand)}"

    def test_lazada_variation_image_from_skuimg(self):
        _, var_rows = parse_lazada_product_files(_LAZADA_DIR)
        with_img = [r for r in var_rows if r.get('variation_image_url')]
        assert len(with_img) > 0, "no Lazada variations have variation_image_url from skuimg"


# ── 3. Idempotency / regression guard ───────────────────────────────────────────

@pytest.mark.skipif(not _HAS_REAL_FILES, reason="real product-info files not present")
class TestIdempotency:

    def _get_mappings(self, db_path):
        """Return dict {(platform, variation_id): (internal_product_id, qty_per_sale)}."""
        conn = sqlite3.connect(db_path)
        rows = conn.execute(
            "SELECT platform, variation_id, internal_product_id, qty_per_sale "
            "FROM platform_skus"
        ).fetchall()
        conn.close()
        return {(r[0], r[1]): (r[2], r[3]) for r in rows}

    def _get_counts(self, db_path):
        conn = sqlite3.connect(db_path)
        skus = conn.execute(
            "SELECT platform, COUNT(*), SUM(CASE WHEN internal_product_id IS NOT NULL THEN 1 ELSE 0 END) "
            "FROM platform_skus GROUP BY platform"
        ).fetchall()
        products = conn.execute(
            "SELECT platform, COUNT(*) FROM platform_products GROUP BY platform"
        ).fetchall()
        conn.close()
        return {r[0]: (r[1], r[2]) for r in skus}, {r[0]: r[1] for r in products}

    def _run_import(self, db_path, platform, folder):
        import config
        import database as db_mod
        orig_path = config.DATABASE_PATH
        try:
            config.DATABASE_PATH = db_path
            db_mod.DATABASE_PATH = db_path

            if platform == 'shopee':
                prod_rows, var_rows = parse_shopee_product_files(folder)
            else:
                prod_rows, var_rows = parse_lazada_product_files(folder)

            models.import_platform_products(platform, prod_rows)
            models.import_platform_skus(platform, var_rows)
        finally:
            config.DATABASE_PATH = orig_path
            db_mod.DATABASE_PATH = orig_path

    def test_shopee_idempotent_run1_vs_run2(self, tmp_db, monkeypatch):
        """Import Shopee twice → mappings byte-identical, row counts stable."""
        # Run 1
        self._run_import(tmp_db, 'shopee', _SHOPEE_DIR)
        mappings_after_run1 = self._get_mappings(tmp_db)
        skus1, prods1 = self._get_counts(tmp_db)

        # Run 2
        self._run_import(tmp_db, 'shopee', _SHOPEE_DIR)
        mappings_after_run2 = self._get_mappings(tmp_db)
        skus2, prods2 = self._get_counts(tmp_db)

        # Counts stable
        assert skus1['shopee'][0] == skus2['shopee'][0], "Shopee sku row count changed on run 2"
        assert prods1.get('shopee', 0) == prods2.get('shopee', 0), \
            "Shopee product row count changed on run 2"

        # No mapping degradation: every (variation_id) with non-NULL internal_product_id in run1
        # must still have the same non-NULL value in run2
        degraded = [
            k for k, (pid1, qty1) in mappings_after_run1.items()
            if k[0] == 'shopee' and pid1 is not None
            and mappings_after_run2.get(k, (None, None))[0] != pid1
        ]
        assert not degraded, \
            f"{len(degraded)} Shopee mappings changed/nulled on run 2: {degraded[:3]}"

        # qty_per_sale also stable for mapped rows
        qty_changed = [
            k for k, (pid1, qty1) in mappings_after_run1.items()
            if k[0] == 'shopee' and pid1 is not None
            and mappings_after_run2.get(k, (None, None))[1] != qty1
        ]
        assert not qty_changed, \
            f"{len(qty_changed)} Shopee qty_per_sale values changed on run 2"

    def test_lazada_idempotent_run1_vs_run2(self, tmp_db, monkeypatch):
        """Import Lazada twice → mappings byte-identical, row counts stable."""
        self._run_import(tmp_db, 'lazada', _LAZADA_DIR)
        mappings_after_run1 = self._get_mappings(tmp_db)
        skus1, prods1 = self._get_counts(tmp_db)

        self._run_import(tmp_db, 'lazada', _LAZADA_DIR)
        mappings_after_run2 = self._get_mappings(tmp_db)
        skus2, prods2 = self._get_counts(tmp_db)

        assert skus1['lazada'][0] == skus2['lazada'][0], "Lazada sku row count changed"
        assert prods1.get('lazada', 0) == prods2.get('lazada', 0), \
            "Lazada product row count changed"

        degraded = [
            k for k, (pid1, qty1) in mappings_after_run1.items()
            if k[0] == 'lazada' and pid1 is not None
            and mappings_after_run2.get(k, (None, None))[0] != pid1
        ]
        assert not degraded, \
            f"{len(degraded)} Lazada mappings changed/nulled on run 2: {degraded[:3]}"

    def test_pre_existing_mappings_preserved(self, tmp_db, monkeypatch):
        """Pre-existing internal_product_id values (from live DB before import) are
        byte-identical after a full import. This is the core regression guard against
        the old DELETE-and-reinsert behavior."""
        # Capture pre-import mapped set (independent signal: direct DB query)
        pre_mappings = self._get_mappings(tmp_db)
        pre_mapped_shopee = {k: v for k, v in pre_mappings.items()
                             if k[0] == 'shopee' and v[0] is not None}
        pre_mapped_lazada = {k: v for k, v in pre_mappings.items()
                             if k[0] == 'lazada' and v[0] is not None}

        # Run full import for both platforms
        self._run_import(tmp_db, 'shopee', _SHOPEE_DIR)
        self._run_import(tmp_db, 'lazada', _LAZADA_DIR)

        post_mappings = self._get_mappings(tmp_db)

        # Every pre-existing mapping must survive with the same pid
        shopee_lost = [
            k for k, (pid, qty) in pre_mapped_shopee.items()
            if post_mappings.get(k, (None, None))[0] != pid
        ]
        lazada_lost = [
            k for k, (pid, qty) in pre_mapped_lazada.items()
            if post_mappings.get(k, (None, None))[0] != pid
        ]

        assert not shopee_lost, \
            (f"{len(shopee_lost)}/{len(pre_mapped_shopee)} Shopee pre-existing mappings lost. "
             f"Examples: {shopee_lost[:3]}")
        assert not lazada_lost, \
            (f"{len(lazada_lost)}/{len(pre_mapped_lazada)} Lazada pre-existing mappings lost. "
             f"Examples: {lazada_lost[:3]}")


# ── 4. _propagate guard ──────────────────────────────────────────────────────────

def test_propagate_never_overwrites_non_null_mapping(tmp_db, monkeypatch):
    """_propagate_listings_to_platform_skus must only set internal_product_id
    where it is currently NULL. A row that already has a valid mapping must be
    left untouched even if ecommerce_listings name-match would suggest a different pid."""
    import config
    import database as db_mod
    monkeypatch.setattr(config, 'DATABASE_PATH', tmp_db)
    monkeypatch.setattr(db_mod, 'DATABASE_PATH', tmp_db)

    conn = sqlite3.connect(tmp_db)
    conn.row_factory = sqlite3.Row

    # Find a shopee row that is already mapped
    mapped_row = conn.execute(
        "SELECT id, variation_id, product_name, variation_name, seller_sku, "
        "internal_product_id, qty_per_sale "
        "FROM platform_skus WHERE platform='shopee' AND internal_product_id IS NOT NULL "
        "LIMIT 1"
    ).fetchone()
    if mapped_row is None:
        pytest.skip("no pre-mapped shopee rows in test DB")

    original_pid = mapped_row['internal_product_id']
    original_qty = mapped_row['qty_per_sale']
    variation_id = mapped_row['variation_id']

    # Ensure an ecommerce_listings row exists that would name-match this product
    # but with a DIFFERENT product_id (to prove we don't overwrite)
    fake_pid = original_pid + 9999 if original_pid else 99999
    conn.execute("""
        INSERT OR IGNORE INTO ecommerce_listings
          (platform, item_name, variation, seller_sku, listing_key, product_id, qty_per_sale)
        VALUES ('shopee', ?, ?, ?, 'test_propagate_guard', ?, 999.0)
    """, (
        mapped_row['product_name'],
        mapped_row['variation_name'],
        mapped_row['seller_sku'],
        fake_pid,
    ))
    conn.commit()

    # Call _propagate directly via models (it uses the connection internally)
    from models import _propagate_listings_to_platform_skus
    _propagate_listings_to_platform_skus(conn, 'shopee')
    conn.commit()

    after = conn.execute(
        "SELECT internal_product_id, qty_per_sale FROM platform_skus "
        "WHERE platform='shopee' AND variation_id=?",
        (variation_id,)
    ).fetchone()
    conn.close()

    assert after['internal_product_id'] == original_pid, \
        (f"_propagate overwrote existing mapping: "
         f"original={original_pid}, after={after['internal_product_id']}")
    # qty_per_sale should also be unchanged (we passed 999.0 as fake, it should not apply)
    assert after['qty_per_sale'] == original_qty, \
        f"_propagate changed qty_per_sale on mapped row: {original_qty} → {after['qty_per_sale']}"


# ── 5. import_platform_products upsert ──────────────────────────────────────────

def test_import_platform_products_upsert(tmp_db, monkeypatch):
    """insert then re-import same product_id_str → fields updated, count stable."""
    import config
    import database as db_mod
    monkeypatch.setattr(config, 'DATABASE_PATH', tmp_db)
    monkeypatch.setattr(db_mod, 'DATABASE_PATH', tmp_db)

    rows1 = [{
        'product_id_str': 'TEST_PROD_1',
        'product_name': 'ชื่อเดิม',
        'name_en': 'Old Name',
        'description': 'desc1',
        'category_name': None,
        'brand': None,
        'cover_image_url': 'https://example.com/img1.jpg',
        'image_urls': '["https://example.com/img1.jpg"]',
        'status': 'active',
        'raw_json': '{}',
    }]

    models.import_platform_products('shopee', rows1)

    count_after_1 = sqlite3.connect(tmp_db).execute(
        "SELECT COUNT(*) FROM platform_products WHERE platform='shopee' AND product_id_str='TEST_PROD_1'"
    ).fetchone()[0]
    assert count_after_1 == 1

    # Re-import with updated fields
    rows2 = [{
        'product_id_str': 'TEST_PROD_1',
        'product_name': 'ชื่อใหม่',
        'name_en': 'New Name',
        'description': 'desc2',
        'cover_image_url': 'https://example.com/img2.jpg',
        'image_urls': '["https://example.com/img2.jpg"]',
        'status': 'active',
        'raw_json': '{"updated": true}',
    }]

    models.import_platform_products('shopee', rows2)

    conn = sqlite3.connect(tmp_db)
    row = conn.execute(
        "SELECT product_name, name_en, description FROM platform_products "
        "WHERE platform='shopee' AND product_id_str='TEST_PROD_1'"
    ).fetchone()
    count_after_2 = conn.execute(
        "SELECT COUNT(*) FROM platform_products WHERE platform='shopee' AND product_id_str='TEST_PROD_1'"
    ).fetchone()[0]
    conn.close()

    assert count_after_2 == 1, "upsert created duplicate row"
    assert row[0] == 'ชื่อใหม่', f"product_name not updated: {row[0]}"
    assert row[1] == 'New Name', f"name_en not updated: {row[1]}"


def test_import_platform_skus_upsert_no_delete(tmp_db, monkeypatch):
    """The new import_platform_skus must NOT delete existing rows.
    A re-import of a subset must leave unmentioned rows intact."""
    import config
    import database as db_mod
    monkeypatch.setattr(config, 'DATABASE_PATH', tmp_db)
    monkeypatch.setattr(db_mod, 'DATABASE_PATH', tmp_db)

    conn = sqlite3.connect(tmp_db)
    # Count rows before
    before_count = conn.execute(
        "SELECT COUNT(*) FROM platform_skus WHERE platform='shopee'"
    ).fetchone()[0]

    # Get one existing variation_id
    existing = conn.execute(
        "SELECT variation_id, product_id_str, product_name FROM platform_skus "
        "WHERE platform='shopee' LIMIT 1"
    ).fetchone()
    conn.close()

    if not existing:
        pytest.skip("no shopee rows in test DB")

    # Re-import just that one row with updated price
    records = [{
        'variation_id': existing[0],
        'product_id_str': existing[1],
        'product_name': existing[2],
        'variation_name': None,
        'parent_sku': None,
        'seller_sku': None,
        'price': 9999.0,
        'special_price': None,
        'stock': 0,
        'raw_json': '{}',
        'weight_kg': None,
        'length_cm': None,
        'width_cm': None,
        'height_cm': None,
        'gtin': None,
        'special_price_start': None,
        'special_price_end': None,
        'variation_image_url': None,
    }]

    models.import_platform_skus('shopee', records)

    conn2 = sqlite3.connect(tmp_db)
    after_count = conn2.execute(
        "SELECT COUNT(*) FROM platform_skus WHERE platform='shopee'"
    ).fetchone()[0]
    updated_price = conn2.execute(
        "SELECT price FROM platform_skus WHERE platform='shopee' AND variation_id=?",
        (existing[0],)
    ).fetchone()[0]
    conn2.close()

    # No rows should be deleted
    assert after_count == before_count, \
        f"import deleted rows: before={before_count}, after={after_count}"
    assert updated_price == 9999.0, f"price not updated by upsert: {updated_price}"
