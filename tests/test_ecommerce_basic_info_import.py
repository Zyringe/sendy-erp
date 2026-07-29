"""ecommerce-revamp Phase 2 — basic-info import path (description refresh).

Covers:
  1. Regression pin — the sheetViews crash is ALREADY fixed. `_read_shopee_file`
     did NOT need a new fix: parse_platform.py's module-level
     `Pane.__init__` monkeypatch (shipped in PR #117, well before this plan
     was written) already strips the non-standard `activePane="bottom_left"`
     value Shopee's exporter writes. Verified 2026-07-27 against every real
     `mass_update_basic_info_*.xlsx` file on record (06-19 through 07-25, 7
     files, 0 crashes) — see the poisoned-fixture test below, which builds a
     synthetic file reproducing the exact real-file XML shape rather than
     committing a binary fixture.
  2. parse_shopee_basic_info / parse_lazada_basic_info — new single-file
     item-grain parsers.
  3. import_platform_products COALESCE fix — a basic-info-only import must
     not wipe status/image_urls populated by an earlier fuller import.
  4. POST /ecommerce/import-info route.
"""
import io
import json
import os
import re
import zipfile

import pandas as pd
import pytest

os.environ.setdefault('SKIP_DB_INIT', '1')
os.environ.setdefault('SECRET_KEY', 'test-only-secret')
os.environ.setdefault('ADMIN_PASSWORD', 'test-only-admin')

import parse_platform as pp
import models

_REAL_SHOPEE_BASIC_INFO = os.path.expanduser(
    '~/Downloads/mass_update_basic_info_74562936_20260725150242.xlsx')
_REAL_LAZADA_BASIC = os.path.expanduser(
    '~/Downloads/basic100522265export1784972109068_0725-17-35-09.xlsx')


# ── Fixture builders ─────────────────────────────────────────────────────────

def _build_poisoned_shopee_basic_info_xlsx():
    """Tiny xlsx mimicking Shopee's mass_update_basic_info shape (2 metadata
    rows, Thai header row 2, 3 instruction rows, 1 data row), with freeze
    panes to produce a <pane> element, then binary-patched so its
    activePane carries the exact non-standard value ('bottom_left', an
    underscore variant) real Shopee exports write and openpyxl's Pane
    descriptor rejects — reproduces the real crash shape without a
    committed binary fixture."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['metadata row 0'])
    ws.append(['metadata row 1'])
    ws.append(['รหัสสินค้า', 'Parent SKU', 'ชื่อสินค้า', 'รายละเอียดสินค้า', 'เหตุผล'])
    ws.append(['instruction'])
    ws.append(['instruction'])
    ws.append(['instruction'])
    ws.append(['900000001', 'PSKU-1', 'ทดสอบสินค้า', 'รายละเอียดทดสอบ', ''])
    ws.freeze_panes = 'A7'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    with zipfile.ZipFile(buf, 'r') as zin:
        data = {n: zin.read(n) for n in zin.namelist()}
    sheet_name = next(n for n in data if n.startswith('xl/worksheets/sheet'))
    patched = re.sub(rb'activePane="[a-zA-Z]+"', b'activePane="bottom_left"',
                      data[sheet_name])
    assert patched != data[sheet_name], 'fixture did not get a pane to poison'
    data[sheet_name] = patched

    out = io.BytesIO()
    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zout:
        for n, content in data.items():
            zout.writestr(n, content)
    out.seek(0)
    return out


def _build_shopee_basic_info_xlsx(rows):
    """rows: list of (product_id_str, parent_sku, product_name, description)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['metadata row 0'])
    ws.append(['metadata row 1'])
    ws.append(['รหัสสินค้า', 'Parent SKU', 'ชื่อสินค้า', 'รายละเอียดสินค้า', 'เหตุผล'])
    ws.append(['instruction'])
    ws.append(['instruction'])
    ws.append(['instruction'])
    for r in rows:
        ws.append(list(r) + [''])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_lazada_basic_xlsx(rows):
    """rows: list of dicts keyed by the real Thai column names this export
    uses (verified against Put's real basic*.xlsx, 2026-07-27 — NOT the
    English headers parse_lazada_product_files' basic-file branch expects;
    Lazada's template evidently localized between when that was written and
    now)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'template'
    cols = ['Product ID', 'catId', 'ชื่อสินค้า', 'ชื่อสินค้าใน En', 'สถานะสินค้า',
            'รูปภาพสินค้า1', 'รูปภาพสินค้า2', 'เงื่อนไขการรับประกัน',
            'ระยะเวลาการรับประกัน', 'คำอธิบายหลัก', 'จุดเด่นของสินค้า']
    ws.append(cols)
    ws.append(['ไม่บังคับการกรอกข้อมูล'] + [''] * (len(cols) - 1))
    ws.append(['คำอธิบาย'] + [''] * (len(cols) - 1))
    ws.append(['*ตัวอย่าง'] + [''] * (len(cols) - 1))
    for r in rows:
        ws.append([r.get(c) for c in cols])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── 1. sheetViews regression pin ─────────────────────────────────────────────

def test_poisoned_sheetviews_shopee_file_does_not_crash():
    xlsx = _build_poisoned_shopee_basic_info_xlsx()
    df = pp._read_shopee_file(xlsx)
    assert len(df) == 1
    assert df.iloc[0]['รหัสสินค้า'] == '900000001'


@pytest.mark.skipif(not os.path.exists(_REAL_SHOPEE_BASIC_INFO),
                     reason="real Shopee basic_info file not present")
def test_real_shopee_basic_info_file_parses_288_rows_with_description():
    with open(_REAL_SHOPEE_BASIC_INFO, 'rb') as fh:
        records = pp.parse_shopee_basic_info(io.BytesIO(fh.read()))
    assert len(records) == 288, f"expected 288, got {len(records)}"
    assert all(r['product_id_str'] for r in records)
    lengths = sorted(len(r['description'] or '') for r in records)
    median = lengths[len(lengths) // 2]
    assert median >= 400, f"median description length {median} < 400"


@pytest.mark.skipif(not os.path.exists(_REAL_LAZADA_BASIC),
                     reason="real Lazada basic file not present")
def test_real_lazada_basic_file_parses_179_rows():
    with open(_REAL_LAZADA_BASIC, 'rb') as fh:
        records = pp.parse_lazada_basic_info(io.BytesIO(fh.read()))
    assert len(records) == 179, f"expected 179, got {len(records)}"
    assert all(r['product_id_str'] for r in records)


# ── 2. parse_shopee_basic_info ───────────────────────────────────────────────

def test_parse_shopee_basic_info_maps_fields():
    xlsx = _build_shopee_basic_info_xlsx([
        ('900000001', 'PSKU-1', 'สินค้าทดสอบ 1', 'รายละเอียดยาวๆ ของสินค้าทดสอบ 1'),
        ('900000002', 'PSKU-2', 'สินค้าทดสอบ 2', ''),
    ])
    records = pp.parse_shopee_basic_info(xlsx)
    assert len(records) == 2
    r0 = records[0]
    assert r0['product_id_str'] == '900000001'
    assert r0['parent_sku'] == 'PSKU-1'
    assert r0['product_name'] == 'สินค้าทดสอบ 1'
    assert r0['description'] == 'รายละเอียดยาวๆ ของสินค้าทดสอบ 1'
    raw = json.loads(r0['raw_json'])
    assert raw['รหัสสินค้า'] == '900000001'
    # image/status/warranty keys never present — basic_info doesn't carry them
    assert 'image_urls' not in r0
    assert 'status' not in r0


def test_parse_shopee_basic_info_drops_instruction_rows():
    xlsx = _build_shopee_basic_info_xlsx([('900000003', 'P3', 'ชื่อ', 'desc')])
    records = pp.parse_shopee_basic_info(xlsx)
    assert len(records) == 1  # the 3 instruction rows must not appear


# ── 3. parse_lazada_basic_info ───────────────────────────────────────────────

def test_parse_lazada_basic_info_maps_fields():
    xlsx = _build_lazada_basic_xlsx([{
        'Product ID': '421038989',
        'ชื่อสินค้า': 'SENDAI ลูกบิดประตู',
        'ชื่อสินค้าใน En': 'SENDAI Door Knob',
        'สถานะสินค้า': 'active',
        'รูปภาพสินค้า1': 'https://th-live.slatic.net/p/img1.jpg',
        'รูปภาพสินค้า2': 'https://th-live.slatic.net/p/img2.jpg',
        'เงื่อนไขการรับประกัน': 'มีการรับประกัน',
        'ระยะเวลาการรับประกัน': '1 ปี',
        'คำอธิบายหลัก': 'คำอธิบายสินค้าอย่างละเอียด',
    }])
    records = pp.parse_lazada_basic_info(xlsx)
    assert len(records) == 1
    r = records[0]
    assert r['product_id_str'] == '421038989'
    assert r['product_name'] == 'SENDAI ลูกบิดประตู'
    assert r['name_en'] == 'SENDAI Door Knob'
    assert r['status'] == 'active'
    assert r['warranty_policy'] == 'มีการรับประกัน'
    assert r['warranty_period'] == '1 ปี'
    assert r['description'] == 'คำอธิบายสินค้าอย่างละเอียด'
    assert r['cover_image_url'] == 'https://th-live.slatic.net/p/img1.jpg'
    assert json.loads(r['image_urls']) == [
        'https://th-live.slatic.net/p/img1.jpg',
        'https://th-live.slatic.net/p/img2.jpg',
    ]


def test_parse_lazada_basic_info_desc_falls_back_to_highlights():
    # The real basic*.xlsx carries คำอธิบายหลัก for only ~74/179 products —
    # the live content for the rest sits in จุดเด่นของสินค้า. Must mirror
    # parse_lazada_product_files' Main-or-Highlights fallback so a
    # basic-only re-import still carries edited Highlights content.
    xlsx = _build_lazada_basic_xlsx([
        {'Product ID': '421100001', 'ชื่อสินค้า': 'มีแต่จุดเด่น',
         'คำอธิบายหลัก': None, 'จุดเด่นของสินค้า': '<p>จุดเด่นเท่านั้น</p>'},
        {'Product ID': '421100002', 'ชื่อสินค้า': 'มีทั้งคู่',
         'คำอธิบายหลัก': 'คำอธิบายหลักชนะ', 'จุดเด่นของสินค้า': 'จุดเด่นแพ้'},
    ])
    r1, r2 = pp.parse_lazada_basic_info(xlsx)
    assert r1['description'] == '<p>จุดเด่นเท่านั้น</p>'
    assert r2['description'] == 'คำอธิบายหลักชนะ'


def test_parse_lazada_basic_info_empty_gallery_is_explicit_empty_list():
    xlsx = _build_lazada_basic_xlsx([{
        'Product ID': '421038990', 'ชื่อสินค้า': 'สินค้าไม่มีรูป',
        'ชื่อสินค้าใน En': None, 'สถานะสินค้า': 'inactive',
        'รูปภาพสินค้า1': None, 'รูปภาพสินค้า2': None,
        'เงื่อนไขการรับประกัน': None, 'ระยะเวลาการรับประกัน': None,
        'คำอธิบายหลัก': None,
    }])
    r = pp.parse_lazada_basic_info(xlsx)[0]
    assert r['cover_image_url'] is None
    assert r['image_urls'] == '[]'   # explicitly parsed empty, not absent


def test_parse_lazada_basic_info_drops_instruction_rows():
    xlsx = _build_lazada_basic_xlsx([{
        'Product ID': '421038991', 'ชื่อสินค้า': 'ชื่อ', 'ชื่อสินค้าใน En': None,
        'สถานะสินค้า': 'active', 'รูปภาพสินค้า1': None, 'รูปภาพสินค้า2': None,
        'เงื่อนไขการรับประกัน': None, 'ระยะเวลาการรับประกัน': None, 'คำอธิบายหลัก': None,
    }])
    records = pp.parse_lazada_basic_info(xlsx)
    assert len(records) == 1  # the 3 help rows must not appear


# ── 4. import_platform_products COALESCE fix ─────────────────────────────────

def test_basic_info_only_import_preserves_image_urls_and_status(tmp_db):
    models.import_platform_products('shopee', [{
        'product_id_str': 'TEST_9001',
        'product_name': 'ชื่อเดิม',
        'status': 'active',
        'cover_image_url': 'https://example.com/cover.jpg',
        'image_urls': '["https://example.com/cover.jpg","https://example.com/2.jpg"]',
        'raw_json': '{}',
    }])

    # Basic-info-only re-import: no status/image_urls keys at all (Shopee shape)
    models.import_platform_products('shopee', [{
        'product_id_str': 'TEST_9001',
        'parent_sku': 'PSKU-9001',
        'product_name': 'ชื่อเดิม',
        'description': 'รายละเอียดใหม่ที่เพิ่งอัปเดต',
        'raw_json': '{"source": "basic_info"}',
    }])

    import sqlite3
    conn = sqlite3.connect(tmp_db)
    row = conn.execute(
        "SELECT status, cover_image_url, image_urls, description FROM platform_products "
        "WHERE platform='shopee' AND product_id_str='TEST_9001'"
    ).fetchone()
    conn.close()

    assert row[0] == 'active', f"status wiped: {row[0]}"
    assert row[1] == 'https://example.com/cover.jpg', f"cover_image_url wiped: {row[1]}"
    assert row[2] == '["https://example.com/cover.jpg","https://example.com/2.jpg"]', \
        f"image_urls wiped: {row[2]}"
    assert row[3] == 'รายละเอียดใหม่ที่เพิ่งอัปเดต', "description not updated"


def test_explicit_empty_image_urls_does_overwrite(tmp_db):
    """The COALESCE guard protects an ABSENT key, not an explicitly-parsed
    empty list — a Lazada basic-info row with a genuinely empty gallery must
    still be able to overwrite a stale non-empty gallery."""
    models.import_platform_products('lazada', [{
        'product_id_str': 'TEST_9002',
        'product_name': 'ชื่อเดิม',
        'image_urls': '["https://example.com/stale.jpg"]',
        'raw_json': '{}',
    }])
    models.import_platform_products('lazada', [{
        'product_id_str': 'TEST_9002',
        'product_name': 'ชื่อเดิม',
        'image_urls': '[]',   # explicitly parsed empty, e.g. product now has no images
        'raw_json': '{}',
    }])

    import sqlite3
    conn = sqlite3.connect(tmp_db)
    row = conn.execute(
        "SELECT image_urls FROM platform_products "
        "WHERE platform='lazada' AND product_id_str='TEST_9002'"
    ).fetchone()
    conn.close()
    assert row[0] == '[]'


# ── 5. POST /ecommerce/import-info route ─────────────────────────────────────

@pytest.fixture
def admin_client(tmp_db):
    from app import app as flask_app
    flask_app.config['TESTING'] = True
    c = flask_app.test_client()
    with c.session_transaction() as sess:
        sess['user_id']  = 1
        sess['username'] = 'test-admin'
        sess['role']     = 'admin'
    return c


def test_import_info_route_shopee_file(admin_client, tmp_db):
    xlsx = _build_shopee_basic_info_xlsx([
        ('900000010', 'P10', 'ชื่อ', 'รายละเอียด'),
    ])
    resp = admin_client.post('/ecommerce/import-info', data={
        'info_files': (xlsx, 'mass_update_basic_info_74562936_20260101000000.xlsx'),
    }, content_type='multipart/form-data', follow_redirects=False)
    assert resp.status_code == 302

    import sqlite3
    conn = sqlite3.connect(tmp_db)
    row = conn.execute(
        "SELECT product_name, description FROM platform_products "
        "WHERE platform='shopee' AND product_id_str='900000010'"
    ).fetchone()
    conn.close()
    assert row == ('ชื่อ', 'รายละเอียด')


def test_import_info_route_lazada_file(admin_client, tmp_db):
    xlsx = _build_lazada_basic_xlsx([{
        'Product ID': '421000099', 'ชื่อสินค้า': 'สินค้า Lazada',
        'ชื่อสินค้าใน En': None, 'สถานะสินค้า': 'active',
        'รูปภาพสินค้า1': None, 'รูปภาพสินค้า2': None,
        'เงื่อนไขการรับประกัน': None, 'ระยะเวลาการรับประกัน': None,
        'คำอธิบายหลัก': 'คำอธิบาย Lazada',
    }])
    resp = admin_client.post('/ecommerce/import-info', data={
        'info_files': (xlsx, 'basic100000000export0000000000000_0101-00-00-00.xlsx'),
    }, content_type='multipart/form-data', follow_redirects=False)
    assert resp.status_code == 302

    import sqlite3
    conn = sqlite3.connect(tmp_db)
    row = conn.execute(
        "SELECT product_name, description FROM platform_products "
        "WHERE platform='lazada' AND product_id_str='421000099'"
    ).fetchone()
    conn.close()
    assert row == ('สินค้า Lazada', 'คำอธิบาย Lazada')


def test_import_info_route_unrecognized_filename_skipped(admin_client, tmp_db):
    xlsx = _build_shopee_basic_info_xlsx([('900000020', 'P20', 'ชื่อ', 'desc')])
    resp = admin_client.post('/ecommerce/import-info', data={
        'info_files': (xlsx, 'random_export.xlsx'),
    }, content_type='multipart/form-data', follow_redirects=True)
    assert resp.status_code == 200
    assert 'ไม่ตรงรูปแบบ'.encode() in resp.data

    import sqlite3
    conn = sqlite3.connect(tmp_db)
    row = conn.execute(
        "SELECT 1 FROM platform_products WHERE product_id_str='900000020'"
    ).fetchone()
    conn.close()
    assert row is None


def test_import_info_route_advances_max_imported_at(admin_client, tmp_db):
    import sqlite3
    conn = sqlite3.connect(tmp_db)
    before = conn.execute(
        "SELECT MAX(imported_at) FROM platform_products WHERE platform='shopee'"
    ).fetchone()[0]
    conn.close()

    xlsx = _build_shopee_basic_info_xlsx([('900000030', 'P30', 'ชื่อ', 'desc')])
    admin_client.post('/ecommerce/import-info', data={
        'info_files': (xlsx, 'mass_update_basic_info_74562936_20260101000001.xlsx'),
    }, content_type='multipart/form-data')

    conn = sqlite3.connect(tmp_db)
    after = conn.execute(
        "SELECT MAX(imported_at) FROM platform_products WHERE platform='shopee'"
    ).fetchone()[0]
    conn.close()
    assert (before or '') <= after
    assert after is not None


def test_import_info_route_no_no_files_flashes_and_redirects(admin_client):
    resp = admin_client.post('/ecommerce/import-info', data={},
                              content_type='multipart/form-data',
                              follow_redirects=True)
    assert resp.status_code == 200
    assert 'กรุณาเลือกไฟล์'.encode() in resp.data
