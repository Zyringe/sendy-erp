"""
tests/test_hr_sync_no_clobber.py
TDD tests for the HR-sync no-clobber rule in inventory_app/import_cashbook.py.

The rule: sync_salary_sheet() must never modify any field on an EXISTING
employee. New-employee creation is unaffected. Sheet/DB mismatches are
surfaced as DIFF warnings; bank_account_no values are masked (PII).

See docs/superpowers/specs/2026-05-21-hr-sync-no-clobber-design.md for
the design.
"""
import sqlite3

import pytest

import import_cashbook as ic


def _seed_companies(conn):
    """Idempotently seed BSN (id=1) + SD (id=2) so employees.company_id FK satisfied."""
    conn.execute(
        "INSERT OR IGNORE INTO companies (id, code, name_th) VALUES (1, 'BSN', 'บุญสวัสดิ์ นำชัย')"
    )
    conn.execute(
        "INSERT OR IGNORE INTO companies (id, code, name_th) VALUES (2, 'SD', 'เซ็นไดเทรดดิ้ง')"
    )
    conn.commit()


def _seed_employee(conn, *, full_name, nickname=None, bank_name=None,
                   bank_account_no=None, emp_code="EMP001"):
    """Insert one row into employees with the schema columns the sync touches."""
    _seed_companies(conn)
    conn.execute(
        """INSERT INTO employees
             (emp_code, full_name, nickname, bank_name, bank_account_no,
              company_id, sso_enrolled, diligence_allowance, is_active,
              start_date, probation_days)
           VALUES (?, ?, ?, ?, ?, 1, 0, 0, 1, NULL, 90)""",
        (emp_code, full_name, nickname, bank_name, bank_account_no),
    )
    conn.commit()


def _parsed_salary(rows):
    """Build the minimum `parsed` dict shape sync_salary_sheet consumes."""
    return {"salary": rows, "advances": [], "transactions": [],
            "categories": {"income": [], "expense": []}, "overview": {}}


# ── Test 2 (กิติยา case) — first failing test ────────────────────────────────

def test_existing_employee_with_null_nickname_not_refilled(empty_db_conn):
    """
    Seed an employee with nickname=NULL (Put manually cleared it). The salary
    sheet has a nickname for the same person. After sync, the DB must remain
    NULL — and a DIFF warning must surface the mismatch.
    """
    conn = empty_db_conn
    _seed_employee(conn, full_name="กิติยา ทดสอบ", nickname=None,
                   emp_code="EMP001")

    parsed = _parsed_salary([{
        "first_name": "กิติยา",
        "last_name":  "ทดสอบ",
        "nickname":   "กี",
        "bank":       None,
        "bank_account_no": None,
        "salary":     12000.0,
        "sso_deduction": 600.0,
        "is_active":  True,
    }])

    result = ic.sync_salary_sheet(parsed, conn)

    db_nickname = conn.execute(
        "SELECT nickname FROM employees WHERE emp_code='EMP001'"
    ).fetchone()[0]
    assert db_nickname is None, \
        f"DB nickname must remain NULL, got {db_nickname!r}"

    assert "EMP001" in result["skipped"]
    assert "EMP001" not in result["updated"]

    diff_warnings = [w for w in result["warnings"] if "DIFF" in w]
    assert len(diff_warnings) >= 1, \
        f"Expected at least one DIFF warning, got: {result['warnings']!r}"
    msg = diff_warnings[0]
    assert "EMP001" in msg
    assert "nickname" in msg


# ── Test 1 — non-NULL custom value (วิภา=หลุย case) ──────────────────────────

def test_existing_employee_with_non_null_nickname_not_modified(empty_db_conn):
    """
    Seed nickname='หลุย' (Put's custom value, different from sheet's 'วิภา').
    DB must remain 'หลุย'; a DIFF warning must surface both values.
    """
    conn = empty_db_conn
    _seed_employee(conn, full_name="วิภา ทดสอบ", nickname="หลุย",
                   emp_code="EMP002")

    parsed = _parsed_salary([{
        "first_name": "วิภา", "last_name": "ทดสอบ",
        "nickname":   "วิภา",
        "bank":       None, "bank_account_no": None,
        "salary": 12000.0, "sso_deduction": 600.0, "is_active": True,
    }])

    result = ic.sync_salary_sheet(parsed, conn)

    db_nickname = conn.execute(
        "SELECT nickname FROM employees WHERE emp_code='EMP002'"
    ).fetchone()[0]
    assert db_nickname == "หลุย", \
        f"DB nickname must remain 'หลุย', got {db_nickname!r}"

    assert "EMP002" in result["skipped"]
    diffs = [w for w in result["warnings"] if "DIFF" in w]
    assert len(diffs) == 1, f"Expected exactly 1 DIFF warning, got: {diffs!r}"
    msg = diffs[0]
    assert "EMP002" in msg
    assert "nickname" in msg
    assert "หลุย" in msg
    assert "วิภา" in msg


# ── Test 3 — sheet matches DB → no warnings, silent skip ─────────────────────

def test_existing_employee_matching_sheet_silent_skip(empty_db_conn):
    """
    Seed all 3 lockable fields identically to the sheet row. sync must emit
    zero DIFF warnings and the emp_code must land in 'skipped'.
    """
    conn = empty_db_conn
    _seed_employee(conn, full_name="สมศักดิ์ ทดสอบ",
                   nickname="ศักดิ์", bank_name="กสิกร",
                   bank_account_no="1234567890", emp_code="EMP003")

    parsed = _parsed_salary([{
        "first_name": "สมศักดิ์", "last_name": "ทดสอบ",
        "nickname":   "ศักดิ์",
        "bank":       "กสิกร",
        "bank_account_no": "1234567890",
        "salary": 15000.0, "sso_deduction": 750.0, "is_active": True,
    }])

    result = ic.sync_salary_sheet(parsed, conn)

    diff_warnings = [w for w in result["warnings"] if "DIFF" in w]
    assert diff_warnings == [], \
        f"Expected zero DIFF warnings, got: {diff_warnings!r}"
    assert "EMP003" in result["skipped"]


# ── Test 6 — bank_account_no values must NEVER appear in warning text ────────

def test_bank_account_no_diff_does_not_leak_raw_values(empty_db_conn):
    """
    Seed employee with bank_account_no='1234567890' AND matching nickname +
    bank_name (so they emit no warning). Sheet has a different account number
    ('9999999999'). The only DIFF warning must be the bank_account_no one,
    and neither '1234567890' nor '9999999999' may appear in any warning.
    """
    conn = empty_db_conn
    _seed_employee(conn, full_name="สมหญิง ทดสอบ",
                   nickname="หญิง", bank_name="ไทยพาณิชย์",
                   bank_account_no="1234567890", emp_code="EMP004")

    parsed = _parsed_salary([{
        "first_name": "สมหญิง", "last_name": "ทดสอบ",
        "nickname":   "หญิง",                       # matches DB
        "bank":       "ไทยพาณิชย์",                 # matches DB
        "bank_account_no": "9999999999",            # diverges
        "salary": 14000.0, "sso_deduction": 700.0, "is_active": True,
    }])

    result = ic.sync_salary_sheet(parsed, conn)

    diff_warnings = [w for w in result["warnings"] if "DIFF" in w]
    assert len(diff_warnings) == 1, \
        f"Expected exactly 1 DIFF warning, got: {diff_warnings!r}"
    msg = diff_warnings[0]
    assert "bank_account_no" in msg
    assert "EMP004" in msg

    # PII regression — neither account number may appear in any warning text.
    for w in result["warnings"]:
        assert "1234567890" not in w, f"DB account leaked into warning: {w!r}"
        assert "9999999999" not in w, f"Sheet account leaked into warning: {w!r}"


# ── Test 4 — new employee still auto-created from sheet ──────────────────────

def test_new_employee_still_auto_created(empty_db_conn):
    """
    Sheet has a person not in DB. The new-employee branch must still run:
    employees row inserted with auto EMP-code, employee_salary_history seed
    row inserted, result['created'] populated, "start_date unknown" warning
    emitted.
    """
    conn = empty_db_conn
    _seed_companies(conn)  # required for FK; no existing employee seeded

    parsed = _parsed_salary([{
        "first_name": "ใหม่", "last_name": "มาเอง",
        "nickname":   "ใหม่",
        "bank":       "กสิกร",
        "bank_account_no": "5551112220",
        "salary": 13000.0, "sso_deduction": 650.0, "is_active": True,
    }])

    result = ic.sync_salary_sheet(parsed, conn)

    row = conn.execute(
        "SELECT emp_code, nickname, bank_name, bank_account_no "
        "FROM employees WHERE full_name='ใหม่ มาเอง'"
    ).fetchone()
    assert row is not None, "New employee must be inserted"
    emp_code = row[0]
    assert emp_code.startswith("EMP")
    assert row[1] == "ใหม่"
    assert row[2] == "กสิกร"
    assert row[3] == "5551112220"

    assert emp_code in result["created"]
    assert emp_code not in result["skipped"]
    assert emp_code not in result["updated"]

    hist = conn.execute(
        "SELECT effective_date, reason FROM employee_salary_history "
        "WHERE employee_id=(SELECT id FROM employees WHERE emp_code=?)",
        (emp_code,),
    ).fetchone()
    assert hist is not None, "Salary history seed row missing for new employee"
    assert hist[1] == "initial"

    start_warnings = [w for w in result["warnings"] if "start_date" in w]
    assert len(start_warnings) >= 1, \
        "Expected 'start_date unknown' warning for new employee"


# ── Test 5 — end-to-end idempotence after Put clears a field in HR UI ────────

def test_re_import_idempotent_after_manual_clear(empty_db_conn):
    """
    Simulate Put's workflow:
      1. First import creates EMP with nickname filled.
      2. Put manually clears the nickname via HR UI (modelled here as a
         direct UPDATE … SET nickname=NULL).
      3. Second import on the same sheet row must NOT refill the nickname.
    """
    conn = empty_db_conn
    _seed_companies(conn)  # required for FK; no existing employee seeded

    parsed = _parsed_salary([{
        "first_name": "ทดสอบ", "last_name": "อิดิ",
        "nickname":   "ทด",
        "bank":       None, "bank_account_no": None,
        "salary": 10000.0, "sso_deduction": 500.0, "is_active": True,
    }])

    # Pass 1 — creates the employee.
    result1 = ic.sync_salary_sheet(parsed, conn)
    emp_code = result1["created"][0]
    row = conn.execute(
        "SELECT nickname FROM employees WHERE emp_code=?", (emp_code,)
    ).fetchone()
    assert row[0] == "ทด", "Initial sync must populate nickname for new employee"

    # Simulate Put clearing via HR UI.
    conn.execute(
        "UPDATE employees SET nickname=NULL WHERE emp_code=?", (emp_code,)
    )
    conn.commit()

    # Pass 2 — must NOT refill.
    result2 = ic.sync_salary_sheet(parsed, conn)
    row = conn.execute(
        "SELECT nickname FROM employees WHERE emp_code=?", (emp_code,)
    ).fetchone()
    assert row[0] is None, \
        f"DB nickname must stay NULL on re-import, got {row[0]!r}"

    assert emp_code in result2["skipped"]
    assert emp_code not in result2["updated"]
    diffs = [w for w in result2["warnings"] if "DIFF" in w and "nickname" in w]
    assert len(diffs) == 1, \
        f"Expected exactly 1 nickname DIFF warning on pass 2, got: {diffs!r}"


# ── Test 7 — advance unmatched when existing employee has NULL nickname ──────
import datetime as _dt
import openpyxl


def _build_minimal_advance_test_wb(path, *, full_first, full_last,
                                    sheet_nickname, advance_raw_name,
                                    advance_amount=500.0):
    """
    Build the smallest workbook import_cashbook can process: 1 account with 1
    transaction, Salary_Sheet with 1 employee, 1 advance row, minimal Setup +
    Overview.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_ov = wb.create_sheet("Overview")
    ws_ov.cell(row=3, column=2, value="รายรับ");  ws_ov.cell(row=3, column=3, value=1000.0)
    ws_ov.cell(row=4, column=2, value="รายจ่าย"); ws_ov.cell(row=4, column=3, value=500.0)
    ws_ov.cell(row=5, column=2, value="คงเหลือ"); ws_ov.cell(row=5, column=3, value=500.0)

    ws = wb.create_sheet("Txn_MAIN")
    for ci, h in enumerate(
        ["วันที่", "ประเภท", "หมวดหมู่", "หมวดหมู่_ผู้ใช้",
         "จำนวนเงิน", "รายละเอียด", "หมายเหตุ"], 1):
        ws.cell(row=1, column=ci, value=h)
    ws.cell(row=2, column=9, value="Bank");           ws.cell(row=2, column=10, value="KBank")
    ws.cell(row=3, column=9, value="Account Number"); ws.cell(row=3, column=10, value="0000000000")
    ws.cell(row=4, column=9, value="Name");           ws.cell(row=4, column=10, value="เจ้าของ")
    ws.cell(row=2, column=1, value=_dt.datetime(2026, 4, 1))
    ws.cell(row=2, column=2, value="รายรับ")
    ws.cell(row=2, column=3, value="เงินฝาก")
    ws.cell(row=2, column=5, value=1000.0)
    ws.cell(row=3, column=1, value=_dt.datetime(2026, 4, 2))
    ws.cell(row=3, column=2, value="รายจ่าย")
    ws.cell(row=3, column=3, value="เงินเดือน")
    ws.cell(row=3, column=5, value=500.0)

    ws_sal = wb.create_sheet("Salary_Sheet")
    ws_sal.cell(row=1, column=1, value="des")
    for ci, h in enumerate(
        ["", "ชื่อ", "นามสกุล", "ชื่อเล่น", "ธนาคาร", "เลขบัญชี",
         "เงินเดือน", "หักประกันสังคม", "เงินเดือนสุทธิ", "is_active"], 1):
        ws_sal.cell(row=2, column=ci, value=h)
    ws_sal.cell(row=3, column=2, value=full_first)
    ws_sal.cell(row=3, column=3, value=full_last)
    ws_sal.cell(row=3, column=4, value=sheet_nickname)
    ws_sal.cell(row=3, column=5, value=None)
    ws_sal.cell(row=3, column=6, value=None)
    ws_sal.cell(row=3, column=7, value=10000.0)
    ws_sal.cell(row=3, column=8, value=500.0)
    ws_sal.cell(row=3, column=9, value=9500.0)
    ws_sal.cell(row=3, column=10, value=True)

    ws_adv = wb.create_sheet("เบิกเงินล่วงหน้า")
    ws_adv.cell(row=2, column=2, value="วันที่")
    ws_adv.cell(row=2, column=3, value="ชื่อ")
    ws_adv.cell(row=2, column=4, value="เบิกเงินล่วงหน้า")
    ws_adv.cell(row=2, column=5, value="หมายเหตุ")
    ws_adv.cell(row=3, column=2, value=_dt.datetime(2026, 4, 10))
    ws_adv.cell(row=3, column=3, value=advance_raw_name)
    ws_adv.cell(row=3, column=4, value=advance_amount)
    ws_adv.cell(row=3, column=5, value=None)

    ws_setup = wb.create_sheet("Setup")
    ws_setup.cell(row=2, column=2, value="รายรับ")
    ws_setup.cell(row=2, column=3, value="รายจ่าย")
    ws_setup.cell(row=3, column=2, value="เงินฝาก")
    ws_setup.cell(row=3, column=3, value="เงินเดือน")

    wb.save(path)


def test_advance_unmatched_when_existing_nickname_null(empty_db_conn, tmp_path):
    """
    Existing employee has full_name='ทดสอบ อิดิ' with nickname=NULL.
    Sheet brings nickname='ทด' AND a salary advance with raw_name='ทด'.

    Expected after import:
      - employee row UNCHANGED (DB nickname still NULL)
      - advance row inserted, but employee_id=NULL (raw_name preserved)
      - DIFF warning for the nickname mismatch emitted
    """
    conn = empty_db_conn
    # Seed employee with NULL nickname.
    _seed_employee(conn, full_name="ทดสอบ อิดิ", nickname=None,
                   emp_code="EMP010")

    wb_path = str(tmp_path / "advance_unmatched.xlsx")
    _build_minimal_advance_test_wb(
        wb_path,
        full_first="ทดสอบ", full_last="อิดิ",
        sheet_nickname="ทด",
        advance_raw_name="ทด",
    )

    ic.import_cashbook(wb_path, conn=conn)

    # 1. Employee row unchanged.
    db_nickname = conn.execute(
        "SELECT nickname FROM employees WHERE emp_code='EMP010'"
    ).fetchone()[0]
    assert db_nickname is None, \
        f"DB nickname must remain NULL after import, got {db_nickname!r}"

    # 2. Advance row inserted, employee_id NULL, raw_name preserved.
    adv_rows = conn.execute(
        "SELECT employee_id, raw_name FROM salary_advances WHERE raw_name='ทด'"
    ).fetchall()
    assert len(adv_rows) == 1, \
        f"Expected exactly 1 advance row with raw_name='ทด', got {adv_rows!r}"
    assert adv_rows[0][0] is None, \
        "Advance must be unmatched (employee_id=NULL) until nickname is set"
    assert adv_rows[0][1] == "ทด"
