"""
tests/test_cashbook_parse.py
TDD tests for inventory_app/parse_cashbook.py

Run RED first (no parser yet), then implement parser to GREEN.

Synthetic fixture: built programmatically with openpyxl in a tmp dir.
Real-file integration test: skipped if /Users/putty/Downloads/NoVat_Account.xlsx absent.
"""
import datetime
import os
from typing import Optional

import pytest

# ── Where is the real file? ────────────────────────────────────────────────────
REAL_FILE = "/Users/putty/Downloads/NoVat_Account.xlsx"

# ── Fixture builder ────────────────────────────────────────────────────────────

def _build_synthetic_wb(path: str) -> None:
    """
    Build a synthetic cashbook xlsx at `path` with just enough structure to
    exercise every code path in parse_cashbook:

    Sheets:
      Overview           — income/expense/balance block
      Txn_392            — numeric code, full IJ sidecar (account no as float)
      Txn_ชาร           — Thai-name code, minimal IJ
      Note >>            — separator (must be ignored)
      Salary_Sheet       — two employees (one active, one inactive)
      เบิกเงินล่วงหน้า  — two advance rows
      Setup              — income + expense category lists
    """
    import openpyxl

    wb = openpyxl.Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # ── Overview ──────────────────────────────────────────────────────────────
    ws_ov = wb.create_sheet("Overview")
    # rows 3-5: label in col B (idx 2), value in col C (idx 3)
    ws_ov.cell(row=3, column=2, value="รายรับ")
    ws_ov.cell(row=3, column=3, value=75000.50)
    ws_ov.cell(row=4, column=2, value="รายจ่าย")
    ws_ov.cell(row=4, column=3, value=30000.00)
    ws_ov.cell(row=5, column=2, value="คงเหลือ")
    ws_ov.cell(row=5, column=3, value=45000.50)

    # ── Txn_392 ───────────────────────────────────────────────────────────────
    ws392 = wb.create_sheet("Txn_392")
    # Header row 1: cols A-G
    headers = ["วันที่", "ประเภท", "หมวดหมู่", "หมวดหมู่_ผู้ใช้", "จำนวนเงิน", "รายละเอียด", "หมายเหตุ"]
    for ci, h in enumerate(headers, 1):
        ws392.cell(row=1, column=ci, value=h)

    # IJ sidecar (col H=blank, I=label, J=value)
    ws392.cell(row=2, column=9, value="Bank");       ws392.cell(row=2, column=10, value="SCB")
    ws392.cell(row=3, column=9, value="Account Number"); ws392.cell(row=3, column=10, value=2322842392.0)
    ws392.cell(row=4, column=9, value="Name");       ws392.cell(row=4, column=10, value="วฤทธิ์ Test")
    ws392.cell(row=5, column=9, value="หมายเหตุ");  ws392.cell(row=5, column=10, value="note text")
    # IJ computed totals block — must be ignored by parser
    ws392.cell(row=8, column=9, value="รายรับ");    ws392.cell(row=8, column=10, value=75000.50)
    ws392.cell(row=9, column=9, value="รายจ่าย");   ws392.cell(row=9, column=10, value=30000.00)
    ws392.cell(row=10, column=9, value="คงเหลือ");  ws392.cell(row=10, column=10, value=45000.50)

    # Row 2: valid income row — amount as plain float
    ws392.cell(row=2, column=1, value=datetime.datetime(2026, 3, 7))
    ws392.cell(row=2, column=2, value="รายรับ")
    ws392.cell(row=2, column=3, value="เงินฝาก")
    ws392.cell(row=2, column=4, value=None)           # user_category blank
    ws392.cell(row=2, column=5, value=50000.00)
    ws392.cell(row=2, column=6, value="เปิดบัญชี")
    ws392.cell(row=2, column=7, value=None)

    # Row 3: valid expense row — amount as string with thousands comma
    ws392.cell(row=3, column=1, value=datetime.datetime(2026, 3, 9))
    ws392.cell(row=3, column=2, value="รายจ่าย")
    ws392.cell(row=3, column=3, value="ซื้อสินค้า")
    ws392.cell(row=3, column=4, value="user1")
    ws392.cell(row=3, column=5, value="25,000.00")    # comma-formatted string
    ws392.cell(row=3, column=6, value="ซื้อพุก")
    ws392.cell(row=3, column=7, value="note1")

    # Row 4: valid income row — description with embedded newline
    ws392.cell(row=4, column=1, value=datetime.datetime(2026, 3, 12))
    ws392.cell(row=4, column=2, value="รายรับ")
    ws392.cell(row=4, column=3, value="ยอดขาย")
    ws392.cell(row=4, column=4, value=None)
    ws392.cell(row=4, column=5, value=25000.50)
    ws392.cell(row=4, column=6, value="ลูกค้า A\nลูกค้า B")   # embedded newline
    ws392.cell(row=4, column=7, value=None)

    # Row 5: blank separator row — must be skipped
    # (all cells None — no writes needed)

    # Row 6: has date but blank ประเภท — must be skipped
    ws392.cell(row=6, column=1, value=datetime.datetime(2026, 3, 15))
    ws392.cell(row=6, column=2, value=None)            # ประเภท blank → skip
    ws392.cell(row=6, column=5, value=100.0)

    # Row 7: has date but ประเภท is not รายรับ/รายจ่าย — must be skipped
    ws392.cell(row=7, column=1, value=datetime.datetime(2026, 3, 16))
    ws392.cell(row=7, column=2, value="ยอดคงเหลือ")   # not รายรับ/รายจ่าย → skip
    ws392.cell(row=7, column=5, value=1000.0)

    # ── Txn_ชาร ───────────────────────────────────────────────────────────────
    ws_th = wb.create_sheet("Txn_ชาร")
    for ci, h in enumerate(headers, 1):
        ws_th.cell(row=1, column=ci, value=h)
    # Minimal IJ (all None) — still creates account entry
    ws_th.cell(row=2, column=9, value="Bank");       ws_th.cell(row=2, column=10, value=None)
    ws_th.cell(row=3, column=9, value="Account Number"); ws_th.cell(row=3, column=10, value=None)
    ws_th.cell(row=4, column=9, value="Name");       ws_th.cell(row=4, column=10, value=None)
    ws_th.cell(row=5, column=9, value="หมายเหตุ");  ws_th.cell(row=5, column=10, value=None)
    # One valid expense row
    ws_th.cell(row=2, column=1, value=datetime.datetime(2026, 4, 1))
    ws_th.cell(row=2, column=2, value="รายจ่าย")
    ws_th.cell(row=2, column=3, value="เงินเดือน")
    ws_th.cell(row=2, column=4, value=None)
    ws_th.cell(row=2, column=5, value=5000.00)
    ws_th.cell(row=2, column=6, value=None)
    ws_th.cell(row=2, column=7, value=None)

    # ── Note >> separator ─────────────────────────────────────────────────────
    ws_note = wb.create_sheet("Note >>")
    ws_note.cell(row=1, column=1, value=None)   # empty

    # ── Salary_Sheet ──────────────────────────────────────────────────────────
    ws_sal = wb.create_sheet("Salary_Sheet")
    ws_sal.cell(row=1, column=1, value="des")
    # Header row 2: B-J (cols 2-10)
    sal_hdr = ["", "ชื่อ", "นามสกุล", "ชื่อเล่น", "ธนาคาร", "เลขบัญชี",
               "เงินเดือน", "หักประกันสังคม", "เงินเดือนสุทธิ", "is_active"]
    for ci, h in enumerate(sal_hdr, 1):
        ws_sal.cell(row=2, column=ci, value=h)
    # Active employee (row 3)
    ws_sal.cell(row=3, column=2, value="สมชาย")
    ws_sal.cell(row=3, column=3, value="ใจดี")
    ws_sal.cell(row=3, column=4, value="โอ")
    ws_sal.cell(row=3, column=5, value="กสิกร")
    ws_sal.cell(row=3, column=6, value=1234567890.0)   # float account no
    ws_sal.cell(row=3, column=7, value=15000.0)
    ws_sal.cell(row=3, column=8, value=750.0)
    ws_sal.cell(row=3, column=9, value=14250.0)
    ws_sal.cell(row=3, column=10, value=True)
    # Inactive employee (row 4)
    ws_sal.cell(row=4, column=2, value="มานี")
    ws_sal.cell(row=4, column=3, value="มีผล")
    ws_sal.cell(row=4, column=4, value=None)          # no nickname
    ws_sal.cell(row=4, column=5, value="ไทยพาณิชย์")
    ws_sal.cell(row=4, column=6, value=9876543210.0)
    ws_sal.cell(row=4, column=7, value=12000.0)
    ws_sal.cell(row=4, column=8, value=0.0)
    ws_sal.cell(row=4, column=9, value=12000.0)
    ws_sal.cell(row=4, column=10, value=False)
    # Row 5 empty — must be skipped

    # ── เบิกเงินล่วงหน้า ──────────────────────────────────────────────────────
    ws_adv = wb.create_sheet("เบิกเงินล่วงหน้า")
    # Row 2: header
    ws_adv.cell(row=2, column=2, value="วันที่")
    ws_adv.cell(row=2, column=3, value="ชื่อ")
    ws_adv.cell(row=2, column=4, value="เบิกเงินล่วงหน้า")
    ws_adv.cell(row=2, column=5, value="หมายเหตุ")
    # Row 3: บอล advance
    ws_adv.cell(row=3, column=2, value=datetime.datetime(2026, 4, 20))
    ws_adv.cell(row=3, column=3, value="บอล")
    ws_adv.cell(row=3, column=4, value=1300.0)
    ws_adv.cell(row=3, column=5, value="3 วัน")
    # Row 4: ริน advance — no note
    ws_adv.cell(row=4, column=2, value=datetime.datetime(2026, 4, 23))
    ws_adv.cell(row=4, column=3, value="ริน")
    ws_adv.cell(row=4, column=4, value=400.0)
    ws_adv.cell(row=4, column=5, value=None)
    # Row 5: no date — must be skipped
    ws_adv.cell(row=5, column=3, value="หลุย")
    ws_adv.cell(row=5, column=4, value=999.0)

    # ── Setup ─────────────────────────────────────────────────────────────────
    ws_setup = wb.create_sheet("Setup")
    # Row 2: header
    ws_setup.cell(row=2, column=2, value="รายรับ")
    ws_setup.cell(row=2, column=3, value="รายจ่าย")
    ws_setup.cell(row=2, column=5, value="ผู้ใช้")
    ws_setup.cell(row=2, column=6, value="ผู้ใช้ (คน)")
    # Income categories: col B rows 3-5
    for i, cat in enumerate(["เงินฝาก", "ดอกเบี้ย", "รายรับอื่นๆ"], 3):
        ws_setup.cell(row=i, column=2, value=cat)
    # Expense categories: col C rows 3-7
    for i, cat in enumerate(["ซื้อสินค้า", "เงินเดือน", "ค่าน้ำมัน", "อื่นๆ"], 3):
        ws_setup.cell(row=i, column=3, value=cat)
    # Users col E/F
    ws_setup.cell(row=3, column=5, value="บัญชี A")
    ws_setup.cell(row=3, column=6, value="คนหนึ่ง")

    wb.save(path)


@pytest.fixture
def synthetic_wb(tmp_path):
    """Build the synthetic workbook, return its path."""
    p = str(tmp_path / "test_cashbook.xlsx")
    _build_synthetic_wb(p)
    return p


# ── Helper: import parse_cashbook (will fail until implemented) ────────────────

def _import_parser():
    from inventory_app import parse_cashbook  # noqa: F401
    return parse_cashbook


# ══════════════════════════════════════════════════════════════════════════════
# Unit tests — synthetic fixture
# ══════════════════════════════════════════════════════════════════════════════

class TestParserImportable:
    """Parser module must exist and expose parse_cashbook()."""

    def test_module_imports(self):
        pc = _import_parser()
        assert hasattr(pc, "parse_cashbook"), "parse_cashbook function must exist"

    def test_returns_expected_top_level_keys(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        for key in ("accounts", "transactions", "salary", "advances", "categories", "overview", "warnings"):
            assert key in result, f"Missing top-level key: {key!r}"


class TestAccountSheetIdentification:
    """Only Txn_* sheets should become accounts; separators/others ignored."""

    def test_account_count(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        assert len(result["accounts"]) == 2, (
            f"Expected 2 accounts (Txn_392, Txn_ชาร), got {len(result['accounts'])}"
        )

    def test_account_codes(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        codes = {a["code"] for a in result["accounts"]}
        assert codes == {"392", "ชาร"}, f"Unexpected account codes: {codes}"

    def test_separator_sheet_not_in_accounts(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        codes = {a["code"] for a in result["accounts"]}
        assert "Note >>" not in codes
        assert "Overview" not in codes
        assert "Salary_Sheet" not in codes

    def test_no_txn_from_separator_sheet(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        sources = {t["source_sheet"] for t in result["transactions"]}
        assert "Note >>" not in sources
        assert "Overview" not in sources


class TestAccountMetadata:
    """IJ sidecar extraction."""

    def _get_account(self, result, code: str) -> dict:
        for a in result["accounts"]:
            if a["code"] == code:
                return a
        raise KeyError(f"No account with code {code!r}")

    def test_bank_name(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        acct = self._get_account(result, "392")
        assert acct["bank_name"] == "SCB"

    def test_account_no_stripped_of_float_zero(self, synthetic_wb):
        """2322842392.0 → '2322842392' (no trailing .0)"""
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        acct = self._get_account(result, "392")
        assert acct["bank_account_no"] == "2322842392", (
            f"Expected '2322842392', got {acct['bank_account_no']!r}"
        )

    def test_account_owner_name(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        acct = self._get_account(result, "392")
        assert acct["account_owner_name"] == "วฤทธิ์ Test"

    def test_account_note(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        acct = self._get_account(result, "392")
        assert acct["note"] == "note text"

    def test_all_none_meta_for_minimal_account(self, synthetic_wb):
        """Txn_ชาร has all-None IJ values."""
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        acct = self._get_account(result, "ชาร")
        assert acct["bank_name"] is None
        assert acct["bank_account_no"] is None
        assert acct["account_owner_name"] is None


class TestTransactionParsing:
    """Core row-parsing logic."""

    def _txns_for(self, result, sheet: str) -> list:
        return [t for t in result["transactions"] if t["source_sheet"] == sheet]

    def test_valid_row_count_txn392(self, synthetic_wb):
        """Txn_392 has 3 valid rows (rows 2,3,4); rows 5,6,7 must be skipped."""
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        txns = self._txns_for(result, "Txn_392")
        assert len(txns) == 3, f"Expected 3 txns from Txn_392, got {len(txns)}"

    def test_income_direction_mapping(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        txns = self._txns_for(result, "Txn_392")
        income_rows = [t for t in txns if t["direction"] == "income"]
        assert len(income_rows) == 2, f"Expected 2 income rows, got {len(income_rows)}"

    def test_expense_direction_mapping(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        txns = self._txns_for(result, "Txn_392")
        expense_rows = [t for t in txns if t["direction"] == "expense"]
        assert len(expense_rows) == 1, f"Expected 1 expense row, got {len(expense_rows)}"

    def test_iso_date_conversion(self, synthetic_wb):
        """datetime(2026,3,7) → '2026-03-07'"""
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        txns = self._txns_for(result, "Txn_392")
        dates = {t["txn_date"] for t in txns}
        assert "2026-03-07" in dates, f"Expected '2026-03-07' in {dates}"

    def test_comma_amount_parsed_to_float(self, synthetic_wb):
        """'25,000.00' string in amount cell → float 25000.0"""
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        txns = self._txns_for(result, "Txn_392")
        expense = [t for t in txns if t["direction"] == "expense"][0]
        assert expense["amount"] == pytest.approx(25000.0), (
            f"Expected 25000.0, got {expense['amount']!r}"
        )

    def test_amount_is_positive(self, synthetic_wb):
        """Amount must be positive regardless of direction."""
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        for t in result["transactions"]:
            assert t["amount"] >= 0, f"Negative amount in txn: {t}"

    def test_newline_preserved_in_description(self, synthetic_wb):
        """Description 'ลูกค้า A\\nลูกค้า B' must keep the embedded newline."""
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        txns = self._txns_for(result, "Txn_392")
        multi = [t for t in txns if t.get("description") and "\n" in t["description"]]
        assert len(multi) == 1, "Expected exactly 1 txn with embedded newline in description"
        assert multi[0]["description"] == "ลูกค้า A\nลูกค้า B"

    def test_blank_ประเภท_row_skipped(self, synthetic_wb):
        """Row with date but ประเภท=None must be skipped (not in transactions)."""
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        txns = self._txns_for(result, "Txn_392")
        # Row 6 has date=2026-03-15 with blank ประเภท
        dates = [t["txn_date"] for t in txns]
        assert "2026-03-15" not in dates, "Row with blank ประเภท must be skipped"

    def test_invalid_ประเภท_row_skipped(self, synthetic_wb):
        """Row with ประเภท='ยอดคงเหลือ' must be skipped."""
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        txns = self._txns_for(result, "Txn_392")
        dates = [t["txn_date"] for t in txns]
        assert "2026-03-16" not in dates, "Row with invalid ประเภท must be skipped"

    def test_account_code_on_each_transaction(self, synthetic_wb):
        """Every transaction must carry the source account_code."""
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        for t in result["transactions"]:
            assert "account_code" in t
            assert t["account_code"] is not None

    def test_source_sheet_on_each_transaction(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        for t in result["transactions"]:
            assert "source_sheet" in t
            assert t["source_sheet"].startswith("Txn_")

    def test_source_row_is_1_based_excel_row(self, synthetic_wb):
        """Row 2 of xlsx (first data row after header) → source_row=2."""
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        txns = self._txns_for(result, "Txn_392")
        first = min(txns, key=lambda t: t["source_row"])
        assert first["source_row"] == 2, (
            f"Expected source_row=2 for first data row, got {first['source_row']}"
        )

    def test_ij_totals_block_not_counted_as_txn(self, synthetic_wb):
        """
        IJ block has 'รายรับ'/'รายจ่าย'/'คงเหลือ' labels in col I/J.
        These must NOT produce transaction records (they're in col I/J, not A-G).
        """
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        # Txn_392 has 3 real transactions; if IJ totals were parsed as txns the count would be higher
        txns = self._txns_for(result, "Txn_392")
        assert len(txns) == 3, (
            f"IJ totals appear to have been parsed as transactions (got {len(txns)} instead of 3)"
        )

    def test_thai_account_code_transactions(self, synthetic_wb):
        """Txn_ชาร has 1 valid expense."""
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        txns = self._txns_for(result, "Txn_ชาร")
        assert len(txns) == 1
        assert txns[0]["direction"] == "expense"
        assert txns[0]["account_code"] == "ชาร"


class TestSalarySheet:
    """Salary_Sheet parsing."""

    def test_employee_count(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        assert len(result["salary"]) == 2, f"Expected 2 employees, got {len(result['salary'])}"

    def test_active_employee(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        active = [e for e in result["salary"] if e["is_active"]]
        assert len(active) == 1
        e = active[0]
        assert e["first_name"] == "สมชาย"
        assert e["last_name"] == "ใจดี"
        assert e["nickname"] == "โอ"
        assert e["bank"] == "กสิกร"
        assert e["bank_account_no"] == "1234567890", (
            f"Expected '1234567890', got {e['bank_account_no']!r}"
        )
        assert e["salary"] == pytest.approx(15000.0)
        assert e["sso_deduction"] == pytest.approx(750.0)
        assert e["net_salary"] == pytest.approx(14250.0)
        assert e["is_active"] is True

    def test_inactive_employee(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        inactive = [e for e in result["salary"] if not e["is_active"]]
        assert len(inactive) == 1
        e = inactive[0]
        assert e["first_name"] == "มานี"
        assert e["nickname"] is None
        assert e["is_active"] is False
        assert e["bank_account_no"] == "9876543210"

    def test_sso_zero_allowed(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        inactive = [e for e in result["salary"] if not e["is_active"]][0]
        assert inactive["sso_deduction"] == pytest.approx(0.0)


class TestAdvancesSheet:
    """เบิกเงินล่วงหน้า parsing."""

    def test_advance_count(self, synthetic_wb):
        """Rows 3-4 valid; row 5 (no date) must be skipped."""
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        assert len(result["advances"]) == 2, (
            f"Expected 2 advances, got {len(result['advances'])}"
        )

    def test_advance_fields(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        ball = next(a for a in result["advances"] if a["raw_name"] == "บอล")
        assert ball["advance_date"] == "2026-04-20"
        assert ball["amount"] == pytest.approx(1300.0)
        assert ball["note"] == "3 วัน"

    def test_advance_no_note_is_none(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        rin = next(a for a in result["advances"] if a["raw_name"] == "ริน")
        assert rin["note"] is None

    def test_advance_no_date_skipped(self, synthetic_wb):
        """Row 5 has ชื่อ='หลุย' but no date — must not appear in advances."""
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        names = [a["raw_name"] for a in result["advances"]]
        assert "หลุย" not in names, "Advance row without date must be skipped"


class TestSetupSheet:
    """Setup sheet category parsing."""

    def test_income_categories(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        cats = result["categories"]
        assert "income" in cats
        assert "เงินฝาก" in cats["income"]
        assert "ดอกเบี้ย" in cats["income"]
        assert "รายรับอื่นๆ" in cats["income"]
        assert len(cats["income"]) == 3

    def test_expense_categories(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        cats = result["categories"]
        assert "expense" in cats
        assert "ซื้อสินค้า" in cats["expense"]
        assert "เงินเดือน" in cats["expense"]
        assert len(cats["expense"]) == 4

    def test_categories_order_preserved(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        assert result["categories"]["income"][0] == "เงินฝาก"
        assert result["categories"]["expense"][0] == "ซื้อสินค้า"


class TestOverviewSheet:
    """Overview block parsing."""

    def test_overview_income(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        assert result["overview"]["income"] == pytest.approx(75000.50)

    def test_overview_expense(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        assert result["overview"]["expense"] == pytest.approx(30000.00)

    def test_overview_balance(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        assert result["overview"]["balance"] == pytest.approx(45000.50)


class TestWarningsList:
    """Skipped rows should surface in warnings."""

    def test_warnings_is_list(self, synthetic_wb):
        pc = _import_parser()
        result = pc.parse_cashbook(synthetic_wb)
        assert isinstance(result["warnings"], list)


class TestVatHelper:
    """is_novat_file helper function."""

    def test_novat_detection(self):
        pc = _import_parser()
        assert pc.is_novat_file("NoVat_Account.xlsx") is True
        assert pc.is_novat_file("/some/path/novat_account.xlsx") is True

    def test_vat_detection(self):
        pc = _import_parser()
        assert pc.is_novat_file("Vat_Account.xlsx") is False
        assert pc.is_novat_file("Account.xlsx") is False


# ══════════════════════════════════════════════════════════════════════════════
# Integration test — real file (skipped if absent)
# ══════════════════════════════════════════════════════════════════════════════

@pytest.mark.skipif(
    not os.path.exists(REAL_FILE),
    reason=f"Real cashbook file not found at {REAL_FILE}",
)
class TestRealFileIntegration:
    """Parse the actual NoVat_Account.xlsx and verify shape + reconciliation."""

    @pytest.fixture(scope="class")
    def real_result(self):
        pc = _import_parser()
        return pc.parse_cashbook(REAL_FILE)

    def test_six_accounts(self, real_result):
        """Exactly 6 Txn_ sheets → 6 accounts."""
        codes = {a["code"] for a in real_result["accounts"]}
        assert codes == {"392", "LEX", "SPX", "ชฎามาศ", "กิติยา", "904"}, (
            f"Unexpected account codes: {codes}"
        )
        assert len(real_result["accounts"]) == 6

    def test_transactions_exist(self, real_result):
        assert len(real_result["transactions"]) > 0

    def test_overview_parsed(self, real_result):
        ov = real_result["overview"]
        assert ov["income"]  == pytest.approx(534952.28,     abs=0.01)
        assert ov["expense"] == pytest.approx(1459572.501,   abs=0.01)
        assert ov["balance"] == pytest.approx(-924620.2215,  abs=0.01)

    def test_reconciliation_balance(self, real_result):
        """
        Σ income − Σ expense across Txn_392/LEX/SPX/ชฎามาศ/กิติยา
        (excluding Txn_904 which is a transfer/passthrough account)
        must reconcile to overview.balance within ฿0.05.

        Txn_904 is excluded because its income = expense (it is a passthrough)
        and the Overview sheet does not include it in its totals.
        """
        PASSTHROUGH_ACCOUNTS = {"904"}
        inc = sum(
            t["amount"]
            for t in real_result["transactions"]
            if t["direction"] == "income"
            and t["account_code"] not in PASSTHROUGH_ACCOUNTS
        )
        exp = sum(
            t["amount"]
            for t in real_result["transactions"]
            if t["direction"] == "expense"
            and t["account_code"] not in PASSTHROUGH_ACCOUNTS
        )
        ov = real_result["overview"]
        balance = inc - exp

        income_delta  = inc - ov["income"]
        expense_delta = exp - ov["expense"]
        balance_delta = balance - ov["balance"]

        # Report deltas regardless (visible in pytest output on failure)
        print(f"\nReconciliation deltas (excl. Txn_904):")
        print(f"  income:  parsed={inc:.4f}  overview={ov['income']:.4f}  delta={income_delta:.4f}")
        print(f"  expense: parsed={exp:.4f}  overview={ov['expense']:.4f}  delta={expense_delta:.4f}")
        print(f"  balance: parsed={balance:.4f}  overview={ov['balance']:.4f}  delta={balance_delta:.4f}")

        assert abs(income_delta)  <= 0.05, f"Income delta ฿{income_delta:.4f} exceeds tolerance"
        assert abs(expense_delta) <= 0.05, f"Expense delta ฿{expense_delta:.4f} exceeds tolerance"
        assert abs(balance_delta) <= 0.05, f"Balance delta ฿{balance_delta:.4f} exceeds tolerance"

    def test_per_account_ij_vs_rows(self, real_result):
        """
        For each account, the I/J sidecar totals (if non-zero) must match the
        summed transaction rows within ฿0.05.

        This is an internal consistency check on the source workbook.
        """
        # Rebuild per-account summed amounts from transaction list
        import openpyxl, datetime as _dt
        wb = openpyxl.load_workbook(REAL_FILE, data_only=True)

        discrepancies = []
        for sheet_name in wb.sheetnames:
            if not sheet_name.startswith("Txn_"):
                continue
            code = sheet_name[4:]
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(min_row=1, values_only=True))

            # IJ sidecar
            sidecar = {}
            for row in rows:
                if len(row) > 8 and row[8] is not None:
                    sidecar[row[8]] = row[9]

            ij_income  = float(sidecar.get("รายรับ",  0) or 0)
            ij_expense = float(sidecar.get("รายจ่าย", 0) or 0)

            # Summed from parsed transactions
            t_income  = sum(t["amount"] for t in real_result["transactions"]
                            if t["account_code"] == code and t["direction"] == "income")
            t_expense = sum(t["amount"] for t in real_result["transactions"]
                            if t["account_code"] == code and t["direction"] == "expense")

            if abs(t_income - ij_income) > 0.05:
                discrepancies.append(
                    f"{sheet_name}: income parsed={t_income:.4f} IJ={ij_income:.4f} delta={t_income-ij_income:.4f}"
                )
            if abs(t_expense - ij_expense) > 0.05:
                discrepancies.append(
                    f"{sheet_name}: expense parsed={t_expense:.4f} IJ={ij_expense:.4f} delta={t_expense-ij_expense:.4f}"
                )

        assert not discrepancies, (
            "Per-account I/J totals disagree with parsed rows:\n" + "\n".join(discrepancies)
        )
