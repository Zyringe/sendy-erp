"""
tests/test_cashbook_import.py
TDD tests for inventory_app/import_cashbook.py

Run RED first (no importer yet), then implement importer until GREEN.

Synthetic fixture: built programmatically with openpyxl.
Real-file integration test: skipped if /Users/putty/Downloads/NoVat_Account.xlsx absent.

Coverage:
  - idempotent re-import (run twice → same counts)
  - transfer detection (income == expense → is_transfer=1)
  - is_transfer clobber-guard (existing 1 not overwritten when heuristic says 0)
  - category upsert (setup + imported)
  - advance nickname matching (matched + unmatched)
  - employee sync: new employees created, existing contract employees NOT clobbered
  - salary_history dedupe on re-run
  - reconciliation dict present with correct keys
  - migration 056 applied (is_transfer column exists)
"""
import datetime
import os
import sqlite3

import pytest
import openpyxl

# ── Path to real file ─────────────────────────────────────────────────────────
REAL_FILE = "/Users/putty/Downloads/NoVat_Account.xlsx"

# ── Synthetic workbook builder ────────────────────────────────────────────────

def _build_import_wb(path, *, include_transfer_account=True):
    """
    Build a minimal synthetic cashbook xlsx suitable for import tests.

    Accounts:
      Txn_MAIN  — normal account, income ≠ expense (income 10000, expense 3000)
      Txn_PASS  — transfer/passthrough: income == expense (both 5000, only if
                  include_transfer_account=True)

    Salary_Sheet has 3 employees:
      - วุฒิพงษ์ แปงนุจา  (EMP001 — nickname 'บอล', bank already seeded)
      - สมใหม่ มาใหม่     (new employee, active, sso 600, salary 14000)
      - อาสา ไม่มีแบงค์   (new employee, active, sso 0, salary 10000)

    เบิกเงินล่วงหน้า:
      - 'บอล'   2026-04-10  2000  (will match EMP001)
      - 'ไม่รู้'  2026-04-15  500   (unmatched name)

    Setup:
      income:  ['เงินฝาก', 'ยอดขาย']
      expense: ['เงินเดือน', 'ค่าใช้จ่าย']

    Overview:
      income  10000.00 (or 15000.00 if transfer included)
      expense  3000.00 (or  8000.00 if transfer included)
      balance  7000.00
    """
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    # ── Overview ──────────────────────────────────────────────────────────────
    ws_ov = wb.create_sheet("Overview")
    if include_transfer_account:
        ov_income  = 10000.0   # excludes PASS account
        ov_expense = 3000.0
    else:
        ov_income  = 10000.0
        ov_expense = 3000.0
    ws_ov.cell(row=3, column=2, value="รายรับ");  ws_ov.cell(row=3, column=3, value=ov_income)
    ws_ov.cell(row=4, column=2, value="รายจ่าย"); ws_ov.cell(row=4, column=3, value=ov_expense)
    ws_ov.cell(row=5, column=2, value="คงเหลือ"); ws_ov.cell(row=5, column=3, value=ov_income - ov_expense)

    # ── Txn_MAIN — normal P&L account ─────────────────────────────────────────
    _write_txn_sheet(wb, "Txn_MAIN", "KBank", "1112223330", "สมศักดิ์ ทดสอบ", [
        # (date, direction, category, amount)
        (datetime.datetime(2026, 4, 1), "รายรับ",  "เงินฝาก",   7000.0),
        (datetime.datetime(2026, 4, 2), "รายรับ",  "ยอดขาย",    3000.0),
        (datetime.datetime(2026, 4, 3), "รายจ่าย", "เงินเดือน", 3000.0),
    ])

    # ── Txn_PASS — transfer/passthrough (income == expense) ───────────────────
    if include_transfer_account:
        _write_txn_sheet(wb, "Txn_PASS", "SCB", "9998887770", "โอนผ่าน", [
            (datetime.datetime(2026, 4, 1), "รายรับ",  "เงินฝาก",   5000.0),
            (datetime.datetime(2026, 4, 1), "รายจ่าย", "โอน",       5000.0),
        ])

    # ── Salary_Sheet ──────────────────────────────────────────────────────────
    ws_sal = wb.create_sheet("Salary_Sheet")
    ws_sal.cell(row=1, column=1, value="des")
    sal_hdr = ["", "ชื่อ", "นามสกุล", "ชื่อเล่น", "ธนาคาร", "เลขบัญชี",
               "เงินเดือน", "หักประกันสังคม", "เงินเดือนสุทธิ", "is_active"]
    for ci, h in enumerate(sal_hdr, 1):
        ws_sal.cell(row=2, column=ci, value=h)

    # row 3: EMP001 วุฒิพงษ์ (existing) — nickname 'บอล'
    ws_sal.cell(row=3, column=2, value="วุฒิพงษ์")
    ws_sal.cell(row=3, column=3, value="แปงนุจา")
    ws_sal.cell(row=3, column=4, value="บอล")
    ws_sal.cell(row=3, column=5, value="กสิกร")
    ws_sal.cell(row=3, column=6, value=1872917469.0)   # account no as float
    ws_sal.cell(row=3, column=7, value=13000.0)
    ws_sal.cell(row=3, column=8, value=650.0)
    ws_sal.cell(row=3, column=9, value=12350.0)
    ws_sal.cell(row=3, column=10, value=True)

    # row 4: new employee สมใหม่ — active, sso=600
    ws_sal.cell(row=4, column=2, value="สมใหม่")
    ws_sal.cell(row=4, column=3, value="มาใหม่")
    ws_sal.cell(row=4, column=4, value="ใหม่")
    ws_sal.cell(row=4, column=5, value="ไทยพาณิชย์")
    ws_sal.cell(row=4, column=6, value=5551234560.0)
    ws_sal.cell(row=4, column=7, value=14000.0)
    ws_sal.cell(row=4, column=8, value=600.0)
    ws_sal.cell(row=4, column=9, value=13400.0)
    ws_sal.cell(row=4, column=10, value=True)

    # row 5: new employee อาสา — active, sso=0
    ws_sal.cell(row=5, column=2, value="อาสา")
    ws_sal.cell(row=5, column=3, value="ไม่มีแบงค์")
    ws_sal.cell(row=5, column=4, value=None)
    ws_sal.cell(row=5, column=5, value=None)
    ws_sal.cell(row=5, column=6, value=None)
    ws_sal.cell(row=5, column=7, value=10000.0)
    ws_sal.cell(row=5, column=8, value=0.0)
    ws_sal.cell(row=5, column=9, value=10000.0)
    ws_sal.cell(row=5, column=10, value=True)

    # ── เบิกเงินล่วงหน้า ──────────────────────────────────────────────────────
    ws_adv = wb.create_sheet("เบิกเงินล่วงหน้า")
    ws_adv.cell(row=2, column=2, value="วันที่")
    ws_adv.cell(row=2, column=3, value="ชื่อ")
    ws_adv.cell(row=2, column=4, value="เบิกเงินล่วงหน้า")
    ws_adv.cell(row=2, column=5, value="หมายเหตุ")

    # advance 1: บอล → matched to EMP001
    ws_adv.cell(row=3, column=2, value=datetime.datetime(2026, 4, 10))
    ws_adv.cell(row=3, column=3, value="บอล")
    ws_adv.cell(row=3, column=4, value=2000.0)
    ws_adv.cell(row=3, column=5, value="ค่าอาหาร")

    # advance 2: ไม่รู้ → unmatched
    ws_adv.cell(row=4, column=2, value=datetime.datetime(2026, 4, 15))
    ws_adv.cell(row=4, column=3, value="ไม่รู้")
    ws_adv.cell(row=4, column=4, value=500.0)
    ws_adv.cell(row=4, column=5, value=None)

    # ── Setup ─────────────────────────────────────────────────────────────────
    ws_setup = wb.create_sheet("Setup")
    ws_setup.cell(row=2, column=2, value="รายรับ")
    ws_setup.cell(row=2, column=3, value="รายจ่าย")
    for i, cat in enumerate(["เงินฝาก", "ยอดขาย"], 3):
        ws_setup.cell(row=i, column=2, value=cat)
    for i, cat in enumerate(["เงินเดือน", "ค่าใช้จ่าย"], 3):
        ws_setup.cell(row=i, column=3, value=cat)

    wb.save(path)


def _write_txn_sheet(wb, sheet_name, bank, acct_no, owner, rows):
    """Helper: create a Txn_* sheet with IJ sidecar and given transaction rows."""
    ws = wb.create_sheet(sheet_name)
    headers = ["วันที่", "ประเภท", "หมวดหมู่", "หมวดหมู่_ผู้ใช้", "จำนวนเงิน", "รายละเอียด", "หมายเหตุ"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)

    # IJ sidecar
    ws.cell(row=2, column=9, value="Bank");          ws.cell(row=2, column=10, value=bank)
    ws.cell(row=3, column=9, value="Account Number"); ws.cell(row=3, column=10, value=acct_no)
    ws.cell(row=4, column=9, value="Name");           ws.cell(row=4, column=10, value=owner)

    for i, (date, direction, category, amount) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=date)
        ws.cell(row=i, column=2, value=direction)
        ws.cell(row=i, column=3, value=category)
        ws.cell(row=i, column=5, value=amount)


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture
def synth_wb(tmp_path):
    """Synthetic workbook path (with PASS transfer account)."""
    p = str(tmp_path / "NoVat_Test.xlsx")
    _build_import_wb(p, include_transfer_account=True)
    return p


@pytest.fixture
def synth_wb_no_transfer(tmp_path):
    """Synthetic workbook path (without PASS account)."""
    p = str(tmp_path / "NoVat_Test_NoPass.xlsx")
    _build_import_wb(p, include_transfer_account=False)
    return p


def _import_cashbook():
    from inventory_app import import_cashbook
    return import_cashbook


# ── Helper: ensure mig 056 + 067 on the tmp_db ───────────────────────────────

def _ensure_mig056(conn):
    """
    Apply migration 056 if cashbook_accounts.is_transfer is missing,
    AND migration 067 if cashbook_transactions still has vat_flag.

    The fixture copies the live DB which may pre-date 067 — without this,
    the tests' INSERT statements (which no longer include vat_flag) collide
    with the NOT NULL CHECK constraint on the legacy column.
    """
    cols = [r[1] for r in conn.execute("PRAGMA table_info(cashbook_accounts)").fetchall()]
    if 'is_transfer' not in cols:
        conn.execute(
            "ALTER TABLE cashbook_accounts ADD COLUMN is_transfer INTEGER NOT NULL DEFAULT 0 "
            "CHECK(is_transfer IN (0,1))"
        )
        conn.commit()

    txn_cols = [r[1] for r in conn.execute("PRAGMA table_info(cashbook_transactions)").fetchall()]
    if 'vat_flag' in txn_cols:
        import os as _os
        mig_path = _os.path.join(
            _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
            'data', 'migrations', '067_drop_cashbook_vat_flag.sql',
        )
        with open(mig_path, 'r') as f:
            conn.executescript(f.read())
        conn.commit()


# ══════════════════════════════════════════════════════════════════════════════
# Unit / functional tests — synthetic workbook
# ══════════════════════════════════════════════════════════════════════════════

class TestModuleImportable:

    def test_import_cashbook_importable(self):
        mod = _import_cashbook()
        assert hasattr(mod, "import_cashbook"), "import_cashbook function must be present"
        assert hasattr(mod, "sync_salary_sheet"), "sync_salary_sheet function must be present"


class TestMigration056:
    """Migration 056 adds is_transfer column to cashbook_accounts."""

    def test_is_transfer_column_exists_after_import(self, synth_wb, tmp_db, tmp_db_conn):
        """Importing via import_cashbook must leave is_transfer column present."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        cols = [r[1] for r in tmp_db_conn.execute(
            "PRAGMA table_info(cashbook_accounts)"
        ).fetchall()]
        assert "is_transfer" in cols, "is_transfer column missing from cashbook_accounts"


class TestAccountUpsert:

    def test_accounts_created(self, synth_wb, tmp_db, tmp_db_conn):
        """MAIN and PASS accounts must appear in cashbook_accounts."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        codes = {r[0] for r in tmp_db_conn.execute(
            "SELECT code FROM cashbook_accounts"
        ).fetchall()}
        assert "MAIN" in codes
        assert "PASS" in codes

    def test_account_bank_meta_stored(self, synth_wb, tmp_db, tmp_db_conn):
        """Bank name / account no / owner should be persisted."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        row = tmp_db_conn.execute(
            "SELECT bank_name, bank_account_no, account_owner_name FROM cashbook_accounts WHERE code='MAIN'"
        ).fetchone()
        assert row is not None
        assert row[0] == "KBank"
        assert row[1] == "1112223330"
        assert row[2] == "สมศักดิ์ ทดสอบ"

    def test_account_upsert_preserves_id(self, synth_wb, tmp_db, tmp_db_conn):
        """Running import twice must not create duplicate account rows."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        id_before = tmp_db_conn.execute(
            "SELECT id FROM cashbook_accounts WHERE code='MAIN'"
        ).fetchone()[0]

        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        id_after = tmp_db_conn.execute(
            "SELECT id FROM cashbook_accounts WHERE code='MAIN'"
        ).fetchone()[0]
        count = tmp_db_conn.execute(
            "SELECT COUNT(*) FROM cashbook_accounts WHERE code='MAIN'"
        ).fetchone()[0]
        assert count == 1, "Re-import must not duplicate accounts"
        assert id_before == id_after, "Account id must be stable across re-imports"


class TestTransferDetection:

    def test_pass_account_is_transfer(self, synth_wb, tmp_db, tmp_db_conn):
        """PASS account (income==expense==5000) → is_transfer=1."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        row = tmp_db_conn.execute(
            "SELECT is_transfer FROM cashbook_accounts WHERE code='PASS'"
        ).fetchone()
        assert row is not None
        assert row[0] == 1, "PASS account (income==expense) must be flagged is_transfer=1"

    def test_main_account_not_transfer(self, synth_wb, tmp_db, tmp_db_conn):
        """MAIN account (income != expense) → is_transfer=0."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        row = tmp_db_conn.execute(
            "SELECT is_transfer FROM cashbook_accounts WHERE code='MAIN'"
        ).fetchone()
        assert row is not None
        assert row[0] == 0, "MAIN account must NOT be flagged is_transfer=1"

    def test_transfer_clobber_guard(self, synth_wb_no_transfer, tmp_db, tmp_db_conn):
        """
        If an account already has is_transfer=1 but the current file's heuristic
        returns False (e.g. partial-month snapshot), the importer must NOT flip it
        to 0 — it must leave it at 1 and emit a warning.
        """
        _ensure_mig056(tmp_db_conn)
        # Manually pre-seed a MAIN account row with is_transfer=1
        tmp_db_conn.execute(
            """INSERT INTO cashbook_accounts (code, bank_name, is_transfer)
               VALUES ('MAIN', 'KBank', 1)"""
        )
        tmp_db_conn.commit()

        mod = _import_cashbook()
        result = mod.import_cashbook(synth_wb_no_transfer, conn=tmp_db_conn)
        tmp_db_conn.commit()

        # is_transfer must remain 1 despite heuristic returning 0
        row = tmp_db_conn.execute(
            "SELECT is_transfer FROM cashbook_accounts WHERE code='MAIN'"
        ).fetchone()
        assert row[0] == 1, "Clobber guard: pre-existing is_transfer=1 must not be reset to 0"

        # A warning must have been emitted
        all_warnings = result.get("warnings", [])
        has_warning = any("is_transfer" in w or "transfer" in w.lower() for w in all_warnings)
        assert has_warning, f"Expected a clobber-guard warning in result['warnings'], got: {all_warnings}"


class TestTransactionInsert:

    def test_transactions_inserted(self, synth_wb, tmp_db, tmp_db_conn):
        """Transactions rows must be stored in cashbook_transactions."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        n = tmp_db_conn.execute("SELECT COUNT(*) FROM cashbook_transactions").fetchone()[0]
        assert n > 0

    def test_main_account_rows(self, synth_wb, tmp_db, tmp_db_conn):
        """MAIN account has 3 valid transaction rows."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        account_id = tmp_db_conn.execute(
            "SELECT id FROM cashbook_accounts WHERE code='MAIN'"
        ).fetchone()[0]
        n = tmp_db_conn.execute(
            "SELECT COUNT(*) FROM cashbook_transactions WHERE account_id=?",
            (account_id,)
        ).fetchone()[0]
        assert n == 3, f"Expected 3 txns for MAIN, got {n}"

    def test_source_file_stored(self, synth_wb, tmp_db, tmp_db_conn):
        """source_file should be the basename of the input path."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        # Scope to the synthetic MAIN/PASS accounts — avoids picking up
        # source_file values from pre-existing live cashbook rows.
        row = tmp_db_conn.execute(
            """SELECT DISTINCT source_file FROM cashbook_transactions
               WHERE account_id IN (
                   SELECT id FROM cashbook_accounts WHERE code IN ('MAIN', 'PASS')
               )
               LIMIT 1"""
        ).fetchone()
        assert row is not None, "No transactions found for synthetic MAIN/PASS accounts"
        assert row[0] == os.path.basename(synth_wb), (
            f"Expected basename {os.path.basename(synth_wb)!r}, got {row[0]!r}"
        )

    def test_import_batch_id_present(self, synth_wb, tmp_db, tmp_db_conn):
        """All rows from one import run share the same import_batch_id."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        # Scope to the synthetic MAIN/PASS accounts so pre-existing live rows
        # (which carry their own batch_id) do not pollute the assertion.
        ids = {r[0] for r in tmp_db_conn.execute(
            """SELECT DISTINCT import_batch_id FROM cashbook_transactions
               WHERE account_id IN (
                   SELECT id FROM cashbook_accounts WHERE code IN ('MAIN', 'PASS')
               )"""
        ).fetchall()}
        assert len(ids) == 1, (
            f"Expected 1 batch_id across all synthetic-batch rows, got {len(ids)}: {ids}"
        )
        assert None not in ids, "import_batch_id must not be NULL"


class TestIdempotency:

    def test_run_twice_same_txn_count(self, synth_wb, tmp_db, tmp_db_conn):
        """Importing the same file twice must not double the transaction count."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        count1 = tmp_db_conn.execute("SELECT COUNT(*) FROM cashbook_transactions").fetchone()[0]

        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        count2 = tmp_db_conn.execute("SELECT COUNT(*) FROM cashbook_transactions").fetchone()[0]

        assert count1 == count2, (
            f"Idempotency broken: first import={count1}, second import={count2}"
        )

    def test_run_twice_same_account_count(self, synth_wb, tmp_db, tmp_db_conn):
        """Re-import must not duplicate cashbook_accounts rows."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        c1 = tmp_db_conn.execute("SELECT COUNT(*) FROM cashbook_accounts").fetchone()[0]

        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        c2 = tmp_db_conn.execute("SELECT COUNT(*) FROM cashbook_accounts").fetchone()[0]

        assert c1 == c2, f"Account count changed on re-import: {c1} → {c2}"

    def test_run_twice_same_advance_count(self, synth_wb, tmp_db, tmp_db_conn):
        """Re-import must not duplicate salary_advances rows."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        c1 = tmp_db_conn.execute("SELECT COUNT(*) FROM salary_advances").fetchone()[0]

        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        c2 = tmp_db_conn.execute("SELECT COUNT(*) FROM salary_advances").fetchone()[0]

        assert c1 == c2, f"salary_advances count changed on re-import: {c1} → {c2}"

    def test_run_twice_same_employee_count(self, synth_wb, tmp_db, tmp_db_conn):
        """Re-import must not create duplicate employees."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        c1 = tmp_db_conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0]

        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        c2 = tmp_db_conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0]

        assert c1 == c2, f"employees count changed on re-import: {c1} → {c2}"


class TestCategoryUpsert:

    def test_setup_categories_present(self, synth_wb, tmp_db, tmp_db_conn):
        """Income categories from Setup sheet → cashbook_categories with source='setup'."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        rows = tmp_db_conn.execute(
            "SELECT name, direction, source FROM cashbook_categories ORDER BY name"
        ).fetchall()
        names = {r[0] for r in rows}
        assert "เงินฝาก" in names
        assert "ยอดขาย" in names
        assert "เงินเดือน" in names
        assert "ค่าใช้จ่าย" in names

    def test_categories_no_duplicates(self, synth_wb, tmp_db, tmp_db_conn):
        """Running import twice must not create duplicate category rows."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        c1 = tmp_db_conn.execute("SELECT COUNT(*) FROM cashbook_categories").fetchone()[0]

        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        c2 = tmp_db_conn.execute("SELECT COUNT(*) FROM cashbook_categories").fetchone()[0]

        assert c1 == c2, f"Category count changed on re-import: {c1} → {c2}"

    def test_imported_category_from_transaction(self, synth_wb, tmp_db, tmp_db_conn):
        """Category 'โอน' appears only in PASS transactions (not in Setup) → source='imported'."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        row = tmp_db_conn.execute(
            "SELECT source FROM cashbook_categories WHERE name='โอน'"
        ).fetchone()
        # 'โอน' is an expense category in PASS account, not in Setup
        assert row is not None, "Category 'โอน' from transactions must be upserted"
        assert row[0] == "imported", f"Expected source='imported', got {row[0]!r}"


class TestSalaryAdvances:

    def test_advances_inserted(self, synth_wb, tmp_db, tmp_db_conn):
        """2 advances (บอล + ไม่รู้) should be in salary_advances for this import batch."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        # Scope to this batch's source_file so advances from any other imported
        # cashbook files (present in the live DB snapshot or future imports) do
        # not affect the count.
        n = tmp_db_conn.execute(
            "SELECT COUNT(*) FROM salary_advances WHERE source_file=?",
            (os.path.basename(synth_wb),),
        ).fetchone()[0]
        assert n == 2, f"Expected 2 advances from this batch, got {n}"

    def test_advance_ball_matched(self, synth_wb, tmp_db, tmp_db_conn):
        """'บอล' nickname → matched to EMP001 (employee_id set).

        Under no-clobber the import will not fill EMP001's nickname from the
        sheet. Pre-seed it here so the advance lookup has something to match.
        This is the same operator step Put performs in HR UI in production.
        """
        _ensure_mig056(tmp_db_conn)
        # The live-DB clone already nicknames EMP005 (วุฒิพงษ์) 'บอล'. The multi-key
        # resolver refuses to guess between two employees sharing a nickname, so
        # clear the collision first — then 'บอล' uniquely identifies EMP001.
        tmp_db_conn.execute("UPDATE employees SET nickname=NULL WHERE nickname='บอล'")
        tmp_db_conn.execute(
            "UPDATE employees SET nickname='บอล' WHERE emp_code='EMP001'"
        )
        tmp_db_conn.commit()
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        row = tmp_db_conn.execute(
            "SELECT employee_id, raw_name FROM salary_advances WHERE raw_name='บอล'"
        ).fetchone()
        assert row is not None, "Advance for 'บอล' not found"
        # EMP001 id=1 in the live DB seed; in tmp_db it's the same
        emp_row = tmp_db_conn.execute(
            "SELECT id FROM employees WHERE emp_code='EMP001'"
        ).fetchone()
        assert emp_row is not None
        assert row[0] == emp_row[0], (
            f"'บอล' should match EMP001 (id={emp_row[0]}), got employee_id={row[0]}"
        )

    def test_advance_unmatched_has_null_employee(self, synth_wb, tmp_db, tmp_db_conn):
        """'ไม่รู้' has no matching nickname → employee_id=NULL, raw_name preserved."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        row = tmp_db_conn.execute(
            "SELECT employee_id, raw_name FROM salary_advances WHERE raw_name='ไม่รู้'"
        ).fetchone()
        assert row is not None, "Advance for 'ไม่รู้' not found"
        assert row[0] is None, f"Unmatched advance must have employee_id=NULL, got {row[0]}"
        assert row[1] == "ไม่รู้"

    def test_unmatched_advance_in_summary(self, synth_wb, tmp_db, tmp_db_conn):
        """Summary dict must report unmatched advances."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        result = mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        adv = result.get("advances", {})
        assert adv.get("unmatched", 0) >= 1, (
            f"Expected at least 1 unmatched advance in summary, got: {adv}"
        )


class TestEmployeeSync:

    def test_new_employees_created(self, synth_wb, tmp_db, tmp_db_conn):
        """Salary_Sheet has 3 rows; 2 exist → 2 new employees should be created."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        before = tmp_db_conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        after = tmp_db_conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
        assert after == before + 2, (
            f"Expected {before+2} employees after import, got {after}"
        )

    def test_new_employee_salary_history(self, synth_wb, tmp_db, tmp_db_conn):
        """New employees should have one salary_history row (effective_date=2026-01-01)."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        row = tmp_db_conn.execute(
            """SELECT esh.monthly_salary, esh.effective_date, esh.reason
               FROM employees e
               JOIN employee_salary_history esh ON esh.employee_id = e.id
               WHERE e.full_name = 'สมใหม่ มาใหม่'"""
        ).fetchone()
        assert row is not None, "New employee สมใหม่ มาใหม่ must have a salary_history row"
        assert row[1] == "2026-01-01", f"Expected effective_date=2026-01-01, got {row[1]!r}"
        assert row[0] == pytest.approx(14000.0)
        assert row[2] == "initial"  # CHECK constraint allows only initial/post_probation/raise/adjust

    def test_new_employee_sso_enrolled(self, synth_wb, tmp_db, tmp_db_conn):
        """New employee with sso_deduction>0 → sso_enrolled=1."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        row = tmp_db_conn.execute(
            "SELECT sso_enrolled FROM employees WHERE full_name='สมใหม่ มาใหม่'"
        ).fetchone()
        assert row is not None
        assert row[0] == 1

    def test_new_employee_no_sso(self, synth_wb, tmp_db, tmp_db_conn):
        """New employee with sso_deduction=0 → sso_enrolled=0."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        row = tmp_db_conn.execute(
            "SELECT sso_enrolled FROM employees WHERE full_name='อาสา ไม่มีแบงค์'"
        ).fetchone()
        assert row is not None
        assert row[0] == 0

    def test_new_employee_diligence_is_zero(self, synth_wb, tmp_db, tmp_db_conn):
        """New employees from salary sheet → diligence_allowance=0 (not in sheet)."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        row = tmp_db_conn.execute(
            "SELECT diligence_allowance FROM employees WHERE full_name='สมใหม่ มาใหม่'"
        ).fetchone()
        assert row is not None
        assert row[0] == pytest.approx(0.0), (
            f"New employee diligence should be 0, got {row[0]}"
        )

    def test_new_employee_company_id_1(self, synth_wb, tmp_db, tmp_db_conn):
        """New employees default to company_id=1 (BSN)."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        row = tmp_db_conn.execute(
            "SELECT company_id FROM employees WHERE full_name='สมใหม่ มาใหม่'"
        ).fetchone()
        assert row is not None
        assert row[0] == 1, f"Expected company_id=1, got {row[0]}"

    def test_emp001_nickname_preserved_under_no_clobber(self, synth_wb, tmp_db, tmp_db_conn):
        """
        No-clobber rule: whatever nickname EMP001 has in the DB before import,
        the import must not change it. Pre-seed NULL → stays NULL; the sheet's
        'บอล' is never written to an existing employee.
        """
        _ensure_mig056(tmp_db_conn)

        # Pre-seed EMP001's nickname to NULL to make the test deterministic
        # (the cloned live DB may have any value; reset it for this assertion).
        tmp_db_conn.execute(
            "UPDATE employees SET nickname=NULL WHERE emp_code='EMP001'"
        )
        tmp_db_conn.commit()

        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()

        row = tmp_db_conn.execute(
            "SELECT nickname FROM employees WHERE emp_code='EMP001'"
        ).fetchone()
        assert row is not None
        assert row[0] is None, (
            f"No-clobber: EMP001 nickname must remain NULL after import, got {row[0]!r}"
        )

    def test_emp001_salary_not_clobbered(self, synth_wb, tmp_db, tmp_db_conn):
        """HARD CONSTRAINT: EMP001's salary_history rows must NOT change."""
        _ensure_mig056(tmp_db_conn)
        # Capture before state
        before = tmp_db_conn.execute(
            """SELECT effective_date, monthly_salary, reason
               FROM employee_salary_history
               WHERE employee_id = (SELECT id FROM employees WHERE emp_code='EMP001')
               ORDER BY effective_date""",
        ).fetchall()

        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()

        after = tmp_db_conn.execute(
            """SELECT effective_date, monthly_salary, reason
               FROM employee_salary_history
               WHERE employee_id = (SELECT id FROM employees WHERE emp_code='EMP001')
               ORDER BY effective_date""",
        ).fetchall()

        assert list(before) == list(after), (
            f"EMP001 salary_history was mutated!\n  before={before}\n  after={after}"
        )

    def test_emp001_diligence_not_clobbered(self, synth_wb, tmp_db, tmp_db_conn):
        """import_cashbook (which updates EMP001, matched by the วุฒิพงษ์ row in
        the synth sheet) must not change its diligence_allowance. Seed a known
        value first — the live-DB EMP001 has drifted away from the mig-054 seed."""
        _ensure_mig056(tmp_db_conn)
        tmp_db_conn.execute(
            "UPDATE employees SET diligence_allowance=500 WHERE emp_code='EMP001'"
        )
        tmp_db_conn.commit()
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        row = tmp_db_conn.execute(
            "SELECT diligence_allowance FROM employees WHERE emp_code='EMP001'"
        ).fetchone()
        assert row is not None
        assert row[0] == pytest.approx(500.0), (
            f"EMP001 diligence must remain 500, got {row[0]}"
        )

    def test_salary_history_dedupe_on_rerun(self, synth_wb, tmp_db, tmp_db_conn):
        """Re-running import must not add duplicate salary_history rows for new employees."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        # Count salary_history for the new employee
        count1 = tmp_db_conn.execute(
            """SELECT COUNT(*) FROM employee_salary_history
               WHERE employee_id = (SELECT id FROM employees WHERE full_name='สมใหม่ มาใหม่')"""
        ).fetchone()[0]

        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        count2 = tmp_db_conn.execute(
            """SELECT COUNT(*) FROM employee_salary_history
               WHERE employee_id = (SELECT id FROM employees WHERE full_name='สมใหม่ มาใหม่')"""
        ).fetchone()[0]

        assert count1 == count2, (
            f"salary_history duplicated on re-run: {count1} → {count2}"
        )

    def test_emp_code_sequential(self, synth_wb, tmp_db, tmp_db_conn):
        """New emp_codes must be allocated sequentially from the current max suffix."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()

        # Capture the highest numeric EMP-suffix currently in the DB before import.
        existing_codes = [
            r[0] for r in tmp_db_conn.execute(
                "SELECT emp_code FROM employees WHERE emp_code LIKE 'EMP%'"
            ).fetchall()
        ]
        max_n_before = 0
        for code in existing_codes:
            try:
                n = int(code[3:])
                if n > max_n_before:
                    max_n_before = n
            except (ValueError, TypeError):
                pass

        mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()

        # The synthetic Salary_Sheet has 2 employees (สมใหม่, อาสา) that are not
        # in the DB prior to import (วุฒิพงษ์ matches EMP001 by name).
        # They must receive max_n_before+1 and max_n_before+2, contiguously.
        expected_first  = "EMP{:03d}".format(max_n_before + 1)
        expected_second = "EMP{:03d}".format(max_n_before + 2)

        new_codes = [
            r[0] for r in tmp_db_conn.execute(
                "SELECT emp_code FROM employees WHERE emp_code > ? ORDER BY emp_code",
                ("EMP{:03d}".format(max_n_before),),
            ).fetchall()
        ]
        assert len(new_codes) == 2, (
            f"Expected exactly 2 new emp_codes after import "
            f"(pre-import max was EMP{max_n_before:03d}), got {new_codes}"
        )
        assert new_codes[0] == expected_first, (
            f"First new code should be {expected_first}, got {new_codes[0]}"
        )
        assert new_codes[1] == expected_second, (
            f"Second new code should be {expected_second}, got {new_codes[1]}"
        )


class TestReconciliationSummary:

    def test_summary_has_reconciliation(self, synth_wb, tmp_db, tmp_db_conn):
        """Return dict must include a 'reconciliation' key."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        result = mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        assert "reconciliation" in result, f"Expected 'reconciliation' in result keys: {list(result.keys())}"

    def test_reconciliation_keys(self, synth_wb, tmp_db, tmp_db_conn):
        """Reconciliation block must have income, expense, balance, overview keys."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        result = mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        rec = result["reconciliation"]
        for key in ("income", "expense", "balance", "overview_income", "overview_expense", "overview_balance"):
            assert key in rec, f"reconciliation missing key {key!r}"

    def test_reconciliation_excludes_transfer(self, synth_wb, tmp_db, tmp_db_conn):
        """Reconciliation income/expense must EXCLUDE is_transfer accounts."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        result = mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        tmp_db_conn.commit()
        rec = result["reconciliation"]
        # MAIN has income=10000, expense=3000; PASS is transfer so excluded
        assert rec["income"] == pytest.approx(10000.0, abs=0.05)
        assert rec["expense"] == pytest.approx(3000.0, abs=0.05)

    def test_warnings_in_result(self, synth_wb, tmp_db, tmp_db_conn):
        """Result must include parser warnings list."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()
        result = mod.import_cashbook(synth_wb, conn=tmp_db_conn)
        assert "warnings" in result
        assert isinstance(result["warnings"], list)


# ══════════════════════════════════════════════════════════════════════════════
# Integration test — real NoVat_Account.xlsx
# ══════════════════════════════════════════════════════════════════════════════

@pytest.mark.skipif(
    not os.path.exists(REAL_FILE),
    reason=f"Real cashbook file not found at {REAL_FILE}",
)
class TestRealFileImport:
    """Import the real NoVat_Account.xlsx into a tmp_db and assert invariants."""

    def test_real_import_run1(self, tmp_db, tmp_db_conn):
        """First import: basic shape checks."""
        _ensure_mig056(tmp_db_conn)
        # Pre-seed EMP001 nickname so advance matching works under no-clobber.
        tmp_db_conn.execute(
            "UPDATE employees SET nickname='บอล' WHERE emp_code='EMP001'"
        )
        tmp_db_conn.commit()
        mod = _import_cashbook()
        result = mod.import_cashbook(REAL_FILE, conn=tmp_db_conn)
        tmp_db_conn.commit()

        # 6 accounts
        codes = {r[0] for r in tmp_db_conn.execute(
            "SELECT code FROM cashbook_accounts"
        ).fetchall()}
        assert codes == {"392", "LEX", "SPX", "ชฎามาศ", "กิติยา", "904"}, (
            f"Unexpected account codes: {codes}"
        )

        # Txn_904 must be is_transfer=1
        row = tmp_db_conn.execute(
            "SELECT is_transfer FROM cashbook_accounts WHERE code='904'"
        ).fetchone()
        assert row is not None
        assert row[0] == 1, "Txn_904 must be flagged is_transfer=1"

        # Reconciliation excluding transfers within ฿0.05 of Overview
        rec = result["reconciliation"]
        assert abs(rec["income"]  - 534952.28)   <= 0.05, f"income delta: {rec['income'] - 534952.28:.4f}"
        assert abs(rec["expense"] - 1459572.50)  <= 0.05, f"expense delta: {rec['expense'] - 1459572.50:.4f}"
        assert abs(rec["balance"] - (-924620.22)) <= 0.05, f"balance delta: {rec['balance'] - (-924620.22):.4f}"

        # 5 employees (EMP001 + EMP002 + 3 new from salary sheet)
        n = tmp_db_conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
        assert n == 5, f"Expected 5 employees, got {n}"

        # EMP001 nickname preserved (no-clobber) — pre-seeded above
        row = tmp_db_conn.execute(
            "SELECT nickname FROM employees WHERE emp_code='EMP001'"
        ).fetchone()
        assert row is not None
        assert row[0] == "บอล", f"EMP001 nickname must be 'บอล' (preserved, not clobbered), got {row[0]!r}"

        # EMP001 diligence_allowance unchanged
        row = tmp_db_conn.execute(
            "SELECT diligence_allowance FROM employees WHERE emp_code='EMP001'"
        ).fetchone()
        assert row[0] == pytest.approx(500.0)

        # EMP002 post_probation salary_history unchanged
        row = tmp_db_conn.execute(
            """SELECT monthly_salary FROM employee_salary_history
               WHERE employee_id=(SELECT id FROM employees WHERE emp_code='EMP002')
               AND effective_date='2026-07-01' AND reason='post_probation'"""
        ).fetchone()
        assert row is not None
        assert row[0] == pytest.approx(15000.0)

        # salary_advances populated — บอล/หลุย/ริน should be matched
        for nickname in ("บอล", "หลุย", "ริน"):
            row = tmp_db_conn.execute(
                """SELECT sa.employee_id FROM salary_advances sa
                   WHERE sa.raw_name=?""",
                (nickname,)
            ).fetchone()
            if row is not None:
                assert row[0] is not None, (
                    f"Advance for '{nickname}' should be matched to an employee"
                )

    def test_real_import_run2_idempotent(self, tmp_db, tmp_db_conn):
        """Second import of the same file → identical row counts."""
        _ensure_mig056(tmp_db_conn)
        mod = _import_cashbook()

        mod.import_cashbook(REAL_FILE, conn=tmp_db_conn)
        tmp_db_conn.commit()
        txn1  = tmp_db_conn.execute("SELECT COUNT(*) FROM cashbook_transactions").fetchone()[0]
        adv1  = tmp_db_conn.execute("SELECT COUNT(*) FROM salary_advances").fetchone()[0]
        emp1  = tmp_db_conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
        acct1 = tmp_db_conn.execute("SELECT COUNT(*) FROM cashbook_accounts").fetchone()[0]

        mod.import_cashbook(REAL_FILE, conn=tmp_db_conn)
        tmp_db_conn.commit()
        txn2  = tmp_db_conn.execute("SELECT COUNT(*) FROM cashbook_transactions").fetchone()[0]
        adv2  = tmp_db_conn.execute("SELECT COUNT(*) FROM salary_advances").fetchone()[0]
        emp2  = tmp_db_conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
        acct2 = tmp_db_conn.execute("SELECT COUNT(*) FROM cashbook_accounts").fetchone()[0]

        assert txn1 == txn2,  f"Transactions not idempotent: {txn1} → {txn2}"
        assert adv1 == adv2,  f"Advances not idempotent: {adv1} → {adv2}"
        assert emp1 == emp2,  f"Employees not idempotent: {emp1} → {emp2}"
        assert acct1 == acct2, f"Accounts not idempotent: {acct1} → {acct2}"
