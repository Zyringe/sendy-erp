"""
parse_platform.py — Import/Export Shopee & Lazada xlsx files

Shopee Mass Update format:
  Row 0: metadata
  Row 1: (filler)
  Row 2: Thai column headers  ← header row
  Row 3-4: instruction rows
  Row 5+: actual data

Lazada Price/Stock Export format:
  Row 0-2: instruction rows
  Row 3+:  actual data (header is the pandas default from read_excel)
"""

import io
import json
import re
import warnings

# ── Patch openpyxl bug (Shopee xlsx has invalid activePane value) ─────────────
try:
    import openpyxl.worksheet.views as _opxl_views
    _orig_pane_init = _opxl_views.Pane.__init__

    def _patched_pane_init(self, **kwargs):
        valid = {'bottomRight', 'bottomLeft', 'topLeft', 'topRight'}
        if kwargs.get('activePane') not in valid:
            kwargs.pop('activePane', None)
        _orig_pane_init(self, **kwargs)

    _opxl_views.Pane.__init__ = _patched_pane_init
except Exception:
    pass


import pandas as pd

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


# ── Shopee ────────────────────────────────────────────────────────────────────

SHOPEE_METADATA_ROWS = 2   # rows before Thai header (rows 0-1)
SHOPEE_SKIP_AFTER_HDR = 3  # instruction rows after header to skip (rows 3-5)


def parse_shopee(file_obj):
    """
    Parse Shopee Mass Update xlsx.
    Returns list of dicts with keys:
      product_id_str, product_name, variation_id, variation_name,
      parent_sku, seller_sku, price, stock, raw_json
    """
    df = pd.read_excel(file_obj, header=SHOPEE_METADATA_ROWS, dtype=str)
    # Drop instruction rows (first SHOPEE_SKIP_AFTER_HDR rows after header)
    df = df.iloc[SHOPEE_SKIP_AFTER_HDR:].reset_index(drop=True)
    # Keep only rows with numeric product ID
    df = df[df['รหัสสินค้า'].str.match(r'^\d+$', na=False)].copy()

    records = []
    for _, row in df.iterrows():
        raw = {k: (None if pd.isna(v) else v) for k, v in row.items()}
        records.append({
            'product_id_str':  raw.get('รหัสสินค้า'),
            'product_name':    raw.get('ชื่อสินค้า') or '',
            'variation_id':    raw.get('รหัสตัวเลือกสินค้า'),
            'variation_name':  raw.get('ชื่อตัวเลือกสินค้า'),
            'parent_sku':      raw.get('Parent SKU'),
            'seller_sku':      raw.get('เลข SKU'),
            'price':           _to_float(raw.get('ราคา')),
            'special_price':   None,
            'stock':           _to_int(raw.get('คลัง')),
            'raw_json':        json.dumps(raw, ensure_ascii=False),
        })
    return records


def export_shopee(rows):
    """
    Generate Shopee Mass Update xlsx (BytesIO) from list of platform_skus rows.
    Preserves original raw_json columns; only price & stock are from DB.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    COLS = [
        'รหัสสินค้า', 'ชื่อสินค้า', 'รหัสตัวเลือกสินค้า', 'ชื่อตัวเลือกสินค้า',
        'Parent SKU', 'เลข SKU', 'ราคา', 'GTIN', 'คลัง', 'จำนวนการซื้อขั้นต่ำ',
        'จำนวนสินค้าสูงสุดที่ซื้อได้',
        'จำนวนสินค้าสูงสุดที่ซื้อได้ - เวลาเริ่มต้น',
        'จำนวนสินค้าสูงสุดที่ซื้อได้ - ระยะเวลา (วัน)',
        'จำนวนสินค้าสูงสุดที่ซื้อได้ - วันที่สิ้นสุด',
        'เหตุผล',
    ]

    wb = openpyxl.Workbook()
    ws = wb.active

    # Header row (row 1 = Shopee's row-2 style, orange background)
    orange = PatternFill('solid', start_color='FF6633')
    hdr_font = Font(bold=True, color='FFFFFF')
    for ci, col in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = orange
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data rows
    for ri, row in enumerate(rows, 2):
        raw = json.loads(row['raw_json']) if row['raw_json'] else {}
        # Override price & stock with current DB values
        raw['ราคา']  = row['price']  if row['price']  is not None else raw.get('ราคา')
        raw['คลัง'] = row['stock'] if row['stock'] is not None else raw.get('คลัง')
        for ci, col in enumerate(COLS, 1):
            ws.cell(row=ri, column=ci, value=raw.get(col))

    # Column widths
    widths = [15, 40, 22, 25, 15, 15, 10, 10, 8, 18, 20, 28, 24, 26, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Lazada ────────────────────────────────────────────────────────────────────

# Column header aliases: (English current, Thai legacy).  _pick() tries each in order.
_LAZADA_FIELD_ALIASES = {
    'product_name':        ('Product Name',        'ชื่อสินค้า'),
    'variation_id':        ('Shop SKU',             'ร้าน sku'),
    'variation_name':      ('Variations Combo',     'Variations Combo'),
    'price':               ('Price',                'ราคา'),
    'special_price':       ('SpecialPrice',         'SpecialPrice'),
    'special_price_start': ('SpecialPrice Start',   'SpecialPrice Start'),
    'special_price_end':   ('SpecialPrice End',     'SpecialPrice End'),
    'seller_sku':          ('SellerSKU',            'SellerSKU'),
    'sku_id':              ('sku.skuId',            'sku.skuId'),
    'category_id':         ('catId',               'catId'),
}


def _pick(raw, *aliases):
    """Return first non-blank value from raw dict matching any alias key."""
    for a in aliases:
        v = raw.get(a)
        if v is not None and str(v) not in ('', 'nan', 'None'):
            return v
    return None


def parse_lazada(file_obj):
    """
    Parse Lazada Price/Stock Export xlsx.
    Handles both English headers (current export) and Thai headers (legacy).
    Returns list of dicts.
    """
    # explicit sheet_name='template'; header on row 0
    df = pd.read_excel(file_obj, sheet_name='template', header=0, dtype=str)
    # Keep only rows where Product ID is all-digits (drops the 3 instruction rows)
    df = df[df['Product ID'].str.match(r'^\d+$', na=False)].copy()

    stock_col = _find_stock_col(df)

    records = []
    for _, row in df.iterrows():
        raw = {k: (None if pd.isna(v) else str(v)) for k, v in row.items()}
        records.append({
            'product_id_str':    raw.get('Product ID'),
            'product_name':      _pick(raw, 'Product Name', 'ชื่อสินค้า') or '',
            'variation_id':      _pick(raw, 'Shop SKU', 'ร้าน sku'),
            'variation_name':    _pick(raw, 'Variations Combo'),
            'parent_sku':        None,
            'seller_sku':        _pick(raw, 'SellerSKU'),
            'price':             _to_float(_pick(raw, 'Price', 'ราคา')),
            'special_price':     _to_float(_pick(raw, 'SpecialPrice')),
            'special_price_start': _pick(raw, 'SpecialPrice Start'),
            'special_price_end':   _pick(raw, 'SpecialPrice End'),
            'stock':             _to_int(raw.get(stock_col)),
            'raw_json':          json.dumps(raw, ensure_ascii=False),
        })
    return records


def _find_stock_col(df):
    """Find the store-stock column (the store-name column, e.g. 'บุญสวัสดิ์นำชัย').

    Extended known-set includes English export headers so Price/Product Name/etc.
    are never mistaken for the stock column.
    """
    known = {
        # Thai (legacy)
        'Product ID', 'catId', 'ชื่อสินค้า', 'currencyCode', 'sku.skuId',
        'status', 'ร้าน sku', 'SpecialPrice', 'SpecialPrice Start',
        'SpecialPrice End', 'ราคา', 'SellerSKU', 'Variations Combo',
        'tr(s-wb-product@md5key)',
        # English (current)
        'Product Name', 'Shop SKU', 'Price',
        'SpecialPrice Start', 'SpecialPrice End',
        # freight / skuimg extras
        'sku.auditStatus', 'sku.skuStatus',
        'Package Weight (kg)', 'Package Length (cm)', 'Package Width (cm)',
        'Package Height (cm)', 'Dangerous Goods', 'Standard',
        'Pre-Order(by Days) Enable', 'Pre-Order(by Days) ShipDays',
        'Images1', 'Images2', 'Images3', 'Images4', 'Images5',
        'Images6', 'Images7', 'Images8',
        'Product Images1', 'White Background Image', 'originalLocalName',
        'Product Name in EN', 'Status', 'Warranty Policy', 'Warranty Period',
        'Warranty Type', 'Main Description', 'Product Highlights',
        'TH_TISI License - License Document1', 'TH_TISI License - License Document2',
        'TH_TISI License - License Document3', 'TH_TISI License - License Code',
        'TH_MANUFACTURER_NAME License - License Code',
        'TH_IMPORTER_NAME License - License Code', 'Size Chart',
        'Brand', 'Place of Origin', 'Material', 'Model',
    }
    for col in df.columns:
        if col not in known:
            return col
    return 'บุญสวัสดิ์นำชัย'  # fallback


# ── Shopee multi-file product parser ─────────────────────────────────────────

def _shopee_find_header_row(file_obj, seek_col='รหัสสินค้า'):
    """Return (header_row_index, df) for a Shopee xlsx by finding the first
    row whose first cell equals seek_col.  Handles the shipping_info extra blank row."""
    raw = pd.read_excel(file_obj, header=None, dtype=str)
    for idx, row in raw.iterrows():
        if str(row.iloc[0]).strip() == seek_col:
            return idx, raw
    # Fallback: try first cell of every row
    for idx, row in raw.iterrows():
        for cell in row:
            if str(cell).strip() == seek_col:
                return idx, raw
    return 2, raw  # last-resort: original constant


def _read_shopee_file(path_or_bytes, seek_col='รหัสสินค้า', skip_after=3):
    """Read a Shopee xlsx, locate Thai header row, drop instruction rows, return DataFrame."""
    hdr_idx, raw = _shopee_find_header_row(path_or_bytes, seek_col)
    # Re-read with the correct header row
    if hasattr(path_or_bytes, 'seek'):
        path_or_bytes.seek(0)
    df = pd.read_excel(path_or_bytes, header=hdr_idx, dtype=str)
    # Drop the 3 instruction rows that follow the header
    df = df.iloc[skip_after:].reset_index(drop=True)
    # Keep rows with numeric product ID
    df = df[df['รหัสสินค้า'].str.match(r'^\d+$', na=False)].copy()
    return df


def parse_shopee_product_files(folder):
    """
    Parse all 5 Shopee product-info xlsx files from folder.

    Returns (product_records, variation_records) where:
      product_records: list of dicts for platform_products (product grain)
      variation_records: list of dicts for platform_skus (variation grain),
                         enriched with weight/dims from shipping_info and
                         variation_image_url from media_info (best-effort).
    """
    import glob
    import os

    def _find(pattern):
        matches = glob.glob(os.path.join(folder, pattern))
        if not matches:
            raise FileNotFoundError(f"No file matching {pattern} in {folder}")
        return matches[0]

    basic_path    = _find('mass_update_basic_info*.xlsx')
    dts_path      = _find('mass_update_dts_info*.xlsx')
    media_path    = _find('mass_update_media_info*.xlsx')
    sales_path    = _find('mass_update_sales_info*.xlsx')
    shipping_path = _find('mass_update_shipping_info*.xlsx')

    # ── Product-grain files ──────────────────────────────────────────────────
    df_basic   = _read_shopee_file(basic_path)
    df_dts     = _read_shopee_file(dts_path)
    df_media   = _read_shopee_file(media_path)

    # Build product dict keyed by product_id_str
    products = {}

    def _raw(row):
        return {k: (None if pd.isna(v) else str(v)) for k, v in row.items()}

    for _, row in df_basic.iterrows():
        r = _raw(row)
        pid = r.get('รหัสสินค้า')
        if not pid:
            continue
        products[pid] = {
            'product_id_str': pid,
            'parent_sku':     r.get('Parent SKU'),
            'product_name':   r.get('ชื่อสินค้า') or '',
            'name_en':        None,
            'description':    r.get('รายละเอียดสินค้า'),
            'category_id_str': None,
            'category_name':  None,
            'brand':          None,
            'place_of_origin': None,
            'material':       None,
            'warranty_policy': None,
            'warranty_period': None,
            'status':         None,
            'cover_image_url': None,
            'image_urls':     '[]',
            'dts_info':       None,
            '_raw_parts':     [r],
        }

    for _, row in df_dts.iterrows():
        r = _raw(row)
        pid = r.get('รหัสสินค้า')
        if not pid or pid not in products:
            continue
        p = products[pid]
        if not p['category_name']:
            p['category_name'] = r.get('หมวดหมู่')
        dts = r.get('ระยะเวลาเตรียมพัสดุ') or r.get('ระยะเวลาเตรียมพัสดุสำหรับสินค้าทั่วไป')
        if not p['dts_info'] and dts:
            p['dts_info'] = dts
        p['_raw_parts'].append(r)

    for _, row in df_media.iterrows():
        r = _raw(row)
        pid = r.get('รหัสสินค้า')
        if not pid:
            continue
        if pid not in products:
            products[pid] = {
                'product_id_str': pid,
                'parent_sku': r.get('Parent SKU'),
                'product_name': r.get('ชื่อสินค้า') or '',
                'name_en': None, 'description': None, 'category_id_str': None,
                'category_name': r.get('หมวดหมู่'),
                'brand': None, 'place_of_origin': None, 'material': None,
                'warranty_policy': None, 'warranty_period': None, 'status': None,
                'cover_image_url': None, 'image_urls': '[]', 'dts_info': None,
                '_raw_parts': [],
            }
        p = products[pid]
        # Cover image
        if not p['cover_image_url']:
            p['cover_image_url'] = r.get('ภาพปก') or None
        # Gallery images (รูปภาพ 1 … รูปภาพ 8)
        gallery = []
        for i in range(1, 9):
            v = r.get(f'รูปภาพ {i}')
            if v and str(v).startswith('http'):
                gallery.append(v)
        if gallery:
            p['image_urls'] = json.dumps(gallery, ensure_ascii=False)
        # Category
        if not p['category_name']:
            p['category_name'] = r.get('หมวดหมู่')
        p['_raw_parts'].append(r)

    # Finalize product records
    product_records = []
    for pid, p in products.items():
        raw_merge = {}
        for part in p.get('_raw_parts', []):
            raw_merge.update(part)
        product_records.append({
            'product_id_str':  p['product_id_str'],
            'parent_sku':      p['parent_sku'],
            'product_name':    p['product_name'],
            'name_en':         p['name_en'],
            'description':     p['description'],
            'category_id_str': p['category_id_str'],
            'category_name':   p['category_name'],
            'brand':           p['brand'],
            'place_of_origin': p['place_of_origin'],
            'material':        p['material'],
            'warranty_policy': p['warranty_policy'],
            'warranty_period': p['warranty_period'],
            'status':          p['status'],
            'cover_image_url': p['cover_image_url'],
            'image_urls':      p['image_urls'],
            'dts_info':        p['dts_info'],
            'raw_json':        json.dumps(raw_merge, ensure_ascii=False),
        })

    # ── Variation-grain files ────────────────────────────────────────────────
    df_sales    = _read_shopee_file(sales_path)
    df_shipping = _read_shopee_file(shipping_path)

    # Shipping keyed by variation_id
    shipping_by_vid = {}
    for _, row in df_shipping.iterrows():
        r = _raw(row)
        vid = r.get('รหัสตัวเลือกสินค้า')
        if vid:
            shipping_by_vid[vid] = r

    variation_records = []
    for _, row in df_sales.iterrows():
        r = _raw(row)
        vid = r.get('รหัสตัวเลือกสินค้า')
        if not vid:
            continue
        pid = r.get('รหัสสินค้า')
        ship = shipping_by_vid.get(vid, {})

        # Shopee per-variation images are not provided in a usable form in the
        # mass-export, so variation_image_url is left NULL for Shopee.
        variation_image_url = None

        raw_merge = dict(r)
        raw_merge.update(ship)
        variation_records.append({
            'product_id_str':      pid,
            'product_name':        r.get('ชื่อสินค้า') or '',
            'variation_id':        vid,
            'variation_name':      r.get('ชื่อตัวเลือกสินค้า'),
            'parent_sku':          r.get('Parent SKU'),
            'seller_sku':          r.get('เลข SKU'),
            'price':               _to_float(r.get('ราคา')),
            'special_price':       None,
            'special_price_start': None,
            'special_price_end':   None,
            'stock':               _to_int(r.get('คลัง')),
            'gtin':                r.get('GTIN') if r.get('GTIN') and str(r.get('GTIN', '')) not in ('', 'None') else None,
            'weight_kg':           _to_float(ship.get('น้ำหนัก')),
            'length_cm':           _to_float(ship.get('ยาว')),
            'width_cm':            _to_float(ship.get('กว้าง')),
            'height_cm':           _to_float(ship.get('สูง')),
            'variation_image_url': variation_image_url,
            'raw_json':            json.dumps(raw_merge, ensure_ascii=False),
        })

    return product_records, variation_records


# ── Lazada multi-file product parser ─────────────────────────────────────────

def _read_lazada_file(path, sheet='template'):
    """Read a Lazada xlsx file from path, returning DataFrame of data rows only.

    Header is on row 0 (English); rows 1-3 are instruction rows with non-digit
    Product IDs that the all-digits filter removes.
    """
    df = pd.read_excel(path, sheet_name=sheet, header=0, dtype=str)
    df = df[df['Product ID'].str.match(r'^\d+$', na=False)].copy()
    return df


def parse_lazada_product_files(folder):
    """
    Parse all 5 Lazada product-info xlsx files from folder.

    Returns (product_records, variation_records).
    """
    import glob
    import os

    def _find(pattern):
        matches = glob.glob(os.path.join(folder, pattern))
        if not matches:
            raise FileNotFoundError(f"No file matching {pattern} in {folder}")
        return matches[0]

    basic_path     = _find('basic*.xlsx')
    pricestock_path = _find('pricestock*.xlsx')
    skuimg_path    = _find('skuimg*.xlsx')
    freight_path   = _find('freight*.xlsx')
    attribute_path = _find('attribute*.xlsx')

    def _raw(row):
        return {k: (None if pd.isna(v) else str(v)) for k, v in row.items()}

    # ── Attribute file: iterate per-category sheets ──────────────────────────
    import openpyxl
    wb_attr = openpyxl.load_workbook(attribute_path, read_only=True)
    attr_sheets = [
        s for s in wb_attr.sheetnames
        if not s.endswith('_hide') and s not in ('INDEX', 'ProcessResult', 'global_hide')
    ]
    wb_attr.close()

    # product_id → {brand, place_of_origin, material, category_name}
    attr_by_pid = {}
    for sheet_name in attr_sheets:
        try:
            df_a = pd.read_excel(attribute_path, sheet_name=sheet_name, header=0, dtype=str)
            data_a = df_a[df_a['Product ID'].str.match(r'^\d+$', na=False)].copy()
        except Exception:
            continue
        for _, row in data_a.iterrows():
            r = _raw(row)
            pid = r.get('Product ID')
            if not pid:
                continue
            if pid not in attr_by_pid:
                attr_by_pid[pid] = {
                    'brand': None, 'place_of_origin': None,
                    'material': None, 'category_name': sheet_name,
                }
            a = attr_by_pid[pid]
            if not a['brand']:
                a['brand'] = r.get('Brand')
            if not a['place_of_origin']:
                a['place_of_origin'] = r.get('Place of Origin')
            if not a['material']:
                a['material'] = r.get('Material')

    # ── Basic file → product grain ───────────────────────────────────────────
    df_basic = _read_lazada_file(basic_path)

    products = {}
    for _, row in df_basic.iterrows():
        r = _raw(row)
        pid = r.get('Product ID')
        if not pid:
            continue

        # Collapse gallery images (Product Images1..8)
        gallery = []
        for i in range(1, 9):
            v = r.get(f'Product Images{i}')
            if v and str(v).startswith('http'):
                gallery.append(v)

        attr = attr_by_pid.get(pid, {})
        description = r.get('Main Description') or r.get('Product Highlights')

        products[pid] = {
            'product_id_str':  pid,
            'parent_sku':      None,
            'product_name':    _pick(r, 'Product Name', 'ชื่อสินค้า') or '',
            'name_en':         r.get('Product Name in EN'),
            'description':     description,
            'category_id_str': r.get('catId'),
            'category_name':   attr.get('category_name'),
            'brand':           attr.get('brand'),
            'place_of_origin': attr.get('place_of_origin'),
            'material':        attr.get('material'),
            'warranty_policy': r.get('Warranty Policy'),
            'warranty_period': r.get('Warranty Period'),
            'status':          r.get('Status'),
            'cover_image_url': gallery[0] if gallery else None,
            'image_urls':      json.dumps(gallery, ensure_ascii=False),
            'dts_info':        None,
            'raw_json':        json.dumps(r, ensure_ascii=False),
        }

    product_records = list(products.values())

    # ── Variation-grain files ────────────────────────────────────────────────
    df_pricestock = _read_lazada_file(pricestock_path)
    df_skuimg     = _read_lazada_file(skuimg_path)
    df_freight    = _read_lazada_file(freight_path)

    # skuimg keyed by Shop SKU
    skuimg_by_vid = {}
    for _, row in df_skuimg.iterrows():
        r = _raw(row)
        vid = _pick(r, 'Shop SKU', 'ร้าน sku')
        if vid:
            skuimg_by_vid[vid] = r

    # freight keyed by synth key: "{Product ID}_TH-{sku.skuId}"
    freight_by_vid = {}
    for _, row in df_freight.iterrows():
        r = _raw(row)
        pid_f = r.get('Product ID')
        sku_id = r.get('sku.skuId')
        if pid_f and sku_id:
            synth = f"{pid_f}_TH-{sku_id}"
            freight_by_vid[synth] = r

    # Find stock column from pricestock df
    stock_col = _find_stock_col(df_pricestock)

    variation_records = []
    for _, row in df_pricestock.iterrows():
        r = _raw(row)
        vid = _pick(r, 'Shop SKU', 'ร้าน sku')
        if not vid:
            continue
        skuimg = skuimg_by_vid.get(vid, {})
        freight = freight_by_vid.get(vid, {})

        # variation_image_url from skuimg Images1
        var_img = skuimg.get('Images1') or None
        if var_img and not str(var_img).startswith('http'):
            var_img = None

        raw_merge = dict(r)
        raw_merge.update(skuimg)
        raw_merge.update(freight)

        variation_records.append({
            'product_id_str':      r.get('Product ID'),
            'product_name':        _pick(r, 'Product Name', 'ชื่อสินค้า') or '',
            'variation_id':        vid,
            'variation_name':      _pick(r, 'Variations Combo'),
            'parent_sku':          None,
            'seller_sku':          _pick(r, 'SellerSKU'),
            'price':               _to_float(_pick(r, 'Price', 'ราคา')),
            'special_price':       _to_float(_pick(r, 'SpecialPrice')),
            'special_price_start': _pick(r, 'SpecialPrice Start'),
            'special_price_end':   _pick(r, 'SpecialPrice End'),
            'stock':               _to_int(r.get(stock_col)),
            'gtin':                None,
            'weight_kg':           _to_float(freight.get('Package Weight (kg)')),
            'length_cm':           _to_float(freight.get('Package Length (cm)')),
            'width_cm':            _to_float(freight.get('Package Width (cm)')),
            'height_cm':           _to_float(freight.get('Package Height (cm)')),
            'variation_image_url': var_img,
            'raw_json':            json.dumps(raw_merge, ensure_ascii=False),
        })

    return product_records, variation_records


def export_lazada(rows):
    """
    Generate Lazada Price/Stock update xlsx (BytesIO).
    Only includes columns Lazada needs for batch price/stock update.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    COLS = ['ร้าน sku', 'ราคา', 'SpecialPrice', 'คลัง (ร้าน)']

    wb = openpyxl.Workbook()
    ws = wb.active

    blue = PatternFill('solid', start_color='003087')
    hdr_font = Font(bold=True, color='FFFFFF')
    for ci, col in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = blue
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    for ri, row in enumerate(rows, 2):
        raw = json.loads(row['raw_json']) if row['raw_json'] else {}
        stock_col = _find_stock_col_raw(raw)
        ws.cell(row=ri, column=1, value=row['variation_id'])
        ws.cell(row=ri, column=2, value=row['price']  if row['price']  is not None else raw.get('ราคา'))
        ws.cell(row=ri, column=3, value=row['special_price'] if row['special_price'] is not None else raw.get('SpecialPrice'))
        ws.cell(row=ri, column=4, value=row['stock'] if row['stock'] is not None else raw.get(stock_col))

    for ci, w in enumerate([25, 12, 14, 12], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Add a full-data sheet
    ws_full = wb.create_sheet('ข้อมูลทั้งหมด')
    if rows:
        all_raw = [json.loads(r['raw_json']) for r in rows if r['raw_json']]
        if all_raw:
            full_cols = list(all_raw[0].keys())
            for ci, col in enumerate(full_cols, 1):
                ws_full.cell(row=1, column=ci, value=col).font = Font(bold=True)
            for ri, raw in enumerate(all_raw, 2):
                price_key = 'ราคา'
                stock_key = _find_stock_col_raw(raw)
                # Override with DB values
                matching = [r for r in rows if json.loads(r['raw_json']).get('ร้าน sku') == raw.get('ร้าน sku')]
                if matching:
                    r = matching[0]
                    raw[price_key] = r['price'] if r['price'] is not None else raw.get(price_key)
                    raw['SpecialPrice'] = r['special_price'] if r['special_price'] is not None else raw.get('SpecialPrice')
                    if stock_key:
                        raw[stock_key] = r['stock'] if r['stock'] is not None else raw.get(stock_key)
                for ci, col in enumerate(full_cols, 1):
                    ws_full.cell(row=ri, column=ci, value=raw.get(col))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _find_stock_col_raw(raw_dict):
    known = {'Product ID', 'catId', 'ชื่อสินค้า', 'currencyCode', 'sku.skuId',
             'status', 'ร้าน sku', 'SpecialPrice', 'SpecialPrice Start',
             'SpecialPrice End', 'ราคา', 'SellerSKU', 'Variations Combo',
             'tr(s-wb-product@md5key)'}
    for k in raw_dict:
        if k not in known:
            return k
    return None


# ── Helpers ───────────────────────────────────────────────────────────────────

def _to_float(val):
    if val is None:
        return None
    try:
        return float(str(val).replace(',', ''))
    except (ValueError, TypeError):
        return None


def _to_int(val):
    if val is None:
        return None
    try:
        return int(float(str(val).replace(',', '')))
    except (ValueError, TypeError):
        return None


# ── Platform Mapping Export/Import ───────────────────────────────────────────

MAPPING_COLS = [
    # Read-only info (col 1-10)
    'platform_sku_id', 'platform', 'รหัสสินค้า (platform)',
    'ชื่อสินค้า (platform)', 'variation_id', 'ชื่อ variation', 'seller_sku',
    'ราคา (platform)', 'ราคาพิเศษ (platform, Lazada)', 'คลัง (platform)',
    # AI suggestion (col 11-13, pre-filled but editable)
    'internal_sku', 'ชื่อสินค้า (ระบบ) — อ่านอย่างเดียว', 'confidence_%',
    # User adjusts (col 14)
    'qty_per_sale',
    # Hint (col 15)
    'คำอธิบาย qty_per_sale',
]

_READONLY_COLS = 10   # cols 1-10 = read-only platform data
_SUGGESTION_COLS = 3  # cols 11-13 = AI suggestions (editable but pre-filled)


def export_mapping(rows, suggestions=None):
    """
    Generate mapping xlsx for user to fill in.
    rows:        list of sqlite3.Row from get_platform_mapping_data()
    suggestions: dict {sku_id -> {suggested_sku, suggested_name, confidence}}
    Returns BytesIO.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    suggestions = suggestions or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mapping'

    # ── Header ──
    fill_info  = PatternFill('solid', start_color='FFF3CD')  # yellow = platform info
    fill_ai    = PatternFill('solid', start_color='D1ECF1')  # blue   = AI suggestion
    fill_edit  = PatternFill('solid', start_color='D4EDDA')  # green  = user edits
    fill_hint  = PatternFill('solid', start_color='F8F9FA')  # grey   = hint
    hdr_font   = Font(bold=True, size=9)

    for ci, col in enumerate(MAPPING_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if ci <= _READONLY_COLS:
            cell.fill = fill_info
        elif ci <= _READONLY_COLS + _SUGGESTION_COLS:
            cell.fill = fill_ai
        elif ci < len(MAPPING_COLS):
            cell.fill = fill_edit
        else:
            cell.fill = fill_hint
    ws.row_dimensions[1].height = 40

    # Row fills by confidence band
    fill_hi   = PatternFill('solid', start_color='F0FFF4')  # ≥80%: pale green
    fill_mid  = PatternFill('solid', start_color='FFFDE7')  # 60-79%: pale yellow
    fill_lo   = PatternFill('solid', start_color='FFF3E0')  # 40-59%: pale orange
    fill_vlo  = PatternFill('solid', start_color='FFF5F5')  # <40%: pale red

    # ── Data rows ──
    for ri, r in enumerate(rows, 2):
        r = dict(r)
        sg = suggestions.get(r['id'], {})
        conf = sg.get('confidence', 0)

        # Determine row background by confidence
        if conf == 100:
            row_fill = PatternFill('solid', start_color='E8F5E9')  # already mapped
        elif conf >= 80:
            row_fill = fill_hi
        elif conf >= 60:
            row_fill = fill_mid
        elif conf >= 40:
            row_fill = fill_lo
        else:
            row_fill = fill_vlo

        # Use existing mapping first, fall back to suggestion
        int_sku  = r.get('internal_sku') or (sg.get('suggested_sku', '') if sg else '')
        int_name = r.get('internal_product_name') or (sg.get('suggested_name', '') if sg else '')
        conf_val = 100 if r.get('internal_sku') else (sg.get('confidence', '') if sg else '')

        vals = [
            r['id'],
            r['platform'],
            r['product_id_str'],
            r['product_name'],
            r['variation_id'],
            r['variation_name'] or '',
            r['seller_sku'] or '',
            r['price'],
            r['special_price'],
            r['stock'],
            # AI / existing suggestion (editable)
            int_sku,
            int_name,
            conf_val,
            # User edits
            r['qty_per_sale'] or 1,
            # Hint
            'จำนวนหน่วยในระบบที่ลดเมื่อขาย 1 ชิ้น (เช่น 0.5 กก., 6 ม้วน, 100 ตัว)',
        ]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(size=9)
            if ci <= _READONLY_COLS:
                cell.fill = row_fill
            elif ci == 12:  # ชื่อสินค้าในระบบ = read-only suggestion
                cell.fill = PatternFill('solid', start_color='EBF5FB')
                cell.font = Font(size=9, italic=True, color='2471A3')
            elif ci == 13:  # confidence%
                # Color the confidence cell
                if conf_val == '' or conf_val == 0:
                    cell.fill = PatternFill('solid', start_color='F5F5F5')
                elif conf_val == 100:
                    cell.fill = PatternFill('solid', start_color='C8E6C9')
                    cell.font = Font(size=9, bold=True, color='1B5E20')
                elif conf_val >= 80:
                    cell.fill = PatternFill('solid', start_color='DCEDC8')
                    cell.font = Font(size=9, bold=True, color='33691E')
                elif conf_val >= 60:
                    cell.fill = PatternFill('solid', start_color='FFF9C4')
                    cell.font = Font(size=9, bold=True, color='F57F17')
                else:
                    cell.fill = PatternFill('solid', start_color='FFCCBC')
                    cell.font = Font(size=9, bold=True, color='BF360C')
            elif ci == len(MAPPING_COLS):  # hint
                cell.fill = PatternFill('solid', start_color='F8F9FA')
                cell.font = Font(size=8, italic=True, color='999999')

    # ── Column widths ──
    widths = [10, 9, 16, 42, 26, 22, 14, 9, 18, 9, 13, 36, 12, 12, 52]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'A2'

    # ── Legend sheet ──
    ws2 = wb.create_sheet('คำอธิบาย')
    legends = [
        ('สีเขียวเข้ม (100%)', 'ผูกกับสินค้าในระบบแล้ว'),
        ('สีเขียวอ่อน (80-99%)', 'AI มั่นใจสูง — น่าจะถูกต้อง'),
        ('สีเหลือง (60-79%)', 'AI คิดว่าน่าจะใช่ — ตรวจสอบด้วย'),
        ('สีส้ม (40-59%)', 'AI ไม่แน่ใจ — ต้องแก้ไขเอง'),
        ('สีแดงอ่อน (<40%)', 'AI ไม่มั่นใจ — กรุณากรอก internal_sku เอง'),
        ('', ''),
        ('internal_sku', 'ใส่ตัวเลข SKU ของสินค้าในระบบ ERP'),
        ('qty_per_sale', 'จำนวนหน่วยในระบบที่ลดเมื่อขาย 1 ชิ้นบน platform'),
        ('ตัวอย่าง qty_per_sale', 'ปุ๊ก #7 (500 กรัม) = 0.5 | สายเอ็น 6 ม้วน = 6 | ลูกรีเวท 100 ตัว = 100'),
    ]
    ws2['A1'] = 'สัญลักษณ์สี'
    ws2['B1'] = 'ความหมาย'
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)
    for ri2, (k, v) in enumerate(legends, 2):
        ws2.cell(row=ri2, column=1, value=k)
        ws2.cell(row=ri2, column=2, value=v)
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 55

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_mapping(file_obj):
    """
    Parse a filled mapping xlsx.
    Returns list of dicts: {platform_sku_id, internal_sku, qty_per_sale}
    """
    df = pd.read_excel(file_obj, dtype=str)
    results = []
    for _, row in df.iterrows():
        raw = {k: (None if pd.isna(v) else str(v).strip()) for k, v in row.items()}
        sku_id = _to_int(raw.get('platform_sku_id'))
        if not sku_id:
            continue
        results.append({
            'platform_sku_id': sku_id,
            'internal_sku':    raw.get('internal_sku') or None,
            'qty_per_sale':    _to_float(raw.get('qty_per_sale')) or 1.0,
        })
    return results


# ── Order File Parsers (for Listing Mapping) ─────────────────────────────────

def parse_shopee_orders(file_obj):
    """
    Parse Shopee All-Orders xlsx (exported from My Orders).
    Extracts unique (item_name, variation) combinations.
    Returns list of dicts for ecommerce_listings.
    """
    import hashlib
    df = pd.read_excel(file_obj, dtype=str)
    if 'ชื่อสินค้า' not in df.columns:
        raise ValueError('ไม่พบคอลัมน์ ชื่อสินค้า — ตรวจสอบว่าใช้ไฟล์คำสั่งซื้อ Shopee')

    df = df.dropna(subset=['ชื่อสินค้า'])
    seen = {}
    for _, row in df.iterrows():
        name = str(row['ชื่อสินค้า']).strip()
        var  = str(row.get('ชื่อตัวเลือก', '') or '').strip() or None
        key  = hashlib.sha256(f"shopee|{name}|{var or ''}".encode()).hexdigest()[:16]
        if key in seen:
            continue
        price      = _to_float(row.get('ราคาขาย'))
        seller_sku = str(row.get('เลขอ้างอิง SKU (SKU Reference No.)', '') or '').strip() or None
        seen[key] = {
            'platform':     'shopee',
            'item_name':    name,
            'variation':    var,
            'seller_sku':   seller_sku,
            'listing_key':  key,
            'sample_price': price,
        }
    return list(seen.values())


_LAZADA_ATTR_PREFIX_RE = re.compile(
    r'^(สี|โทนสี|Color Family|Color family|Color|Variation\d*|Item|ขนาด|จำนวน|'
    r'Power Tools Battery Voltage|สินค้า)\s*:\s*',
    re.IGNORECASE,
)


def parse_lazada_orders(file_obj):
    """
    Parse Lazada Orders xlsx (exported from Seller Center).
    Extracts unique (itemName, variation) combinations.
    Returns list of dicts for ecommerce_listings.

    Lazada has changed the variation attribute label over time
    (โทนสี: → สี: → Color Family: → Color family:). We strip these prefixes
    before hashing so the same option doesn't create duplicate listings
    across years.
    """
    import hashlib
    df = pd.read_excel(file_obj, dtype=str)
    if 'itemName' not in df.columns:
        raise ValueError('ไม่พบคอลัมน์ itemName — ตรวจสอบว่าใช้ไฟล์คำสั่งซื้อ Lazada')

    df = df.dropna(subset=['itemName'])
    seen = {}
    for _, row in df.iterrows():
        name = str(row['itemName']).strip()
        var  = str(row.get('variation', '') or '').strip() or None
        if var:
            var = _LAZADA_ATTR_PREFIX_RE.sub('', var).strip() or None
        key  = hashlib.sha256(f"lazada|{name}|{var or ''}".encode()).hexdigest()[:16]
        if key in seen:
            continue
        price      = _to_float(row.get('unitPrice'))
        seller_sku = str(row.get('sellerSku', '') or '').strip() or None
        seen[key] = {
            'platform':     'lazada',
            'item_name':    name,
            'variation':    var,
            'seller_sku':   seller_sku,
            'listing_key':  key,
            'sample_price': price,
        }
    return list(seen.values())


# ── Listing Mapping Export/Import ─────────────────────────────────────────────

LISTING_MAPPING_COLS = [
    'listing_id', 'platform', 'ชื่อสินค้า (platform)', 'ตัวเลือก (platform)',
    'seller_sku', 'ราคาตัวอย่าง (฿)',
    'internal_sku', 'ชื่อสินค้า (ERP) — อ่านอย่างเดียว', 'confidence_%',
    'qty_per_sale', 'คำอธิบาย qty_per_sale',
]


def export_listing_mapping(rows, suggestions=None, unmatched_only=False):
    """
    Generate mapping xlsx for user to fill in internal_sku.
    rows: list of ecommerce_listings rows (dicts or sqlite3.Row)
    suggestions: dict {listing_id -> {suggested_sku, suggested_name, confidence}}
    unmatched_only: True → only export rows without product_id
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    suggestions = suggestions or {}
    rows = [dict(r) for r in rows]
    if unmatched_only:
        rows = [r for r in rows if not r.get('product_id')]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mapping'

    fill_info = PatternFill('solid', start_color='FFF3CD')
    fill_edit = PatternFill('solid', start_color='D4EDDA')
    fill_ai   = PatternFill('solid', start_color='D1ECF1')
    hdr_font  = Font(bold=True, size=9)

    for ci, col in enumerate(LISTING_MAPPING_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if col in ('internal_sku',):
            cell.fill = fill_edit
        elif 'ERP' in col or 'confidence' in col:
            cell.fill = fill_ai
        else:
            cell.fill = fill_info
    ws.row_dimensions[1].height = 36

    fill_mapped = PatternFill('solid', start_color='E8F5E9')
    fill_hi     = PatternFill('solid', start_color='F0FFF4')
    fill_mid    = PatternFill('solid', start_color='FFFDE7')
    fill_lo     = PatternFill('solid', start_color='FFF3E0')
    fill_vlo    = PatternFill('solid', start_color='FFF5F5')

    for ri, r in enumerate(rows, 2):
        sg   = suggestions.get(r['id'], {})
        conf = sg.get('confidence', 0)

        int_sku  = r.get('sku') or (sg.get('suggested_sku', '') if sg else '')
        int_name = r.get('product_name') or (sg.get('suggested_name', '') if sg else '')
        conf_val = 100 if r.get('product_id') else (sg.get('confidence', '') if sg else '')

        if r.get('product_id'):
            row_fill = fill_mapped
        elif conf >= 80:
            row_fill = fill_hi
        elif conf >= 60:
            row_fill = fill_mid
        elif conf >= 40:
            row_fill = fill_lo
        else:
            row_fill = fill_vlo

        qty_per_sale = r.get('qty_per_sale') or 1

        vals = [
            r['id'], r['platform'], r['item_name'], r.get('variation') or '',
            r.get('seller_sku') or '', r.get('sample_price'),
            int_sku, int_name, conf_val,
            qty_per_sale,
            'จำนวนหน่วยในระบบที่ลดเมื่อขาย 1 ชิ้น เช่น ลูกรีเวท 25 ดอก/ถุง → ใส่ 25',
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(size=9)
            col_name = LISTING_MAPPING_COLS[ci - 1]
            if col_name in ('internal_sku', 'qty_per_sale'):
                cell.fill = fill_edit
            elif 'ERP' in col_name:
                cell.fill = PatternFill('solid', start_color='EBF5FB')
                cell.font = Font(size=9, italic=True, color='2471A3')
            elif col_name == 'confidence_%':
                if conf_val == '' or conf_val == 0:
                    cell.fill = PatternFill('solid', start_color='F5F5F5')
                elif conf_val == 100:
                    cell.fill = PatternFill('solid', start_color='C8E6C9')
                    cell.font = Font(size=9, bold=True, color='1B5E20')
                elif conf_val >= 80:
                    cell.fill = PatternFill('solid', start_color='DCEDC8')
                    cell.font = Font(size=9, bold=True, color='33691E')
                elif conf_val >= 60:
                    cell.fill = PatternFill('solid', start_color='FFF9C4')
                    cell.font = Font(size=9, bold=True, color='F57F17')
                else:
                    cell.fill = PatternFill('solid', start_color='FFCCBC')
                    cell.font = Font(size=9, bold=True, color='BF360C')
            elif col_name == 'คำอธิบาย qty_per_sale':
                cell.fill = PatternFill('solid', start_color='F8F9FA')
                cell.font = Font(size=8, italic=True, color='999999')
            else:
                cell.fill = row_fill

    widths = [10, 9, 50, 28, 16, 14, 13, 40, 12, 12, 55]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_listing_mapping(file_obj):
    """
    Parse filled listing mapping xlsx.
    Returns list of dicts: {listing_id, internal_sku, qty_per_sale}
    """
    df = pd.read_excel(file_obj, dtype=str)
    results = []
    for _, row in df.iterrows():
        raw = {k: (None if pd.isna(v) else str(v).strip()) for k, v in row.items()}
        lid = _to_int(raw.get('listing_id'))
        if not lid:
            continue
        results.append({
            'listing_id':    lid,
            'internal_sku':  raw.get('internal_sku') or None,
            'qty_per_sale':  _to_float(raw.get('qty_per_sale')) or 1.0,
        })
    return results
