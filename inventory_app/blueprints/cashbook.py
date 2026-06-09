"""Cashbook blueprint — รายรับ/รายจ่าย dashboard, ledger, import, export.

Access control
--------------
  admin   : full access (GET + import POST + export)
  manager : read-only (GET dashboard/ledger/export); POST /cashbook/import is
            blocked by before_request (not in _MANAGER_POST_OK) + belt-and-braces
            abort(403) inside the POST handler.
  staff   : blocked entirely — before_request redirects any cashbook.* endpoint.

Python 3.9 — no `X | None` union syntax.
"""
from __future__ import annotations

import io
import os
import tempfile
from typing import Optional

import openpyxl
from flask import (Blueprint, abort, flash, redirect, render_template,
                   request, session, url_for, make_response)

import database
import import_cashbook as cashbook_mod
from parse_cashbook import parse_cashbook

bp_cashbook = Blueprint("cashbook", __name__, url_prefix="/cashbook")


# ── Helpers ───────────────────────────────────────────────────────────────────

def _require_admin():
    if session.get("role") != "admin":
        abort(403)


def _fmt_baht(val) -> str:
    try:
        return f"฿{float(val):,.2f}"
    except (TypeError, ValueError):
        return "฿0.00"


# Categories that are capital / inter-account movements, NOT operating income or
# expense. Excluded from the headline P&L, category summary, monthly chart and the
# by-tag report. They are still real money (so they DO count toward account balance)
# and remain visible in the per-account ledger.
TRANSFER_CATEGORIES = ("เงินทุน/เงินโอน",)


def _tcat_ph():
    """Placeholder string + params for the transfer-category list."""
    return ",".join("?" * len(TRANSFER_CATEGORIES)), list(TRANSFER_CATEGORIES)


def _get_accounts_with_totals(conn):
    """
    One dict per active account. `income`/`expense` are OPERATING totals (transfer
    categories excluded) so they sum to the headline P&L. `transfer_in`/`transfer_out`
    hold the excluded capital movements. `balance` is true cash = operating + transfers.
    """
    ph, params = _tcat_ph()
    rows = conn.execute(f"""
        SELECT
            a.id,
            a.code,
            a.display_name,
            a.account_owner_name,
            a.bank_name,
            a.bank_account_no,
            a.note AS account_note,
            a.is_transfer,
            COALESCE(SUM(CASE WHEN t.direction='income'  AND COALESCE(t.category,'') NOT IN ({ph}) THEN t.amount ELSE 0 END), 0) AS income,
            COALESCE(SUM(CASE WHEN t.direction='expense' AND COALESCE(t.category,'') NOT IN ({ph}) THEN t.amount ELSE 0 END), 0) AS expense,
            COALESCE(SUM(CASE WHEN t.direction='income'  AND COALESCE(t.category,'') IN ({ph}) THEN t.amount ELSE 0 END), 0) AS transfer_in,
            COALESCE(SUM(CASE WHEN t.direction='expense' AND COALESCE(t.category,'') IN ({ph}) THEN t.amount ELSE 0 END), 0) AS transfer_out,
            COUNT(t.id) AS txn_count
        FROM cashbook_accounts a
        LEFT JOIN cashbook_transactions t ON t.account_id = a.id
        WHERE a.is_active = 1
        GROUP BY a.id
        ORDER BY a.is_transfer ASC, a.sort_order ASC, a.id ASC
    """, params * 4).fetchall()

    result = []
    for r in rows:
        d = dict(r)
        d["balance"] = (d["income"] + d["transfer_in"]) - (d["expense"] + d["transfer_out"])
        result.append(d)
    return result


def _get_monthly_summary(conn, exclude_transfer: bool = True):
    """Monthly operating income/expense (transfer categories always excluded;
    transfer accounts excluded when exclude_transfer)."""
    ph, params = _tcat_ph()
    acct_clause = "AND a.is_transfer = 0" if exclude_transfer else ""
    rows = conn.execute(f"""
        SELECT
            strftime('%Y-%m', t.txn_date) AS month,
            SUM(CASE WHEN t.direction='income'  THEN t.amount ELSE 0 END) AS income,
            SUM(CASE WHEN t.direction='expense' THEN t.amount ELSE 0 END) AS expense
        FROM cashbook_transactions t
        JOIN cashbook_accounts a ON a.id = t.account_id
        WHERE COALESCE(t.category,'') NOT IN ({ph}) {acct_clause}
        GROUP BY month
        ORDER BY month ASC
    """, params).fetchall()
    return [dict(r) for r in rows]


def _get_category_summary(conn):
    """Income and expense totals by category, excluding transfer accounts AND
    transfer categories."""
    ph, params = _tcat_ph()
    rows = conn.execute(f"""
        SELECT
            t.direction,
            COALESCE(t.category, '(ไม่ระบุ)') AS category,
            SUM(t.amount) AS total
        FROM cashbook_transactions t
        JOIN cashbook_accounts a ON a.id = t.account_id
        WHERE a.is_transfer = 0 AND COALESCE(t.category,'') NOT IN ({ph})
        GROUP BY t.direction, category
        ORDER BY t.direction DESC, total DESC
    """, params).fetchall()
    income_cats = [dict(r) for r in rows if r["direction"] == "income"]
    expense_cats = [dict(r) for r in rows if r["direction"] == "expense"]
    return income_cats, expense_cats


def _get_tag_summary(conn):
    """Operating EXPENSE grouped by ผู้ใช้ tag (user_category). Excludes transfer
    accounts, transfer categories and untagged rows."""
    ph, params = _tcat_ph()
    rows = conn.execute(f"""
        SELECT
            t.user_category AS tag,
            SUM(t.amount)    AS total,
            COUNT(*)         AS n
        FROM cashbook_transactions t
        JOIN cashbook_accounts a ON a.id = t.account_id
        WHERE a.is_transfer = 0
          AND t.direction = 'expense'
          AND COALESCE(t.category,'') NOT IN ({ph})
          AND t.user_category IS NOT NULL AND t.user_category != ''
        GROUP BY t.user_category
        ORDER BY total DESC
    """, params).fetchall()
    return [dict(r) for r in rows]


# ── Routes ────────────────────────────────────────────────────────────────────

@bp_cashbook.route("/")
def dashboard():
    conn = database.get_connection()
    try:
        accounts      = _get_accounts_with_totals(conn)
        monthly       = _get_monthly_summary(conn, exclude_transfer=True)
        income_cats, expense_cats = _get_category_summary(conn)
        tag_summary   = _get_tag_summary(conn)
    finally:
        conn.close()

    # Headline P&L excludes transfer accounts AND transfer categories (income/expense
    # are already operating-only from _get_accounts_with_totals).
    op_accounts = [a for a in accounts if not a["is_transfer"]]
    tr_accounts  = [a for a in accounts if a["is_transfer"]]

    total_income  = sum(a["income"]  for a in op_accounts)
    total_expense = sum(a["expense"] for a in op_accounts)
    # คงเหลือ = actual cash on hand = sum of true-cash account balances (which include
    # capital transfers). This reconciles with the per-account balance column. It is
    # deliberately NOT income − expense: transfers fund the gap (see disclosure note).
    total_balance = sum(a["balance"] for a in op_accounts)
    # Capital/inter-account movements excluded from the P&L (disclosure figure)
    transfer_total = sum(a["transfer_in"] + a["transfer_out"] for a in op_accounts)

    return render_template(
        "cashbook/dashboard.html",
        accounts=accounts,
        op_accounts=op_accounts,
        tr_accounts=tr_accounts,
        total_income=total_income,
        total_expense=total_expense,
        total_balance=total_balance,
        transfer_total=transfer_total,
        monthly=monthly,
        income_cats=income_cats,
        expense_cats=expense_cats,
        tag_summary=tag_summary,
    )


@bp_cashbook.route("/account/<int:account_id>")
def account_ledger(account_id):
    conn = database.get_connection()
    try:
        acct = conn.execute(
            "SELECT * FROM cashbook_accounts WHERE id=?", (account_id,)
        ).fetchone()
        if acct is None:
            flash("ไม่พบบัญชีนี้ในระบบ", "danger")
            return redirect(url_for("cashbook.dashboard"))

        month_filter = request.args.get("month", "").strip()
        dir_filter   = request.args.get("dir", "").strip()
        page         = max(1, int(request.args.get("page", 1)))
        per_page     = 50

        params = [account_id]
        where  = ["t.account_id=?"]

        if month_filter:
            where.append("strftime('%Y-%m', t.txn_date)=?")
            params.append(month_filter)
        if dir_filter in ("income", "expense"):
            where.append("t.direction=?")
            params.append(dir_filter)

        where_sql = " AND ".join(where)
        total_count = conn.execute(
            f"SELECT COUNT(*) FROM cashbook_transactions t WHERE {where_sql}",
            params,
        ).fetchone()[0]

        offset = (page - 1) * per_page
        rows = conn.execute(
            f"""SELECT t.* FROM cashbook_transactions t
                WHERE {where_sql}
                ORDER BY t.txn_date ASC, t.id ASC
                LIMIT ? OFFSET ?""",
            params + [per_page, offset],
        ).fetchall()

        # Running totals for displayed page
        sum_income  = conn.execute(
            f"SELECT COALESCE(SUM(amount),0) FROM cashbook_transactions t "
            f"WHERE {where_sql} AND direction='income'",
            params,
        ).fetchone()[0]
        sum_expense = conn.execute(
            f"SELECT COALESCE(SUM(amount),0) FROM cashbook_transactions t "
            f"WHERE {where_sql} AND direction='expense'",
            params,
        ).fetchone()[0]

        # Available months for filter dropdown
        months = conn.execute(
            """SELECT DISTINCT strftime('%Y-%m', txn_date) AS m
               FROM cashbook_transactions
               WHERE account_id=?
               ORDER BY m""",
            (account_id,),
        ).fetchall()
    finally:
        conn.close()

    total_pages = max(1, (total_count + per_page - 1) // per_page)

    return render_template(
        "cashbook/account_ledger.html",
        acct=dict(acct),
        rows=[dict(r) for r in rows],
        month_filter=month_filter,
        dir_filter=dir_filter,
        page=page,
        per_page=per_page,
        total_count=total_count,
        total_pages=total_pages,
        sum_income=sum_income,
        sum_expense=sum_expense,
        balance=sum_income - sum_expense,
        months=[r["m"] for r in months],
    )


@bp_cashbook.route("/import", methods=["GET", "POST"])
def import_view():
    if request.method == "POST":
        # Belt-and-braces admin check (before_request handles manager redirect;
        # this catches any bypass attempt).
        if session.get("role") != "admin":
            abort(403)

        f = request.files.get("cashbook_file")
        if not f or not f.filename.endswith(".xlsx"):
            flash("กรุณาเลือกไฟล์ .xlsx", "danger")
            return redirect(url_for("cashbook.import_view"))

        # Save to temp file
        tmp = tempfile.NamedTemporaryFile(
            suffix=".xlsx", delete=False,
            dir=os.path.join(os.path.dirname(os.path.dirname(__file__)), "imports")
        )
        try:
            f.save(tmp.name)
            tmp.close()

            summary = cashbook_mod.import_cashbook(tmp.name)
        finally:
            try:
                os.unlink(tmp.name)
            except OSError:
                pass

        return render_template("cashbook/import_result.html", summary=summary,
                               filename=f.filename)

    return render_template("cashbook/import.html")


@bp_cashbook.route("/export")
def export_view():
    conn = database.get_connection()
    try:
        xlsx_bytes = _build_export_xlsx(conn)
    finally:
        conn.close()

    filename = "cashbook_export.xlsx"
    response = make_response(xlsx_bytes)
    response.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ── Export builder ────────────────────────────────────────────────────────────

def _build_export_xlsx(conn) -> bytes:
    """
    Build a round-trip-compatible .xlsx in the same multi-sheet format the
    parser expects.  Sheets:
      Overview              — computed P&L (excluding transfer accounts)
      Txn_<code>            — one per account (header + rows + I/J sidecar)
      Salary_Sheet          — from employees + latest salary_history
      เบิกเงินล่วงหน้า      — from salary_advances
      Setup                 — cashbook_categories
    """
    wb = openpyxl.Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)

    # ── 1. Overview ───────────────────────────────────────────────────────────
    ws_ov = wb.create_sheet("Overview")
    accounts_raw = conn.execute(
        "SELECT id, code, is_transfer FROM cashbook_accounts WHERE is_active=1"
    ).fetchall()
    non_transfer_ids = {r["id"] for r in accounts_raw if r["is_transfer"] == 0}

    income_total  = 0.0
    expense_total = 0.0
    if non_transfer_ids:
        placeholders = ",".join("?" * len(non_transfer_ids))
        row_ov = conn.execute(
            f"""SELECT
                  COALESCE(SUM(CASE WHEN direction='income'  THEN amount ELSE 0 END), 0) AS inc,
                  COALESCE(SUM(CASE WHEN direction='expense' THEN amount ELSE 0 END), 0) AS exp
                FROM cashbook_transactions
                WHERE account_id IN ({placeholders})""",
            list(non_transfer_ids),
        ).fetchone()
        if row_ov:
            income_total  = row_ov["inc"] or 0.0
            expense_total = row_ov["exp"] or 0.0

    balance_total = income_total - expense_total
    # Layout: row 3=รายรับ, row 4=รายจ่าย, row 5=คงเหลือ  (cols B/C = indices 1/2)
    ws_ov.cell(row=3, column=2, value="รายรับ")
    ws_ov.cell(row=3, column=3, value=income_total)
    ws_ov.cell(row=4, column=2, value="รายจ่าย")
    ws_ov.cell(row=4, column=3, value=expense_total)
    ws_ov.cell(row=5, column=2, value="คงเหลือ")
    ws_ov.cell(row=5, column=3, value=balance_total)

    # ── 2. Txn_<code> sheets ─────────────────────────────────────────────────
    accounts = conn.execute(
        "SELECT * FROM cashbook_accounts WHERE is_active=1 ORDER BY sort_order, id"
    ).fetchall()

    for acct in accounts:
        code = acct["code"]
        ws = wb.create_sheet(f"Txn_{code}")

        # Header row (row 1)
        headers = ["วันที่", "ประเภท", "หมวดหมู่", "หมวดหมู่_ผู้ใช้",
                   "จำนวนเงิน", "รายละเอียด", "หมายเหตุ"]
        for ci, h in enumerate(headers, start=1):
            ws.cell(row=1, column=ci, value=h)

        # I/J sidecar meta (cols I=9, J=10)
        sidecar = [
            ("Bank",           acct["bank_name"]),
            ("Account Number", acct["bank_account_no"]),
            ("Name",           acct["account_owner_name"]),
            ("หมายเหตุ",       acct["account_note"] if "account_note" in acct.keys() else acct["note"]),
        ]
        for si, (label, value) in enumerate(sidecar, start=1):
            ws.cell(row=si, column=9, value=label)
            ws.cell(row=si, column=10, value=value)

        # Transaction rows
        txns = conn.execute(
            """SELECT txn_date, direction, category, user_category,
                      amount, description, note
               FROM cashbook_transactions
               WHERE account_id=?
               ORDER BY txn_date ASC, id ASC""",
            (acct["id"],),
        ).fetchall()

        for ri, txn in enumerate(txns, start=2):
            direction_th = "รายรับ" if txn["direction"] == "income" else "รายจ่าย"
            # Parse date string to datetime.date so openpyxl writes a real date
            import datetime as _dt
            try:
                d = _dt.date.fromisoformat(txn["txn_date"])
            except (TypeError, ValueError):
                d = txn["txn_date"]
            ws.cell(row=ri, column=1, value=d)
            ws.cell(row=ri, column=2, value=direction_th)
            ws.cell(row=ri, column=3, value=txn["category"])
            ws.cell(row=ri, column=4, value=txn["user_category"])
            ws.cell(row=ri, column=5, value=txn["amount"])
            ws.cell(row=ri, column=6, value=txn["description"])
            ws.cell(row=ri, column=7, value=txn["note"])

        # I/J summary totals block (appended after meta rows, matching source format)
        meta_end = len(sidecar) + 1
        inc_sum = sum(t["amount"] for t in txns if t["direction"] == "income")
        exp_sum = sum(t["amount"] for t in txns if t["direction"] == "expense")
        totals = [("รายรับ", inc_sum), ("รายจ่าย", exp_sum),
                  ("คงเหลือ", inc_sum - exp_sum)]
        for ti, (label, val) in enumerate(totals, start=meta_end):
            ws.cell(row=ti, column=9, value=label)
            ws.cell(row=ti, column=10, value=val)

    # ── 3. Salary_Sheet ───────────────────────────────────────────────────────
    ws_sal = wb.create_sheet("Salary_Sheet")
    # Row 1: 'des' marker  (parser checks min_row=3 so rows 1-2 are header)
    ws_sal.cell(row=1, column=1, value="des")
    # Row 2: header
    sal_headers = ["", "ชื่อ", "นามสกุล", "ชื่อเล่น", "ธนาคาร", "เลขบัญชี",
                   "เงินเดือน", "หักประกันสังคม", "เงินเดือนสุทธิ", "is_active"]
    for ci, h in enumerate(sal_headers, start=1):
        ws_sal.cell(row=2, column=ci, value=h)

    # Employee rows from DB
    employees = conn.execute("""
        SELECT e.id, e.full_name, e.nickname, e.bank_name, e.bank_account_no, e.is_active,
               COALESCE(sh.monthly_salary, 0) AS salary,
               COALESCE(e.sso_enrolled, 0) AS sso_enrolled
        FROM employees e
        LEFT JOIN (
            SELECT employee_id,
                   monthly_salary,
                   ROW_NUMBER() OVER (PARTITION BY employee_id ORDER BY effective_date DESC) AS rn
            FROM employee_salary_history
        ) sh ON sh.employee_id = e.id AND sh.rn = 1
        WHERE e.is_active = 1
        ORDER BY e.emp_code
    """).fetchall()

    for ri, emp in enumerate(employees, start=3):
        salary = float(emp["salary"] or 0.0)
        # SSO ~750 (standard Thai deduction) if enrolled; else 0
        sso = 750.0 if emp["sso_enrolled"] else 0.0
        net = salary - sso

        # Split full_name into first/last
        parts = str(emp["full_name"] or "").split(" ", 1)
        first = parts[0] if parts else ""
        last  = parts[1] if len(parts) > 1 else ""

        ws_sal.cell(row=ri, column=2, value=first)
        ws_sal.cell(row=ri, column=3, value=last)
        ws_sal.cell(row=ri, column=4, value=emp["nickname"])
        ws_sal.cell(row=ri, column=5, value=emp["bank_name"])
        ws_sal.cell(row=ri, column=6, value=emp["bank_account_no"])
        ws_sal.cell(row=ri, column=7, value=salary)
        ws_sal.cell(row=ri, column=8, value=sso)
        ws_sal.cell(row=ri, column=9, value=net)
        ws_sal.cell(row=ri, column=10, value=emp["is_active"])

    # ── 4. เบิกเงินล่วงหน้า ───────────────────────────────────────────────────
    ws_adv = wb.create_sheet("เบิกเงินล่วงหน้า")
    # Row 1: filler; Row 2: header (cols B-E = 2-5)
    ws_adv.cell(row=2, column=2, value="วันที่")
    ws_adv.cell(row=2, column=3, value="ชื่อ")
    ws_adv.cell(row=2, column=4, value="เบิกเงินล่วงหน้า")
    ws_adv.cell(row=2, column=5, value="หมายเหตุ")

    advances = conn.execute(
        """SELECT sa.advance_date, COALESCE(e.nickname, sa.raw_name) AS name,
                  sa.amount, sa.note
           FROM salary_advances sa
           LEFT JOIN employees e ON e.id = sa.employee_id
           ORDER BY sa.advance_date ASC, sa.id ASC"""
    ).fetchall()

    import datetime as _dt2
    for ri, adv in enumerate(advances, start=3):
        try:
            d = _dt2.date.fromisoformat(str(adv["advance_date"])[:10])
        except (TypeError, ValueError):
            d = adv["advance_date"]
        ws_adv.cell(row=ri, column=2, value=d)
        ws_adv.cell(row=ri, column=3, value=adv["name"])
        ws_adv.cell(row=ri, column=4, value=adv["amount"])
        ws_adv.cell(row=ri, column=5, value=adv["note"])

    # ── 5. Setup ──────────────────────────────────────────────────────────────
    ws_setup = wb.create_sheet("Setup")
    # Row 2: header (B=รายรับ, C=รายจ่าย, E=ผู้ใช้, F=ผู้ใช้ (คน))
    ws_setup.cell(row=2, column=2, value="รายรับ")
    ws_setup.cell(row=2, column=3, value="รายจ่าย")
    ws_setup.cell(row=2, column=5, value="ผู้ใช้")
    ws_setup.cell(row=2, column=6, value="ผู้ใช้ (คน)")

    income_cats = conn.execute(
        "SELECT name FROM cashbook_categories WHERE direction='income' AND is_active=1 ORDER BY sort_order, id"
    ).fetchall()
    expense_cats = conn.execute(
        "SELECT name FROM cashbook_categories WHERE direction='expense' AND is_active=1 ORDER BY sort_order, id"
    ).fetchall()

    max_rows = max(len(income_cats), len(expense_cats), 1)
    for i in range(max_rows):
        row = i + 3
        if i < len(income_cats):
            ws_setup.cell(row=row, column=2, value=income_cats[i]["name"])
        if i < len(expense_cats):
            ws_setup.cell(row=row, column=3, value=expense_cats[i]["name"])

    # ── Serialise ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
