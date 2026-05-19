"""
parse_cashbook.py — Parser for the Sendy Cashbook Excel module.

Handles multi-sheet cashbook workbooks exported from the Cashbook module.
Pure parser — returns normalised Python data structures, no DB writes.

Workbook structure
──────────────────
  Overview              — small block with รายรับ / รายจ่าย / คงเหลือ
  Txn_<code>            — account ledger sheets (identified by name prefix)
  Salary_Sheet          — employee salary records
  เบิกเงินล่วงหน้า     — salary advance records
  Setup                 — income/expense category lists
  Note >>               — separator (ignored)
  Deprecated >>         — separator (ignored)

All other sheets (e.g. "Overview") are skipped by the Txn loop.

Transaction validity (per-row):
  - col A must be a datetime object (real date, not a string or None)
  - col B must be 'รายรับ' or 'รายจ่าย'
  - col E (amount) must be non-empty and parseable to float

I/J sidecar (cols I and J = 0-indexed 8 and 9):
  Label in col I, value in col J.
  Recognised account-meta labels: Bank, Account Number, Name, หมายเหตุ
  Computed totals (รายรับ / รายจ่าย / คงเหลือ) are present too — ignored.

BSN note on encoding: this parser reads .xlsx files (openpyxl), not CSV.
No cp874 encoding concern here. UTF-8 DB storage applies as usual.
"""

import datetime
from typing import Any, Dict, List, Optional

import openpyxl


# ── Public helpers ────────────────────────────────────────────────────────────

def is_novat_file(path: str) -> bool:
    """
    Return True if the filename (basename) contains 'novat' (case-insensitive).
    Call this in the importer to decide whether to tag transactions as VAT-exempt.
    Parser itself is file-agnostic; vat_flag is NOT embedded in row data.
    """
    import os
    basename = os.path.basename(path).lower()
    return "novat" in basename


# ── Internal helpers ──────────────────────────────────────────────────────────

def _clean_account_no(raw: Any) -> Optional[str]:
    """
    Normalise a bank account number cell value.

    openpyxl often reads long numeric account numbers as floats (e.g. 2322842392.0).
    Strip the trailing '.0' and return as a string.  Non-numeric / None → None.
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # Strip trailing '.0' from float representation
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    # Also handle cases like '1234567890.0' → '1234567890'
    try:
        f = float(s.replace(",", ""))
        if f == int(f):
            return str(int(f))
        return s
    except (ValueError, TypeError):
        return s


def _parse_amount(raw: Any) -> Optional[float]:
    """
    Parse an amount cell that may be int, float, or a string with thousands commas.
    Returns the positive float value, or None if not parseable.
    Amount is always positive; direction is conveyed separately via 'direction' field.
    """
    if raw is None:
        return None
    s = str(raw).strip().replace(",", "").replace(" ", "")
    if not s:
        return None
    try:
        return abs(float(s))
    except (ValueError, TypeError):
        return None


def _date_to_iso(val: Any) -> Optional[str]:
    """Convert a datetime or date object to ISO YYYY-MM-DD string."""
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-%d")
    return None


def _str_or_none(val: Any) -> Optional[str]:
    """Return stripped string or None for blank/None values."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _bool_from_cell(val: Any) -> bool:
    """
    Parse is_active field.
    True/False booleans pass through.  'True'/'False' strings handled.
    Any other truthy value → True; falsy → False.
    """
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().lower() == "true"
    return bool(val)


# ── Sheet parsers ─────────────────────────────────────────────────────────────

def _parse_overview(ws) -> Dict[str, float]:
    """
    Extract the Overview block.

    Expected layout (from real file inspection):
      row 3: None, 'รายรับ',  <income_value>,  ...
      row 4: None, 'รายจ่าย', <expense_value>, ...
      row 5: None, 'คงเหลือ', <balance_value>, ...

    Scans ALL rows for these labels in any column pair (label, value) so that
    minor row-offset changes don't break parsing.
    """
    label_map = {
        "รายรับ":  "income",
        "รายจ่าย": "expense",
        "คงเหลือ": "balance",
    }
    result: Dict[str, float] = {"income": 0.0, "expense": 0.0, "balance": 0.0}

    for row in ws.iter_rows(values_only=True):
        for ci, cell in enumerate(row):
            if cell in label_map:
                # Value expected in the next column
                if ci + 1 < len(row) and row[ci + 1] is not None:
                    try:
                        result[label_map[cell]] = float(row[ci + 1])
                    except (ValueError, TypeError):
                        pass
    return result


def _parse_txn_sheet(ws, account_code: str, sheet_name: str, warnings: List[str]):
    """
    Parse one Txn_* ledger sheet.

    Returns:
      account_meta  — dict with bank/account_no/owner/note
      transactions  — list of normalised transaction dicts
    """
    # Account meta from I/J sidecar
    META_LABELS = {
        "Bank":           "bank_name",
        "Account Number": "bank_account_no",
        "Name":           "account_owner_name",
        "หมายเหตุ":       "note",
    }
    # Labels that signal the IJ computed-totals block — ignored for meta
    TOTALS_LABELS = {"รายรับ", "รายจ่าย", "คงเหลือ"}

    meta: Dict[str, Any] = {
        "bank_name":          None,
        "bank_account_no":    None,
        "account_owner_name": None,
        "note":               None,
    }
    transactions: List[Dict[str, Any]] = []

    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    for row_idx, row in enumerate(all_rows, start=1):
        # Pad row to at least 10 elements so index access is safe
        row = tuple(row) + (None,) * max(0, 10 - len(row))

        # ── I/J sidecar (cols I=index 8, J=index 9) ──────────────────────────
        label_ij = row[8]
        value_ij = row[9]
        if label_ij is not None and label_ij not in TOTALS_LABELS:
            key = META_LABELS.get(label_ij)
            if key is not None:
                if key == "bank_account_no":
                    meta[key] = _clean_account_no(value_ij)
                else:
                    meta[key] = _str_or_none(value_ij)

        # ── Transaction columns A-G (indices 0-6) ────────────────────────────
        if row_idx == 1:
            # Header row — skip
            continue

        col_a = row[0]   # วันที่
        col_b = row[1]   # ประเภท
        col_c = row[2]   # หมวดหมู่
        col_d = row[3]   # หมวดหมู่_ผู้ใช้
        col_e = row[4]   # จำนวนเงิน
        col_f = row[5]   # รายละเอียด
        col_g = row[6]   # หมายเหตุ

        # Skip blank rows
        if col_a is None and col_b is None and col_e is None:
            continue

        # Validate: col_a must be a real date
        if not isinstance(col_a, (datetime.datetime, datetime.date)):
            # Has some data but no valid date — skip with warning if it looked like a row
            if col_a is not None or col_b is not None:
                warnings.append(
                    f"{sheet_name} row {row_idx}: skipped — col A is not a date "
                    f"(got {col_a!r})"
                )
            continue

        # Validate: col_b must be รายรับ or รายจ่าย
        if col_b not in ("รายรับ", "รายจ่าย"):
            warnings.append(
                f"{sheet_name} row {row_idx}: skipped — ประเภท is {col_b!r} "
                f"(expected รายรับ or รายจ่าย)"
            )
            continue

        # Validate: col_e must be parseable
        amount = _parse_amount(col_e)
        if amount is None:
            warnings.append(
                f"{sheet_name} row {row_idx}: skipped — amount not parseable: {col_e!r}"
            )
            continue

        direction = "income" if col_b == "รายรับ" else "expense"

        # Description (col F) — preserve embedded newlines
        description = _str_or_none(col_f)
        if col_f is not None and "\n" in str(col_f):
            description = str(col_f)  # keep exactly as-is, including newlines

        transactions.append({
            "account_code":  account_code,
            "source_sheet":  sheet_name,
            "source_row":    row_idx,
            "txn_date":      _date_to_iso(col_a),
            "direction":     direction,
            "category":      _str_or_none(col_c),
            "user_category": _str_or_none(col_d),
            "amount":        amount,
            "description":   description,
            "note":          _str_or_none(col_g),
        })

    # Build account dict
    account = {
        "code":               account_code,
        "bank_name":          meta["bank_name"],
        "bank_account_no":    meta["bank_account_no"],
        "account_owner_name": meta["account_owner_name"],
        "note":               meta["note"],
    }

    return account, transactions


def _parse_salary_sheet(ws, warnings: List[str]) -> List[Dict[str, Any]]:
    """
    Parse Salary_Sheet.

    Layout:
      row 1: ['des', ...]
      row 2: ['', ชื่อ, นามสกุล, ชื่อเล่น, ธนาคาร, เลขบัญชี, เงินเดือน,
               หักประกันสังคม, เงินเดือนสุทธิ, is_active]  (cols A-J, 1-indexed)
      row 3+: data (cols B-J = indices 1-9)
    """
    employees: List[Dict[str, Any]] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        # cols are 0-indexed; B=1, C=2, D=3, E=4, F=5, G=6, H=7, I=8, J=9
        row = tuple(row) + (None,) * max(0, 10 - len(row))

        first_name = _str_or_none(row[1])
        if not first_name:
            continue   # skip blank rows

        last_name     = _str_or_none(row[2])
        nickname      = _str_or_none(row[3])
        bank          = _str_or_none(row[4])
        bank_acct_raw = row[5]
        salary_raw    = row[6]
        sso_raw       = row[7]
        net_raw       = row[8]
        is_active_raw = row[9]

        try:
            salary = float(salary_raw) if salary_raw is not None else 0.0
        except (ValueError, TypeError):
            salary = 0.0
            warnings.append(f"Salary_Sheet row {row_idx}: salary not parseable: {salary_raw!r}")

        try:
            sso = float(sso_raw) if sso_raw is not None else 0.0
        except (ValueError, TypeError):
            sso = 0.0

        try:
            net = float(net_raw) if net_raw is not None else 0.0
        except (ValueError, TypeError):
            net = 0.0

        employees.append({
            "first_name":      first_name,
            "last_name":       last_name,
            "nickname":        nickname,
            "bank":            bank,
            "bank_account_no": _clean_account_no(bank_acct_raw),
            "salary":          salary,
            "sso_deduction":   sso,
            "net_salary":      net,
            "is_active":       _bool_from_cell(is_active_raw),
        })

    return employees


def _parse_advances_sheet(ws, warnings: List[str]) -> List[Dict[str, Any]]:
    """
    Parse เบิกเงินล่วงหน้า (salary advances).

    Layout:
      row 1: (empty filler)
      row 2: ['', วันที่, ชื่อ, เบิกเงินล่วงหน้า, หมายเหตุ]  (cols A-E)
      row 3+: data in cols B-E (0-indexed 1-4)
    """
    advances: List[Dict[str, Any]] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        row = tuple(row) + (None,) * max(0, 5 - len(row))

        date_val   = row[1]  # col B
        name_val   = row[2]  # col C
        amount_raw = row[3]  # col D
        note_val   = row[4]  # col E

        # Skip row if no date
        if date_val is None:
            continue
        if not isinstance(date_val, (datetime.datetime, datetime.date)):
            continue

        amount = _parse_amount(amount_raw)
        if amount is None:
            warnings.append(
                f"เบิกเงินล่วงหน้า row {row_idx}: skipped — amount not parseable: {amount_raw!r}"
            )
            continue

        advances.append({
            "advance_date": _date_to_iso(date_val),
            "raw_name":     _str_or_none(name_val),
            "amount":       amount,
            "note":         _str_or_none(note_val),
        })

    return advances


def _parse_setup_sheet(ws) -> Dict[str, List[str]]:
    """
    Parse Setup sheet category lists.

    Layout:
      row 2: header (col B='รายรับ', col C='รายจ่าย', col E='ผู้ใช้', col F='ผู้ใช้ (คน)')
      row 3+: col B = income categories, col C = expense categories
              (stop collecting at first blank in each column independently)

    Returns dict with keys 'income' and 'expense' (order-preserving, deduped).
    setup_accounts (col E) and setup_users (col F) are also extracted.
    """
    income_cats: List[str] = []
    expense_cats: List[str] = []
    setup_accounts: List[str] = []
    setup_users: List[str] = []

    seen_income: set = set()
    seen_expense: set = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        row = tuple(row) + (None,) * max(0, 6 - len(row))
        # col B = index 1, col C = index 2, col E = index 4, col F = index 5

        inc_val  = _str_or_none(row[1])
        exp_val  = _str_or_none(row[2])
        acct_val = _str_or_none(row[4])
        user_val = _str_or_none(row[5])

        if inc_val and inc_val not in seen_income:
            income_cats.append(inc_val)
            seen_income.add(inc_val)

        if exp_val and exp_val not in seen_expense:
            expense_cats.append(exp_val)
            seen_expense.add(exp_val)

        if acct_val:
            setup_accounts.append(acct_val)
        if user_val:
            setup_users.append(user_val)

    return {
        "income":          income_cats,
        "expense":         expense_cats,
        "setup_accounts":  setup_accounts,
        "setup_users":     setup_users,
    }


# ── Main entry point ──────────────────────────────────────────────────────────

def parse_cashbook(path: str) -> Dict[str, Any]:
    """
    Parse a cashbook Excel workbook (NoVat or Vat variant).

    Parameters
    ----------
    path : str
        Absolute path to the .xlsx file.

    Returns
    -------
    dict with keys:
      accounts     — list of account dicts (one per Txn_* sheet)
      transactions — list of normalised transaction dicts
      salary       — list of employee salary dicts
      advances     — list of salary-advance dicts
      categories   — {'income': [...], 'expense': [...]}
      overview     — {'income': float, 'expense': float, 'balance': float}
      warnings     — list of human-readable skip/anomaly messages

    Notes
    -----
    - does NOT set vat_flag on rows; call is_novat_file(path) in the importer.
    - does NOT write to any database.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    warnings_list: List[str] = []
    accounts: List[Dict[str, Any]] = []
    transactions: List[Dict[str, Any]] = []

    salary: List[Dict[str, Any]] = []
    advances: List[Dict[str, Any]] = []
    categories: Dict[str, Any] = {"income": [], "expense": []}
    overview: Dict[str, float] = {"income": 0.0, "expense": 0.0, "balance": 0.0}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if sheet_name.startswith("Txn_"):
            # Account code = sheet name with 'Txn_' prefix stripped
            account_code = sheet_name[4:]
            acct, txns = _parse_txn_sheet(ws, account_code, sheet_name, warnings_list)
            accounts.append(acct)
            transactions.extend(txns)

        elif sheet_name == "Overview":
            overview = _parse_overview(ws)

        elif sheet_name == "Salary_Sheet":
            salary = _parse_salary_sheet(ws, warnings_list)

        elif sheet_name == "เบิกเงินล่วงหน้า":
            advances = _parse_advances_sheet(ws, warnings_list)

        elif sheet_name == "Setup":
            setup = _parse_setup_sheet(ws)
            categories = {
                "income":  setup["income"],
                "expense": setup["expense"],
            }

        # All other sheets (Note >>, Deprecated >>, etc.) are silently ignored

    return {
        "accounts":     accounts,
        "transactions": transactions,
        "salary":       salary,
        "advances":     advances,
        "categories":   categories,
        "overview":     overview,
        "warnings":     warnings_list,
    }
